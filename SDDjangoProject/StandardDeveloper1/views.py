from django.shortcuts import render
from django.http import HttpResponse
from py2neo import Graph
import openpyxl
from datetime import datetime
import pandas as pd 
import numpy as np
import sys
from django.http import HttpResponse, JsonResponse
import json
import re
import os

graph = Graph('http://neo4j:letsgowings@localhost:7474/db/data/')


#graph = Graph('http://neo4j:letsgowings@10.0.0.10:7474/db/data/')

# Create your views here.
def index(request):
	return render(request,'StandardDeveloper1/index.html')


def NewStudy(request):
	StudyName=request.GET["StudyName"]
	StudyDescription=request.GET['StudyDescription']
	ProtocolName=request.GET['ProtocolName']
	parentstandard=request.GET['ModelName']
	parentversion=request.GET['ModelVersion']
	Documents=json.loads(request.GET['Documents'])
	Dictionaries=json.loads(request.GET['ExtDicts'])

	# Extract the model name and version
	ModelRL=graph.cypher.execute('match (:Standard {Name:"'+parentstandard+'",Version:"'+parentversion+'"})-[:BasedOn]->(m:Model) return m.name as Name,m.version as Version')
	ModelS=pd.Series(ModelRL[0],index=ModelRL.columns)
	# Create a node for the new study and attach it to the standard to which the parent is attached

	if parentstandard == 'ADAM':
		stmt= 'match (a:Standard {Name:"'+parentstandard+'",Version:"'+parentversion+'"}) create (a)<-[:BasedOn]-(c:Study {Name:"'+StudyName+'",Description:"'+StudyDescription+'",\
			ProtocolName:"'+ProtocolName+'"}) '

		if Documents:
			stmt=stmt+'with c create (c)-[:ContainsDocs]->(docs:Documents) '
			for x in Documents:
				if x['name'] == 'Annotated Case Report Form':
					doctype='acrf'
				else:
					doctype='suppdoc'

				stmt=stmt+'with c,docs create (docs)-[:ContainDoc]->(:Document {Type:"'+doctype+'",Name:"'+x['name']+'",File:"'+x['file']+'"}) '

		if Dictionaries:
			stmt=stmt+'with c create (c)-[:ContainsDictionaries]->(dicts:Dictionaries) '
			for x in Dictionaries:
				stmt=stmt+'with c,dicts create (dicts)-[:ContainDict]->(:Dictionary {Name:"'+x['name']+'",Description:"'+x['description']+'",Version:"'+x['version']+'"}) '

		print 'NEWSTUDY STMT: '+stmt ;

		tx = graph.cypher.begin()
		tx.append(stmt)
		tx.commit()

		return render(request,'StandardDeveloper1/FirstStudyHome.html',{'Study':StudyName,'StandardName':parentstandard,'StandardVersion':parentversion,'ModelName':ModelS['Name'],'ModelVersion':ModelS['Version']})

def Back2Study(request):
	StudyName=request.POST["Study"]
	StandardName=request.POST['StandardName']
	StandardVersion=request.POST['StandardVersion']
	ModelName=request.POST['ModelName']
	ModelVersion=request.POST['ModelVersion']
	return render(request,'StandardDeveloper1/SubsequentStudyHome.html',{'Study':StudyName,'StandardName':StandardName,'StandardVersion':StandardVersion,'ModelName':ModelName,'ModelVersion':ModelVersion})

def GetStudyInfo(request):
	Study=request.GET['Study']
	RL=graph.cypher.execute('match (a:Study {Name:"'+Study+'"}) return a.Description as Description,a.ProtocolName as ProtocolName')
	DF=pd.DataFrame(RL.records,columns=RL.columns)
	return HttpResponse(DF.to_json(orient='records'),content_type='application/json')

def GetDicts(request):
	Study=request.GET['Study']
	RL=graph.cypher.execute('match (a:Study {Name:"'+Study+'"})--(:Dictionaries)--(dict:Dictionary) return dict.Name as Name,dict.Description as Description,dict.Version as Version')
	DF=pd.DataFrame(RL.records,columns=RL.columns)
	return HttpResponse(DF.to_json(orient='records'),content_type='application/json')

def GetDocs(request):
	Study=request.GET['Study']
	RL=graph.cypher.execute('match (a:Study {Name:"'+Study+'"})--(:Documents)--(doc:Document) return doc.Name as Name,doc.File as File')
	DF=pd.DataFrame(RL.records,columns=RL.columns)
	return HttpResponse(DF.to_json(orient='records'),content_type='application/json')



def NewMerge(request):
	Study=request.POST['Study']
	StandardName=request.POST['StandardName']
	StandardVersion=request.POST['StandardVersion']
	IGDName=request.POST['DSName']
	Class=request.POST['Class']

	Instruction="We're now ready to define "+IGDName+" predecessors from other data sets.  "
	Instruction=Instruction+"Fill in the following information about one of the merge data sets.  "
	Instruction=Instruction+"Then click the button to begin defining predecessor variables from that data set."

	return render(request,'StandardDeveloper1/predsource.html', \
		{'RSType':'MERGE','Action':'Add','NewStudy':'N','Study':Study,'Instruction':Instruction,'IGDName':IGDName,'StandardName':StandardName,'StandardVersion':StandardVersion,'Class':Class})

def NewModel(request):
	Study=request.POST['Study']
	StandardName=request.POST['StandardName']
	StandardVersion=request.POST['StandardVersion']
	DSName=request.POST['DSName']
	Class=request.POST['Class']
	Model=request.POST['add_model']


	Instruction="We're now ready to define other "+DSName+" model variables.  "
	Instruction=Instruction+"As with predecessors, we'll start by choosing a model variable.  "
	Instruction=Instruction+"We'll then define the metadata along with methods for derived variables."

	if Model == "Add Model variable":
		ReturnTo='datasethome'
	else:
		ReturnTo='modelvarlist'

	Lists=PrepModelLists(StandardName,StandardVersion,Study,DSName,Class)
	Lists['Study']=Study
	Lists['StandardName']=StandardName
	Lists['StandardVersion']=StandardVersion
	Lists['DSName']=DSName
	Lists['Class']=Class
	Lists['Instruction']=Instruction
	Lists['Action']='Add'
	Lists['ReturnTo']=ReturnTo

	return render(request,'StandardDeveloper1/modelvarlist.html', Lists)

def NewDS(request):
	Study=request.POST["Study"]
	StandardName=request.POST["StandardName"]
	StandardVersion=request.POST["StandardVersion"]
	ModelName=request.POST['ModelName']
	ModelVersion=request.POST['ModelVersion']
	IGDSource=request.POST['IGDSource']
	MDDic=json.loads(request.POST['MD'])
	RecordSourceList=json.loads(request.POST['RecordSources'])

	DSName=MDDic['Name']
	Class=MDDic['Class']

	if IGDSource == 'Standard':
		stmt='match (study:Study {Name:"'+Study+'"})-[:BasedOn]->(:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})-[:ItemGroupRef]->(end:ItemGroupDef \
			{Name:"'+DSName+'"})'
	else:
		stmt='match (study:Study {Name:"'+Study+'"})-[:BasedOn]->(:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})-[:BasedOn]->(:Model)-[:ItemGroupRef]->(end:ItemGroupDef {Name:"'+Class+'"})'

	stmt=stmt+' create (study)-[:ItemGroupRef]->(igd:ItemGroupDef {Name:"'+DSName+'",Label:"'+MDDic['Label']+'",Repeating:"'+MDDic['Repeating']+'",Structure:"'+MDDic['Structure']+'",IsReferenceData:"'+MDDic['Reference']+'"}) \
		-[:BasedOn]->(end) '

	for x in RecordSourceList:
		model=x['model']
		if model == 'ADAM':
			# Connect to already existing nodes
			 stmt=stmt+'with study,igd match (study)--(igdR:ItemGroupDef {Name:"'+x['dataset']+'"}) merge (igd)-[:RecordSource {Subset:"'+x['subset']+'"}]->(igdR) '
		else:
			# For now, we're just creating a new SDTM ItemGroupDef node to connect to
			stmt=stmt+'with study,igd merge (igd)-[:RecordSource {Subset:"'+x['subset']+'"}]->(:ItemGroupDef {Name:"'+x['dataset']+'"}) '

	if Class == 'BASIC DATA STRUCTURE':
		# Create parameter variables

		# Calculate new Method OID and WhereClause OID
		maxOIDRL=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)--(md1:MethodDef) with max(md1.OID) as max1 \
			optional match (:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)--(:ValueListDef)--(:ItemDef)--(md2:MethodDef) return max1,max(md2.OID) \
			as max2')
		if maxOIDRL:
			moids=pd.Series(maxOIDRL[0],index=maxOIDRL.columns)
			MethodOID=max(moids['max1'],moids['max2'])
		else:
			MethodOID=0
			

		WCOIDmax=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)--(:ValueListDef)--(:ItemDef)--(w:WhereClauseDef) return max(w.OID) as OID')
		if WCOIDmax:
			woids=pd.Series(WCOIDmax[0],index=WCOIDmax.columns)
			if woids['OID']:
				WCOID=woids['OID']
			else:
				WCOID=0
		else:
			WCOID=0
			

		withlist = ['study','igd']

		stmt=stmt+'with '+', '.join(withlist)+' match (:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})-[:BasedOn]->(:Model)--(igdm:ItemGroupDef {Name:"BASIC DATA STRUCTURE"}) '

		withlist.append('igdm')

		ParmDefList=json.loads(request.POST['Parmdefs'])

		# Create PARAMCD
		
		# Find PARAMCD in the model
		stmt=stmt+'with '+', '.join(withlist)+' match (igdm)--(idm:ItemDef {Name:"PARAMCD"}) '
		withlist.append('idm')
		# See if See Parameter Page method exists in the study
		PPOIDRL=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)--(m:MethodDef {Description:"See Parameter Page"}) return m.OID as PPOID')
		if PPOIDRL:
			PPOIDS=pd.Series(PPOIDRL[0],index=PPOIDRL.columns)
			IRMOID=PPOIDS['PPOID']
		else:
			MethodOID=MethodOID+1
			IRMOID=MethodOID

		# Calculate SASLength
		Length=0
		for x,y in enumerate(ParmDefList):
			Length=max(Length,len(y['paramcd']))


		# Create the PARAMCD ItemDef node
		stmt=stmt+'with '+', '.join(withlist)+' create (igd)-[:ItemRef {Mandatory:"Yes",MethodOID:'+str(IRMOID)+',OrderNumber:1}]->(id:ItemDef)-\
			[:CodeListRef]->(cl:CodeList {Extensible:"Yes",DataType:"text",Name:"'+DSName+' Parameter Code "}) set id.Name=idm.Name,id.Label=idm.Label,id.SASType=idm.SASType,\
			id.SASLength='+str(Length)+',id.DataType=idm.DataType,id.Origin=idm.Origin '

		withlist.append('id')
		withlist.append('cl')

		# BasedOn
		stmt=stmt+'with '+', '.join(withlist)+' create (id)-[:BasedOn]->(idm) '
		withlist.remove('idm')

		# Method
		if PPOIDRL:
			stmt=stmt+'with '+', '.join(withlist)+' match (:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)--(m:MethodDef {OID:'+str(IRMOID)+'}) with '+', '.join(withlist)+',m limit 1 \
				create (id)-[:MethodRef]->(m) '

		else:
			stmt=stmt+'with '+', '.join(withlist)+' create (id)-[:MethodRef]->(m:MethodDef {OID:'+str(IRMOID)+',Description:"See Parameter Page"}) '

		withlist.append('m')
		withlist.remove('id')

		# Add codelist items
		for x,y in enumerate(ParmDefList):
			stmt=stmt+'with '+', '.join(withlist)+' create (cl)-[:ContainsCodeListItem]->(:CodeListItem {CodedValue:"'+y['paramcd']+'",Decode:"'+y['param']+'"}) '
		withlist.remove('cl')






		# Create PARAM
		
		# Find PARAM in the model
		stmt=stmt+'with '+', '.join(withlist)+' match (igdm)--(idm:ItemDef {Name:"PARAM"}) '
		withlist.append('idm')

		# Calculate SASLength
		Length=0
		for x,y in enumerate(ParmDefList):
			Length=max(Length,len(y['param']))

		# Create the PARAM ItemDef node
		stmt=stmt+'with '+', '.join(withlist)+' create (igd)-[:ItemRef {Mandatory:"Yes",MethodOID:'+str(IRMOID)+',OrderNumber:2}]->(id:ItemDef)-\
			[:CodeListRef]->(cl:CodeList {Extensible:"Yes",DataType:"text",Name:"'+DSName+' Parameter Code "}) set id.Name=idm.Name,id.Label=idm.Label,id.SASType=idm.SASType,\
			id.SASLength='+str(Length)+',id.DataType=idm.DataType,id.Origin=idm.Origin '

		withlist.append('id')
		withlist.append('cl')

		# BasedOn
		stmt=stmt+'with '+', '.join(withlist)+' create (id)-[:BasedOn]->(idm) '
		withlist.remove('idm')

		# Method
		stmt=stmt+'with '+', '.join(withlist)+' create (id)-[:MethodRef]->(m) '
		withlist.remove('id')

		# Add codelist items
		for x,y in enumerate(ParmDefList):
			stmt=stmt+'with '+', '.join(withlist)+' create (cl)-[:ContainsCodeListItem]->(:CodeListItem {CodedValue:"'+y['param']+'"}) '
		withlist.remove('cl')






		# Create PARCAT and PARCATN variables if applicable
		if (request.POST['Parcats']):
			ParcatList=json.loads(request.POST['Parcats'])
			# Find PARCATy in the model
			stmt=stmt+'with '+', '.join(withlist)+' match (igdm)--(idm:ItemDef {Name:"PARCATy"}) '
			withlist.append('idm')

			for x,y in enumerate(ParcatList):
				# Create the PARCAT ItemDef nodes
				stmt=stmt+'with '+', '.join(withlist)+' create (igd)-[:ItemRef {Mandatory:"Yes",MethodOID:'+str(IRMOID)+',OrderNumber:'+str(x+5)+'}]->(id:ItemDef)-\
					[:CodeListRef]->(cl:CodeList {Extensible:"Yes",DataType:"text",Name:"'+DSName+' Parameter Category '+str(x+1)+'"}) '

				withlist.append('id')
				withlist.append('cl')

				# Method
				stmt=stmt+'with '+', '.join(withlist)+' create (id)-[:MethodRef]->(m) '

				stmt=stmt+'set id.Name="'+y['Name']+'",id.Label="Parameter Category '+str(x+1)+'",id.Comment="'+y['Comment']+'",id.SASLength="'+str(y['SASLength'])+'",id.SASType=idm.SASType,\
					id.DataType=idm.DataType,id.Origin=idm.Origin '

				# Based on
				stmt=stmt+'with '+', '.join(withlist)+' create (id)-[:BasedOn]->(idm) '
				withlist.remove('id')

				# Add codelist items
				for y1 in y['CT']:
					stmt=stmt+'with '+', '.join(withlist)+' create (cl)-[:ContainsCodeListItem]->(cli:CodeListItem {CodedValue:"'+y1['term']+'"}) '
					withlist.append('cli')
					# Connect PARAMCD codelist items to corresponding PARCAT codelist items
					for x2 in ParmDefList:
						if x2['parcats'][x] == y1['term']:
							stmt=stmt+'with '+', '.join(withlist)+' match (study)--(igd)--(:ItemDef {Name:"PARAMCD"})--(:CodeList)--(pcli:CodeListItem {CodedValue:"'+x2['paramcd']+'"}) \
								create (pcli)-[:ParameterAttribute {Type:"PARCAT'+str(x+1)+'"}]->(cli) '
					withlist.remove('cli')


				withlist.remove('cl')

				# Create PARCATN
				if y['ParcatNTF']:
					stmt=stmt+'with '+', '.join(withlist)+' match (igdm)--(idmn:ItemDef {Name:"PARCATyN"}) '
					withlist.append('idmn')
					stmt=stmt+'with '+', '.join(withlist)+' create (igd)-[:ItemRef {Mandatory:"Yes",MethodOID:'+str(IRMOID)+',OrderNumber:'+str(x+5+len(ParcatList))+'}]->(id:ItemDef)-\
						[:CodeListRef]->(cl:CodeList {Extensible:"Yes",DataType:"integer",Name:"'+DSName+' Parameter Category '+str(x+1)+' (N)"}) with '+', '.join(withlist)+', id, cl \
						create (id)-[:MethodRef]->(m) set id.Name="'+y['Name']+'N",id.Label="Parameter Category '+str(x+1)+' (N)",id.SASLength="8",id.SASType=idmn.SASType,\
						id.DataType=idmn.DataType,id.Origin=idmn.Origin  '
					withlist.append('id')
					withlist.append('cl')

					# Based on
					stmt=stmt+'with '+', '.join(withlist)+' create (id)-[:BasedOn]->(idmn) '
					withlist.remove('id')
					withlist.remove('idmn')

					# Codelist items
					for x1,y1 in enumerate(y['CT']):
						stmt=stmt+'with '+', '.join(withlist)+' create (cl)-[:ContainsCodeListItem]->(:CodeListItem {Decode:"'+y1['term']+'",\
							CodedValue:'+str(x1+1)+'}) '

					withlist.remove('cl')
					
			withlist.remove('idm')


		# Create DTYPEs

		# Make a dictionary of parameter codelists, where keys are parm codes, and values are sets that hold DTYPE values
		# UNIQUECODELISTS is a list of unique lists of dictionaries.  Each list of dictionaries represents a CT for a parameter
		parmdtypes={}
		uniquecodelists=[]
		for x in ParmDefList:
			if len(x['dtypes']) > 0:
				dtypelist=x['dtypes']
				dtypelistsort=sorted(dtypelist,key=lambda k: k['dtypevalue'])
				parmdtypes[x['paramcd']]=dtypelistsort 
				if dtypelistsort not in uniquecodelists:
					uniquecodelists.append(dtypelistsort)

		# If DTYPEs have been provided, then build the query
		if len(uniquecodelists) > 0:
			# Find DTYPE in the model
			stmt=stmt+'with '+', '.join(withlist)+' match (igdm)--(idm:ItemDef {Name:"DTYPE"}) '
			withlist.append('idm')

			# Create the DTYPE ItemDef node
			stmt=stmt+'with '+', '.join(withlist)+' create (igd)-[:ItemRef {Mandatory:"No",MethodOID:'+str(IRMOID)+',OrderNumber:'+str(5+2*len(ParcatList))+'}]->(id:ItemDef)-[:BasedOn]->(idm) \
				set id.Name=idm.Name,id.Label=idm.Label,id.SASType=idm.SASType,id.DataType=idm.DataType,id.Origin="Assigned"  create \
				(id)-[:ValueListRef]->(vl:ValueListDef) '

			withlist.remove('idm')
			withlist.append('vl')

			# Create a value-level ItemDef for each unique codelist in dtypecodelists.  Under each, create a codelist and a where clause
			for x,y in enumerate(uniquecodelists):
				# Determine which parameters used the current codelist
				WhichParms=[]
				WCOID=WCOID+1
				for x1,y1 in parmdtypes.iteritems():
					if y == y1:
						WhichParms.append(x1)
				# Determine Length for this ItemDef
				Length=0
				for x1 in y:
					Length=max(Length,len(x1))

				stmt=stmt+'with '+', '.join(withlist)+' create (vl)-[:ItemRef {Mandatory:"No",WCRef:'+str(WCOID)+'}]->(idv:ItemDef {Name:"DTYPE'+str(x+1)+'",Label:"Derivation Type '+str(x+1)+'",DataType:"text",\
					Length:'+str(Length)+',Origin:"Assigned"}) '
				withlist.append('idv')

				# Create codelist
				stmt=stmt+'with '+', '.join(withlist)+' create (idv)-[:CodeListRef]->(cl:CodeList {Extensible:"Yes",DataType:"text",Name:"Derivation Type '+str(x+1)+'"}) '
				withlist.append('cl')

				# Create codelist items
				for x1 in y:
					stmt=stmt+'with '+', '.join(withlist)+' create (cl)-[:ContainsCodeListItem]->(:CodeListItem {CodedValue:"'+x1['dtypevalue']+'",Decode:"'+x1['dtypedesc']+'"}) '
				withlist.remove('cl')

				# Create Where Clause
				if len(WhichParms)>1:
					operator="in"
					checkvalue=', '.join(WhichParms)
				else:
					operator="eq"
					checkvalue=WhichParms[0]

				stmt=stmt+'with '+', '.join(withlist)+' create (idv)-[:WhereClauseRef]->(wc:WhereClauseDef {OID:'+str(WCOID)+'}) with '+', '.join(withlist)+',wc match (study)--(igd)--(idpc:ItemDef {Name:"PARAMCD"}) \
					create (wc)-[:RangeCheck {Operator:"'+operator+'",CheckValue:"'+checkvalue+'"}]->(idpc) '
				withlist.remove('idv')






	print "NEWDS STMT: "+stmt
	tx=graph.cypher.begin()
	tx.append(stmt)
	tx.commit()

	# Determine the variable groups to be presented individually
	VGList=VarGroupList(DSName,Class)

	return render(request,'StandardDeveloper1/variablelist.html', {'Study':Study,'DSName':DSName,'StandardName':StandardName,'StandardVersion':StandardVersion,'Class':Class,\
		'VarGroup':VGList[0],'NextVarGroup':VGList[1],'IGDSource':IGDSource,'ModelName':ModelName,'ModelVersion':ModelVersion})

def RecordSource(request):
	DSName=request.POST['DSName']
	Study=request.POST['Study']
	StandardName=request.POST['StandardName']
	StandardVersion=request.POST['StandardVersion']
	Class=request.POST['Class']
	IGDSource=request.POST['IGDSource']
	Action=request.POST['Action']

	# Store record sources in the database
	Sources=json.loads(request.POST['Sources'])
	SourcesList=Sources['sources']
	stmt='match (study:Study {Name:"'+Study+'"})--(igd:ItemGroupDef {Name:"'+DSName+'"}) '
	for x in SourcesList:
		model=x['Model']
		if model == 'ADaM':
			# Connect to already existing nodes
			 stmt=stmt+'with study,igd match (study)--(igdR:ItemGroupDef {Name:"'+x['Dataset']+'"}) merge (igd)-[:RecordSource {Subset:"'+x['Subset']+'"}]->(igdR) '
		else:
			# For now, we're just creating a new SDTM ItemGroupDef node to connect to
			stmt=stmt+'with study,igd merge (igd)-[:RecordSource {Subset:"'+x['Subset']+'"}]->(:ItemGroupDef {Name:"'+x['Dataset']+'"}) '
	
	print "STMT: "+stmt
	tx=graph.cypher.begin()
	tx.append(stmt)
	tx.commit()

	# Determine the variable groups to be presented individually
	VGList=VarGroupList(DSName,Class)

	return render(request,'StandardDeveloper1/variablelist.html', {'Study':Study,'DSName':DSName,'StandardName':StandardName,'StandardVersion':StandardVersion,'Class':Class,'VarGroup':VGList[0],'NextVarGroup':VGList[1],'IGDSource':IGDSource,'Action':Action})

def QueryParameters(request):
	Study=request.POST["Study"]
	StandardName=request.POST["StandardName"]
	StandardVersion=request.POST["StandardVersion"]
	DSName=request.POST["DSName"]

	return render(request,'StandardDeveloper1/ParameterDefs.html',{'Action':'Edit','Study':Study,'IGDName':DSName,'StandardName':StandardName,'StandardVersion':StandardVersion})

def EditParameters(request):
	pass

def AddParms(request):
	Study=request.POST["Study"]
	StandardName=request.POST["StandardName"]
	StandardVersion=request.POST["StandardVersion"]
	DSName=request.POST["DSName"]
	VarSource=''
	ParmsDict=json.loads(request.POST["parms"])
	ParmsList=ParmsDict['parms']


	# Get model metadata for the necessary parameter variables
	# First determine which variables and create a string of their names separated by commas enclosed in brackets
	varsString="['PARAMCD','PARAM'"
	if 'paramn' in ParmsList[0]:
		varsString=varsString+",'PARAMN'"
	if 'parcats' in ParmsList[0]:
		varsString=varsString+",'PARCATy'"
	for x,y in enumerate(ParmsList):
		if 'dtypes' in ParmsList[x]:
			varsString=varsString+",'DTYPE'"
			break

	varsString=varsString+']'

	# Now get the standard/model metadata for these variables
	ParmMDRL = graph.cypher.execute('match (:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})-[:ItemGroupRef]->(:ItemGroupDef \
		{Name:"'+DSName+'"})-[:ItemRef]->(a:ItemDef) where a.Name in '+varsString+' return a.Name as Name,a.Label as Label,a.DataType as DataType, \
		a.Origin as Origin,a.SASType as SASType,a.MaxLength as SASLength order by a.Name desc')

	if ParmMDRL:
		VarSource='Standard'
		BasedOnString='match (:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})-[:ItemGroupRef]->(ms:ItemGroupDef {Name:"'+DSName+'"}) '

	else:
		ParmMDRL = graph.cypher.execute('match (:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})-[:BasedOn]->(m:Model)--(:ItemGroupDef \
			{Name:"BASIC DATA STRUCTURE"})-[:ItemRef]->(a:ItemDef) where a.Name in '+varsString+' return a.Name as Name,a.Label as Label,a.DataType as DataType, \
			a.Origin as Origin,a.SASType as SASType,a.MaxLength as SASLength,m.name as ModelName,m.version as ModelVersion order by a.Name desc')
		VarSource='Model'
		ModelName=ParmMDRL[0]['ModelName']
		ModelVersion=ParmMDRL[0]['ModelVersion']
		BasedOnString='match (:Model {name:"'+ModelName+'",version:"'+ModelVersion+'"})-[:ItemGroupRef]->(ms:ItemGroupDef {Name:"BASIC DATA STRUCTURE"}) '

	print 'PARMMDRL: ('+VarSource+')'
	print ParmMDRL

	# Start generating the query
	stmt='match (study:Study {Name:"'+Study+'"})--(igd:ItemGroupDef {Name:"'+DSName+'"}) with igd,study '+BasedOnString
	withlist=['igd','ms','study']

	# Calculate new Method OID and WhereClause OID
	maxOIDRL=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)--(md1:MethodDef) with max(md1.OID) as max1 \
		optional match (:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)--(:ValueListDef)--(:ItemDef)--(md2:MethodDef) return max1,max(md2.OID) \
		as max2')
	if maxOIDRL[0][0]:
		MethodOID=max(maxOIDRL[0]['max1'],maxOIDRL[0]['max2'])+1
	else:
		MethodOID=0

	WCOIDmax=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)--(:ValueListDef)--(:ItemDef)--(w:WhereClauseDef) return max(w.OID) as OID')[0][0]
	if not WCOIDmax:
		WCOIDmax=0


	for x in ParmMDRL:
		if x['Name'] == 'PARCATy':

			for xk in ParmsList[0]['parcats']:
				# For each PARCAT...
				# create the itemdef, codelist, and methoddef nodes
				WhichCat=xk[6:]
				MethodOID=MethodOID+1

				stmt=stmt+WithClause(withlist)+'create (igd)-[:ItemRef {Mandatory:"Yes",SourceRef:"'+DSName+'",MethodOID:'+str(MethodOID)+'}]->(id:ItemDef {Name:"'+xk.upper()+'",\
					Label:"Parameter Category '+WhichCat+'",Origin:"'+x['Origin']+'",SASType:"'+x['SASType']+'",DataType:"'+x['DataType']+'",SASLength:"'+x['SASLength']+'"})-\
					[:CodeListRef]->(cl:CodeList {Extensible:"Yes",DataType:"text",Name:"'+DSName+' Parameter Category '+WhichCat+'"}) '+WithClause(withlist)+', id, cl \
					create (id)-[:MethodRef]->(:MethodDef {OID:'+str(MethodOID)+',Description:"See Parameter Page"}) '

				withlist.append('cl')
				withlist.append('id')

				# Create the BASEDON relationship
				if VarSource == 'Standard':
					MSString=xk
				else:
					MSString='PARCATy'

				stmt=stmt+WithClause(withlist)+'match (ms)-[:ItemRef]->(idms:ItemDef {Name:"'+MSString+'"}) '+WithClause(withlist)+',idms create (id)-[:BasedOn]->(idms) '

				withlist.remove('id')

				# Create codelistitem and methodcondition nodes
				# pcatCL is a list of values for PARCATxk (e.g. "parcat2"), across all parameters
				pcatCL = []
				for y in ParmsList:
					pcatCL.append(y['parcats'][xk])

				for y1 in set(pcatCL):
					stmt=stmt+WithClause(withlist)+'create (cl)-[:ContainsCodeListItem]->(:CodeListItem {CodedValue:"'+y1+'"}) '

				# If there's only one value for PARCATxk...
				# if len(set(pcatCL)) == 1:
				# 	stmt=stmt+'set mt.Description="Set to '+pcatCL[0]+'" '+WithClause(withlist)+'create (cl)-[:ContainsCodeListItem]->(:CodeListItem {CodedValue:"'+pcatCL[0]+'"}) '

				# else:
				# 	CodeMaps={}
				# 	for x1,y1 in enumerate(set(pcatCL)):
				# 		stmt=stmt+WithClause(withlist)+'create (cl)-[:ContainsCodeListItem]->(:CodeListItem {CodedValue:"'+y1+'"}) '
				# 		# Prepare for generating method by determining for each value in the PARCATxk codelist, which PARAMCD values map to it.
				# 		# The CodeMaps dictionary will map each unique value of PARCATxk to a list of PARAMCD values that map to it
				# 		CodeMaps[y1]=[]
				# 		for y in ParmsList:
				# 			if y['parcats'][xk] == y1:
				# 				CodeMaps[y1].append(y['paramcd'])

				# 		# Connect the codelist item to a method condition, and connect the condition back to the method def.
				# 		stmt=stmt+'<-[:IfThen {MethodOID:'+str(MethodOID)+'}]-(:MethodCondition {'
				# 		if x1 < len(set(pcatCL))-1:
				# 			stmt=stmt+'Order:'+str(x1+1)+',ElseFL:"N"'
				# 			if len(CodeMaps[y1]) == 1:
				# 				stmt=stmt+',If:"PARAMCD eq '+CodeMaps[y1][0]+'" '
				# 			else:
				# 				stmt=stmt+',If:"PARAMCD in ('
				# 				for y2 in CodeMaps[y1]:
				# 					stmt=stmt+y2+' '
				# 				stmt=stmt+')" '

				# 		else:
				# 			stmt=stmt+'Order:99999,ElseFL:"Y"'

				# 		stmt=stmt+'})<-[:ContainsConditions]-(mt) '

				withlist.remove('cl')

		else:
			MethodOID=MethodOID+1
			stmt=stmt+WithClause(withlist)+'create (igd)-[:ItemRef {Mandatory:"Yes",SourceRef:"'+DSName+'",MethodOID:'+str(MethodOID)+'}]->(id:ItemDef {Name:"'+x['Name']+'",\
				Label:"'+x['Label']+'",Origin:"'+x['Origin']+'",SASType:"'+x['SASType']+'",DataType:"'+x['DataType']+'",SASLength:"'+x['SASLength']+'"}) '

			if x['Name'] == 'DTYPE' and len(ParmsList) == 1:
				# Only one parameter
				stmt=stmt+'-[:CodeListRef]->(cl:CodeList {Extensible:"Yes",DataType:"text",Name:"DTYPE1"}) '

				withlist.append('cl')
				withlist.append('id')

				# Create CodeListItems
				for x1,y1 in ParmsList[0]['dtypes'].iteritems():
					stmt=stmt+WithClause(withlist)+'create (cl)-[:ContainsCodeListItem]->(:CodeListItem {CodedValue:"'+x1+'",Decode:"'+y1+'"}) '
				withlist.remove('cl')

				# Create the method 
				stmt=stmt+WithClause(withlist)+'create (id)-[:MethodRef]->(:MethodDef {OID:'+str(MethodOID)+',Description:"See Parameter Page"}) '

				# Create the BASEDON relationship
				stmt=stmt+WithClause(withlist)+'match (ms)-[:ItemRef]->(idms:ItemDef {Name:"DTYPE"}) '+WithClause(withlist)+',idms create (id)-[:BasedOn]->(idms) '
				withliist.remove('id')

			elif x['Name'] == 'DTYPE':
				stmt=stmt+'-[:ValueListRef]->(vl:ValueListDef) '
				withlist.append('vl')
				# First, create a list of dictionaries, each representing a unique codelist
				CodeLists=[]
				for x1,y1 in enumerate(ParmsList):
					if 'dtypes' in ParmsList[x1]:
						if ParmsList[x1]['dtypes'] not in CodeLists:
							CodeLists.append(ParmsList[x1]['dtypes'])

				# Create the BASEDON relationship
				withlist.append('id')
				stmt=stmt+WithClause(withlist)+'match (ms)-[:ItemRef]->(idms:ItemDef {Name:"DTYPE"}) '+WithClause(withlist)+',idms create (id)-[:BasedOn]->(idms) '

				# Create the method 
				stmt=stmt+WithClause(withlist)+'create (id)-[:MethodRef]->(:MethodDef {OID:'+str(MethodOID)+',Description:"See Parameter Page"}) '
				withlist.remove('id')

				# For each codelist, determine which PARAMCDs are associated with it, create the ItemDef, CodeList, and Where clause
				for x1,y1 in enumerate(CodeLists):
					CLNbr=x1+1
					WhichParms=[]
					for x2,y2 in enumerate(ParmsList):
						if 'dtypes' in ParmsList[x2]:
							if ParmsList[x2]['dtypes'] == y1:
								WhichParms.append(ParmsList[x2]['paramcd'])

					# determine Length attribute
					length=0
					for x2 in CodeLists[x1]:
						length=max(length,len(x2))

					# generate ItemDef, CodeList, Where Clause
					stmt=stmt+WithClause(withlist)+'create (vl)-[:ItemRef {Mandatory:"No",OrderNumber:'+str(CLNbr)+',WCRef:'+str(WCOIDmax)+'}]\
						->(id:ItemDef {Name:"DTYPE'+str(CLNbr)+'",Label:"Derivation Type '+str(CLNbr)+'",Origin:"Assigned",DataType:"text",\
						Length:'+str(length)+'})-[:CodeListRef]->(cl:CodeList {Extensible:"Yes",DataType:"text",Name:"DTYPE'+str(CLNbr)+'"}) '

					withlist.append('id')
					withlist.append('cl')

					# Now add the CodeList Items
					for x2,y2 in CodeLists[x1].iteritems():
						stmt=stmt+WithClause(withlist)+'create (cl)-[:ContainsCodeListItem]->(:CodeListItem {CodedValue:"'+x2+'",Decode:"'+y2+'"}) '

					# Add the where clause
					withlist.remove('cl')
					stmt=stmt+WithClause(withlist)+'create (id)-[:WhereClauseRef]->(:WhereClauseDef {OID:'+str(WCOIDmax)+'})-[:RangeRef]->(rc:RangeCheck {Operator:"'

					if len(WhichParms)>1:
						stmt=stmt+'IN'
					else:
						stmt=stmt+'EQ'
					stmt=stmt+'"}) '

					withlist.append('rc')
					withlist.remove('id')

					for x2 in WhichParms:
						stmt=stmt+WithClause(withlist)+'create (rc)-[:CheckRef]->(:CheckValue {Value:"'+x2+'"}) '
					stmt=stmt+WithClause(withlist)+'match (igd)--(id:ItemDef {Name:"PARAMCD"}) create (rc)-[:Range2Item]->(id) '
					withlist.remove('rc')

			else:
				# Create the CodeList
				stmt=stmt+'-[:CodeListRef]->(cl:CodeList {Extensible:"Yes", '
				if x['Name'] == 'PARAMN':
					stmt=stmt+'DataType:"integer", Name:"'+DSName+' Parameter Number"'
				else:
					stmt=stmt+'DataType:"text", Name:"'+DSName+' Parameter '
					if x['Name'] == 'PARAMCD':
						stmt=stmt+'Code"'
					else:
						stmt=stmt+'Description"'
				stmt=stmt+'}) '

				withlist.append('cl')
				withlist.append('id')

				# Create the method 
				stmt=stmt+WithClause(withlist)+'create (id)-[:MethodRef]->(mt:MethodDef {OID:'+str(MethodOID)+',Description:"See Parameter Page"}) '

				# Add Sources
				if x['Name'] == 'PARAMCD':
					SourceNameList=[]
					for y in ParmsList:
						for x1 in y['sources']:
							sourcelist=x1.split('.')
							SourceNameList.append(sourcelist[0]+'.'+sourcelist[1])
					SourceNameSet=set(SourceNameList)

					# Derive SourceOrder
					maxsrcorder=graph.cypher.execute('match (s:Study {Name:"'+Study+'"})--(igd:ItemGroupDef)--(id1:ItemDef)-[f1:FromSource]->(s1:Source) return distinct max(s1.Order) as maxOrder')[0][0]
					if maxsrcorder:
						SourceOrder=maxsrcorder+1
					else:
						SourceOrder=1

					# Create source nodes and relationships from PARAMCD to them
					for x1,y in enumerate(SourceNameSet):
						stmt=stmt+WithClause(withlist)+'create (id)-[:FromSource {SourceRef:"'+DSName+'", Type:"MAIN"}]->(:Source {Name:"'+y+'",Order:'+str(SourceOrder+x1)+'}) '

				# Create the BASEDON relationship
				stmt=stmt+WithClause(withlist)+'match (ms)-[:ItemRef]->(idms:ItemDef {Name:"'+x['Name']+'"}) '+WithClause(withlist)+',idms create (id)-[:BasedOn]->(idms) '
				

				# Create CodeListItems and attach them to Method conditions
				for x1,y in enumerate(ParmsList):
					stmt=stmt+WithClause(withlist)+'create (cl)-[:ContainsCodeListItem]->(cli:CodeListItem {CodedValue:"'+y[x['Name'].lower()]+'" '
					if x['Name'] != 'PARAM':
						stmt=stmt+', Decode:"'+y['param']+'"'
					stmt=stmt+'}) '

					# if x['Name'] != 'PARAMCD' and len(ParmsList) > 1:
					# 	stmt=stmt+'<-[:IfThen {MethodOID:'+str(MethodOID)+'}]-(:MethodCondition {'

					# 	if x1 < len(ParmsList)-1:
					# 		stmt=stmt+'Order:'+str(x1+1)+',ElseFL:"N",If:"PARAMCD = '+y['paramcd']+'"'
					# 	else:
					# 		stmt=stmt+'Order:99999,ElseFL:"Y"'

					# 	stmt=stmt+'})<-[:ContainsConditions]-(mt) '

					# elif x['Name'] != 'PARAMCD':
					# 	stmt=stmt+'set mt.Description="Set to '+y[x['Name'].lower()]+'" '

					# Connect PARAMCD CODELISTITEMS to their sources
					if x['Name'] == 'PARAMCD':
						withlist.append('cli')
						for y1 in y['sources']:
							SourceName=y1.split('.')[0]+'.'+y1.split('.')[1]
							SourceTest=y1.split('.')[2]
							stmt=stmt+WithClause(withlist)+'match (id)-[:FromSource]->(src:Source {Name:"'+SourceName+'"}) create (cli)-[:ParmSource {Test:"'+SourceTest+'"}]->(src) '
						withlist.remove('cli')

				withlist.remove('cl')
				withlist.remove('id')

	# Create ParameterAttribute relationships from parameter nodes (codelist items of PARAMCD) to their corresponding attributes (e.g. PARCATx, PARAMN, DTYPE)
	for x in ParmsList:
		# if 'dtypes' in x or 'parcats' in x or 'paramn' in x:
		if 'parcats' in x or 'paramn' in x:
			stmt=stmt+'with study, igd match (study)--(igd)--(:ItemDef {Name:"PARAMCD"})--(:CodeList)--(pcd:CodeListItem {CodedValue:"'+x['paramcd']+'"}) '

			if 'parcats' in x:
				for x1,y1 in x['parcats'].iteritems():
					stmt=stmt+'with study,igd,pcd match (igd)--(:ItemDef {Name:"'+x1.upper()+'"})--(:CodeList)--(cli:CodeListItem {CodedValue:"'+y1+'"})\
						with study,igd,pcd,cli create (pcd)-[:ParameterAttribute {Type:"'+x1.upper()+'"}]->(cli) '

			# if 'dtypes' in x:
			# 	stmt=stmt+'with study,igd,pcd match (igd)--(:ItemDef {Name:"DTYPE"})--(:ValueListDef)--(id:ItemDef)--(:WhereClauseDef)--(:RangeCheck)--\
			# 		(cv:CheckValue {Value:"'+x['paramcd']+'"}) with study,igd,pcd,cv,id match (id)--(:CodeList)--(cli:CodeListItem) with study,igd,pcd,cli \
			# 		create (pcd)-[:ParameterAttribute {Type:"DTYPE"}]->(cli) ' 

			if 'paramn' in x:
				stmt=stmt+'with study,igd,pcd match (igd)--(:ItemDef {Name:"PARAMN"})--(:CodeList)--(cli:CodeListItem {CodedValue:"'+x['paramn']+'"})\
					with study,igd,pcd,cli create (pcd)-[:ParameterAttribute {Type:"PARAMN"}]->(cli) '

	print 'ADDPARMS STMT: '+stmt
	tx=graph.cypher.begin()
	tx.append(stmt)
	tx.commit()

	PMPDict=PrepMainPredList(StandardName,StandardVersion,Study,DSName,'BASIC DATA STRUCTURE',VarSource)
	PMPDict['SourceJoin']=''
	PMPDict['SourceName']=SourceName
	PMPDict['SourceOrder']=SourceOrder
	PMPDict['SourceDescription']=''
	PMPDict['ReturnTo']='mainpredlist'
	PMPDict['RSType']='MAIN'
	PMPDict['DSName']=DSName
	PMPDict['Class']='BASIC DATA STRUCTURE'
	PMPDict['Study']=Study
	PMPDict['StandardName']=StandardName
	PMPDict['StandardVersion']=StandardVersion
	Instruction='Predecessor variables are those that are copied from one data set to another.  '
	Instruction=Instruction+'Define predecessors from '+SourceName+'.'

	return render(request,'StandardDeveloper1/mainpredlist.html', PMPDict)


def QueryStudy(request):
	# Get the default standard to which the study is attached

	# When editing a current study
	if 'indexchoice' in request.GET:
		StandardName=request.GET['indexchoice'][4:]
		Study=request.GET['studies']
		StandardVersion=graph.cypher.execute('match (a:Study {Name:"'+Study+'"})-[:BasedOn]->(b:Standard {Name:"'+StandardName+'"}) return b.Version as Version')[0][0]

	# When finished defining the previous data set
	else:
		StandardName=request.GET['StandardName']
		StandardVersion=request.GET['StandardVersion']
		Study=request.GET['Study']

	if "make_spec" in request.GET:
		GenerateADaMSpec(Study)
		message="Spec.xlsx generated"

	elif "make_define" in request.GET:
		GenerateDefine()
		message="Define.xml generated"

	else:
		message=""

	# Get Standard data sets
	StandardDSRL=QStandardDS(StandardName,StandardVersion)

	# Get Study data sets
	StudyDSRL=QStudyDS(Study)

	# Get a list of standard data sets that does not overlap with the list of study data sets
	StandardList=EliminateDups(StandardDSRL,StudyDSRL)

	# Get the classes for the model
	Classes=graph.cypher.execute('match (:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})-[:BasedOn]->(:Model)-[:ItemGroupRef]->(a:ItemGroupDef)-\
		[:ItemRef]->(:ItemDef) return distinct a.Name')
	# Turn Classes into proper case
	ClassList=[]
	for x1 in Classes:
		Class=''
		for x2 in x1[0].split():
			Class=Class+x2[0:1]+x2[1:].lower()+' '
		ClassList.append(Class.strip())

	return render(request,'StandardDeveloper1/studyhome.html',{'AddDatasets':StandardList,'StudyDatasets':StudyDSRL,'Study':Study,'StandardName':StandardName,'StandardVersion':StandardVersion,'ClassList':ClassList,'Message':message})

def GenerateADaMSpec(request):
	StudyName=request.POST["Study"]
	StandardName=request.POST['StandardName']
	StandardVersion=request.POST['StandardVersion']
	SpecFileName=request.POST['specfilename']

	# Generate Dataset-level spec
	datasetsRL=graph.cypher.execute('match (:Study {Name:"'+StudyName+'"})-[:ItemGroupRef]->(igd:ItemGroupDef)-[rs:RecordSource]->(igdr:ItemGroupDef) \
		optional match (igd)-[:BasedOn]->(igdm:ItemGroupDef)--(:Model) optional match (igd)-[:BasedOn]->(igds:ItemGroupDef)-[:BasedOn]->(igdmm:ItemGroupDef)--(:Model) \
		with igd.Name as Name,igd.Label as Label, case when igdm.Name is not null then igdm.Name else igdmm.Name end as Class,\
		collect(case when rs.Subset <> "" then igdr.Name+" ("+rs.Subset+")" else igdr.Name end) as rsources return Name,Label,Class,\
		reduce(s="",x in rsources|s+", "+x) as RecordSources')
	datasetsDF=pd.DataFrame(datasetsRL.records,columns=datasetsRL.columns)

	print 'DATASETSDF: '
	print datasetsDF 

	# Get variable-level metadata
	stmt='match (:Study {Name:"'+StudyName+'"})-[:ItemGroupRef]->(igd:ItemGroupDef)-[ir:ItemRef]->(id:ItemDef)--(m:MethodDef) where ir.MethodOID=m.OID \
		optional match (id)-[js:JoinSource]->(igdj:ItemGroupDef) \
		optional match (m)--(mc1:MethodCondition)-[it1:IfThen]->(mt:MethodThen) where m.OID=it1.MethodOID \
		optional match (m)--(mc2:MethodCondition)-[it2:IfThen]-(cli:CodeListItem) where m.OID=it2.MethodOID \
		with igd.Name as DSName,id.Name as VarName,id.Label as Label,id.SASType as Type,id.SASLength as Length,ir.Order as VarOrder,case when m.Description is not null then m.Description \
		when mc1.Order=1 then "If "+mc1.If+" then "+mt.Then when mc1.Order=99999 then "; Else "+mt.Then when mc1.Order is not null then "; else if "+mc1.If+" then "+mt.Then \
		when mc2.Order=1 then "If "+mc2.If+" then "+cli.CodedValue when mc2.Order=99999 then " else "+cli.CodedValue else "; else if "+mc2.If+" then "+cli.CodedValue end as instruction,\
		case when mc1.Order is not null then mc1.Order else  mc2.Order end as OrderNew,collect(igdj.Name) as jsnames \
		order by DSName,VarOrder,VarName,OrderNew \
		with DSName,VarOrder,VarName,Label,Type,Length,jsnames,collect(instruction) as icoll \
		return DSName,VarName,Label,Type,Length,\
		reduce(s=head(jsnames),x in jsnames|s+", "+x) as JoinSources,\
		reduce(inst="",x in icoll|inst+" "+x) as Programming order by DSName,VarOrder,VarName'

	print 'VARIABLES STMT: '+stmt
	variablesRL=graph.cypher.execute(stmt)
	variablesDF=pd.DataFrame(variablesRL.records,columns=variablesRL.columns)

	# Get VLM
	stmt='match (:Study {Name:"'+StudyName+'"})-[:ItemGroupRef]->(igd:ItemGroupDef)-[:ItemRef]->(id0:ItemDef)--(:ValueListDef)-[ir:ItemRef]->(id:ItemDef)--(:WhereClauseDef)--(rc:RangeCheck)--(cv:CheckValue) \
		with igd,id0,id,rc,ir,reduce(s="", x in collect(cv.Value)|s+" "+x) as Values match (rc)-[:Range2Item]->(id1:ItemDef) with igd,id0,id,id1,rc,ir,Values match (id)--(m:MethodDef) where ir.MethodOID=m.OID \
		optional match (m)--(mc1:MethodCondition)-[it1:IfThen]->(mt:MethodThen) where m.OID=it1.MethodOID optional match (m)--(mc2:MethodCondition)-[it2:IfThen]-(cli:CodeListItem) where m.OID=it2.MethodOID \
		with igd.Name as DSName,id0.Name as VarName,id.Name as VLMName,collect(id1.Name+" "+rc.Operator+" "+Values) as condition, \
		case when m.Description is not null then m.Description when mc1.Order=1 then "If "+mc1.If+" then "+mt.Then when mc1.Order=99999 then "; Else "+mt.Then when mc1.Order is not null then "; else if "+mc1.If+" then "+mt.Then \
		when mc2.Order=1 then "If "+mc2.If+" then "+cli.CodedValue when mc2.Order=99999 then " else "+cli.CodedValue else "; else if "+mc2.If+" then "+cli.CodedValue end as instruction,\
		case when mc1.Order is not null then mc1.Order else  mc2.Order end as OrderNew order by DSName,VarName,OrderNew with DSName,VarName,VLMName,condition,collect(instruction) as icoll \
		return DSName,VarName,VLMName,reduce(s=head(condition), x in tail(condition)|s+" and "+x) as Condition,reduce(inst="",x in icoll|inst+" "+x) as Programming order by DSName,VarName'

	print 'VLM STMT: '+stmt
	VLMRL=graph.cypher.execute(stmt)
	VLMDF=pd.DataFrame(VLMRL.records,columns=VLMRL.columns)

	# Get parameter information
	parms1RL=graph.cypher.execute('match (:Study {Name:"'+StudyName+'"})--(igd:ItemGroupDef)--(:ItemDef {Name:"PARAMCD"})--(cl:CodeList)--(cli:CodeListItem) \
		optional match (cli)-[:ParameterAttribute {Type:"PARAMN"}]->(pn:CodeListItem) with igd.Name as DSNAME,cli.CodedValue as PARAMCD,cli.Decode as PARAM,pn.CodedValue as PARAMN \
		return DSNAME,PARAMCD,PARAM,PARAMN')
	parms1DF=pd.DataFrame(parms1RL.records,columns=parms1RL.columns)

	# Get DTYPEs for each parameter
	dtypesRL=graph.cypher.execute('match (:Study {Name:"'+StudyName+'"})--(igd:ItemGroupDef)--(:ItemDef {Name:"DTYPE"})--(:ValueListDef)--(id:ItemDef)--(:WhereClauseDef)-[rc:RangeCheck]->(:ItemDef) match (id)--(:CodeList)--(cli:CodeListItem)\
		 with igd.Name as DSNAME,rc.CheckValue as PARAMCD,collect(cli.CodedValue) as dtypes return DSNAME,PARAMCD,reduce(s=head(dtypes), x in tail(dtypes)|s+", "+x) as DTYPE')
	dtypesDF=pd.DataFrame(dtypesRL.records,columns=dtypesRL.columns)

	parmsDF=pd.merge(parms1DF,dtypesDF,how='left',on=['DSNAME','PARAMCD'])


	# Put each PARCAT in its own column
	# This will be accomplished by determining which PARCATs there are, and then iteratively querying for each one
	parcatsRL=graph.cypher.execute('match (:Study {Name:"'+StudyName+'"})--(igd:ItemGroupDef)--(:ItemDef {Name:"PARAMCD"})--(cl:CodeList)--(cli:CodeListItem)-[pa:ParameterAttribute]->(:CodeListItem) \
		where substring(pa.Type,0,6)="PARCAT" return distinct pa.Type as PARCAT order by pa.Type')

	for x in parcatsRL:
		parcatRL=graph.cypher.execute('match (:Study {Name:"'+StudyName+'"})--(igd:ItemGroupDef)--(:ItemDef {Name:"PARAMCD"})--(cl:CodeList)--(cli:CodeListItem)-[pa:ParameterAttribute {Type:"'+x[0]+'"}]->(pc:CodeListItem) return \
			igd.Name as DSNAME,cli.CodedValue as PARAMCD,pc.CodedValue as '+x[0])
		parcatDF=pd.DataFrame(parcatRL.records,columns=parcatRL.columns)
		parmsDF=pd.merge(parmsDF,parcatDF,how='left',on=['DSNAME','PARAMCD'])

	# Write spec
	with pd.ExcelWriter(SpecFileName+'.xlsx') as writer:
		datasetsDF.to_excel(writer,sheet_name='Datasets',index=False)
		VarHeaders=['Dataset','Variable','Variable Label','Type','Length','Sources','Main Programming']
		variablesDF[variablesDF['DSName'] == 'ADSL'].to_excel(writer,sheet_name='ADSL',index=False,header=VarHeaders)

		# Now add non-ADSL main spec pages
		# Determine how many DerivedMethodRef DTYPE/PARAMCD combinations there are.  Then make each a column by iterating through the list, query for each combination, merge with variablesRL
		DmethodsRL=graph.cypher.execute('match (:Study {Name:"'+StudyName+'"})-[:ItemGroupRef]->(igd:ItemGroupDef)-[ir:ItemRef]->(id:ItemDef)-[dmr:DerivedMethodRef]-(dmd:DerivedMethodDef) return distinct igd.Name as DSNAME,dmr.PARAMCD as PARAMCD,dmr.DTYPE as DTYPE')
		DmethodsDF=pd.DataFrame(DmethodsRL.records,columns=DmethodsRL.columns)

		for x,y in datasetsDF[datasetsDF['Name'] != 'ADSL'].iterrows():
			MoreVarHeaders=list(VarHeaders)
			DmethodsDSDF=DmethodsDF[DmethodsDF['DSNAME'] == y['Name']]
			variablesDSDF=variablesDF[variablesDF['DSName'] == y['Name']]
			for x1,y1 in DmethodsDSDF.iterrows():
				if not pd.isnull(y1['DTYPE']):
					DmethodRL=graph.cypher.execute('match (:Study {Name:"'+StudyName+'"})-[:ItemGroupRef]->(igd:ItemGroupDef)-[ir:ItemRef]->(id:ItemDef)-[dmr:DerivedMethodRef {PARAMCD:"'+y1['PARAMCD']+'",DTYPE:"'+y1['DTYPE']+'"}]-(dmd:DerivedMethodDef) \
						where igd.Name=dmr.DS return igd.Name as DSName,id.Name as VarName,case when dmd.Description is null then dmd.Type else dmd.Description end as '+y1['PARAMCD']+'_'+y1['DTYPE'])
					MoreVarHeaders.append('PARAMCD='+y1['PARAMCD']+', DTYPE='+y1['DTYPE'])
				else:
					DmethodRL=graph.cypher.execute('match (:Study {Name:"'+StudyName+'"})-[:ItemGroupRef]->(igd:ItemGroupDef)-[ir:ItemRef]->(id:ItemDef)-[dmr:DerivedMethodRef {PARAMCD:"'+y1['PARAMCD']+'"}]-(dmd:DerivedMethodDef) \
						where igd.Name=dmr.DS and not exists(dmr.DTYPE) return dmr.DS as DSName,id.Name as VarName,case when dmd.Description is null then dmd.Type else dmd.Description end as '+y1['PARAMCD'])
					MoreVarHeaders.append('PARAMCD='+y1['PARAMCD'])

				DmethodDF=pd.DataFrame(DmethodRL.records,columns=DmethodRL.columns)
				variablesDSDF=pd.merge(variablesDSDF,DmethodDF,how='left',on=['DSName','VarName'])

			variablesDSDF.to_excel(writer,sheet_name=y['Name'],index=False,header=MoreVarHeaders)
		
			# Now print the Parameters Page
			if y['Class'] == 'BASIC DATA STRUCTURE':
				# subset to just the current data set
				DSDF=parmsDF[parmsDF['DSNAME'] == y['Name']]
				# Determine if PARAMN exists in this data set
				PARAMNExist=1-pd.isnull(DSDF.iloc[0]['PARAMN'])
				# Determine if any parameter has a DTYPE. 
				DTYPEExist=False
				for x1,y1 in DSDF.iterrows():
					if not pd.isnull(y1['DTYPE']):
						DTYPEExist=True
				# What to sort by
				if PARAMNExist:
					SortVar='PARAMN'
				else:
					SortVar='PARAMCD'
				# Columns to include
				cols=['PARAMCD','PARAM']
				if PARAMNExist:
					cols.append('PARAMN')
				for x1 in parcatsRL:
					if not pd.isnull(DSDF.iloc[0][x1[0]]):
						cols.append(x1[0])
						print cols
				if DTYPEExist:
					cols.append('DTYPE')
				DSDF=DSDF[cols].sort_values(by=SortVar)

				DSDF.to_excel(writer,sheet_name=y['Name']+' Parameters',index=False)

				# and now print the VLM page
				VLMDF[['DSName','VarName','Condition','Programming']].to_excel(writer,sheet_name=y['Name']+' Value Level',index=False,header=['Dataset','Variable','Condition','Programming'])

	#return render(request,'StandardDeveloper1/SubsequentStudyHome.html',{'Study':StudyName,'StandardName':StandardName,'StandardVersion':StandardVersion})
	with open(SpecFileName+'.xlsx', 'rb') as fh:
		response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
		response['Content-Disposition'] = 'inline; filename=' + os.path.basename(SpecFileName+'.xlsx')
		return response

def QueryStudyDS(request):
	StandardName=request.GET['StandardName']
	StandardVersion=request.GET['StandardVersion']
	Study=request.GET['Study']

	# First determine the study data set of interest
	for key,val in request.GET.iteritems():
		if key[0:5] == 'edit_':
			DSName=key[5:]

	# Get data set metadata
	MDDS=QStudyDSMD(Study,DSName)

	# Get data set class
	Class=QDSClass(Study,DSName)

	# Get Variable list
	Variables=QStudyDSVarList(Study,DSName)[0]

	return render(request,'StandardDeveloper1/datasethome.html',{'MDDS':MDDS,'Variables':Variables,'Class':Class,'Study':Study,'DSName':DSName,'StandardName':StandardName,'StandardVersion':StandardVersion})

def EditDS(request):
	MDDic=json.loads(request.POST['MD'])
	RSList=json.loads(request.POST['RecordSources'])
	Study=request.POST['Study']
	StandardName=request.POST['StandardName']
	StandardVersion=request.POST['StandardVersion']


	OldRSRL=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(igd:ItemGroupDef {Name:"'+MDDic['Name']+'"})-[rs:RecordSource]->(igdr:ItemGroupDef) \
		with igdr,rs match (x)--(igdr) return igdr.Name as dataset,rs.Subset as subset,count(x) as count')
	OldRSDF=pd.DataFrame(OldRSRL.records,columns=OldRSRL.columns)

	NewRSDF=pd.DataFrame(RSList)

	RSDF=pd.merge(OldRSDF,NewRSDF,how='outer',on=['dataset','subset'],indicator=True)
	RSChange=False
	for x,y in RSDF.iterrows():
		if y['_merge'] != 'both':
			RSChange=True

	if RSChange == True:
		stmt='match (:Study {Name:"'+Study+'"})--(igd:ItemGroupDef {Name:"'+MDDic['Name']+'"}) '
		for x,y in RSDF.iterrows():
			if y['_merge'] == 'left_only':
				stmt=stmt+'with igd match (igd)-[rs:RecordSource {Subset:"'+y['subset']+'"}]->(igdr:ItemGroupDef {Name:"'+y['dataset']+'"}) '
				if y['count'] == 1:
					stmt=stmt+'detach delete igdr '
				else:
					stmt=stmt+'delete rs '

			elif y['_merge'] == 'right_only':
				if y['model'] == 'SDTM':
					stmt=stmt+'with igd create (igd)-[:RecordSource {Subset:"'+y['subset']+'"}]->(:ItemGroupDef {Name:"'+y['dataset']+'"}) '
				else:
					stmt=stmt+'with igd match (:Study)-[:ItemGroupRef]->(igda:ItemGroupDef {Name:"'+y['dataset']+'"}) create (igd)-[:RecordSource {Subset:"'+y['subset']+'"}]->(igda) '

	tx=graph.cypher.begin()
	tx.append(stmt)
	tx.commit()

	return render(request,'StandardDeveloper1/SubsequentStudyHome.html',{'Study':Study,'StandardName':StandardName,'StandardVersion':StandardVersion})



	# stmt='match (:Study {Name:"'+Study+'"})--(igd:ItemGroupDef {Name:"'+MDDic['Name']+'"})-[rs:RecordSource]->(igdr:ItemGroupDef) set igd.Name="'+MDDic['Name']+'",\
	# 	igd.Label="'+MDDic['Label']+'",igd.IsReferenceData="'+MDDic['Reference']+'",igd.Repeating="'+MDDic['Repeating']+'",igd.Structure="'+MDDic['Structure']+'" \
	# 	'



def ESDS(request):
	StandardName=request.POST['StandardName']
	StandardVersion=request.POST['StandardVersion']
	Study=request.POST['Study']
	DSName=request.POST['DSName']
	Label=request.POST['Label']
	Repeat=request.POST['Repeat']
	Structure=request.POST['Structure']
	Reference=request.POST['Reference']
	BeforeDSName=''
	BeforeLabel=''
	BeforeRepeat=''
	BeforeStructure=''
	BeforeReference=''
	if 'BeforeDSName' in request.POST:
		BeforeDSName=request.POST['BeforeDSName']
	if 'BeforeLabel' in request.POST:
		BeforeLabel=request.POST['BeforeLabel']
	if 'BeforeRepeat' in request.POST:
		BeforeRepeat=request.POST['BeforeRepeat']
	if 'BeforeStructure' in request.POST:
		BeforeStructure=request.POST['BeforeStructure']
	if 'BeforeReference' in request.POST:
		BeforeReference=request.POST['BeforeReference']

	print 'REQUEST.POST: '
	print request.POST

	# Edit the data set
	Comma=''
	if BeforeDSName:
		stmt='match (a:Study {Name:"'+Study+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+BeforeDSName+'"}) set '
		if BeforeDSName != DSName:
			stmt=stmt+'b.Name = "'+DSName+'"'
			Comma=', '
	else:
		stmt='match (a:Study {Name:"'+Study+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+DSName+'"}) set '

	if BeforeLabel and BeforeLabel != Label:
		stmt=stmt+Comma+'b.Label="'+Label+'"'
		Comma=', '

	if BeforeRepeat and BeforeRepeat != Repeat:
		stmt=stmt+Comma+'b.Repeat="'+Repeat+'"'
		Comma=', '

	if BeforeStructure and BeforeStructure != Structure:
		stmt=stmt+Comma+'b.Structure="'+Structure+'"'
		Comma=', '

	if BeforeReference and BeforeReference != Reference:
		stmt=stmt+Comma+'b.Reference="'+Reference+'"'
		Comma=', '

	print 'STMT: '+stmt 
	if Comma:
		tx=graph.cypher.begin()
		tx.append(stmt)
		tx.commit()

	# Get data set metadata
	MDDS=QStudyDSMD(Study,DSName)

	# Get data set class
	Class=QDSClass(Study,DSName)

	# Get Variable list
	Variables=QStudyDSVarList(Study,DSName)[0]

	return render(request,'StandardDeveloper1/datasethome.html',{'MDDS':MDDS,'Variables':Variables,'Class':Class,'Study':Study,'DSName':DSName,'StandardName':StandardName,'StandardVersion':StandardVersion})


def NewSource(request):
	RSType=request.POST['RSType']
	IGDName=request.POST['DSName']
	Study=request.POST['Study']
	StandardName=request.POST['StandardName']
	StandardVersion=request.POST['StandardVersion']
	ModelClass=request.POST['Class']
	SourceName=request.POST['SourceName']
	SourceDescription=request.POST['SourceDescription']
	VarSource=''

	if RSType == 'MAIN':
		PMPDict=PrepMainPredList(StandardName,StandardVersion,Study,IGDName,ModelClass,VarSource)
		form='mainpredlist'
		PMPDict['SourceJoin']=''
		PMPDict['ReturnTo']='mainpredlist'
		Instruction='Predecessor variables are those that are copied from one data set to another.  '
		Instruction=Instruction+'Define predecessors from '+SourceName+'.'
	elif RSType == 'MERGE':
		form='mergepredotherlist'
		Instruction='Predecessor variables are those that are copied from one data set to another.  '
		Instruction=Instruction+'Define predecessors from '+SourceName+'.'
		SourceJoin=request.POST["SourceJoin"]
		PMPDict={'StudyVarsRL':QStudyVars(Study,IGDName,'d.Name="'+SourceName+'" and d.Description="'+SourceDescription+'"')}
		PMPDict['SourceJoin']=SourceJoin
		PMPDict['ReturnTo']='mergepredotherlist'
	elif RSType == 'MODEL' or (not RSType and request.POST['RSType_Button'] == 'Add Custom variable'):
		if RSType == 'MODEL':
			ReturnTo='mergepredotherlist'
		else:
			ReturnTo='datasethome'
		RSType='OTHER'
		form='mergepredotherlist'
		Instruction='ADaM allows for the definition of custom variables under guidelines outlined in the IG. '
		Instruction=Instruction+'Define your next custom variable. '
		SourceName=''
		SourceDescription=''
		PMPDict={}
		PMPDict['SourceJoin']=''
		PMPDict['ReturnTo']=ReturnTo


	# Determine SourceOrder by adding 1 to the max SourceOrder that currently exists
	maxOrder=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)-[:FromSource]->(a:Source) return max(a.Order) as Order')
	if maxOrder[0][0]:
		SourceOrder=maxOrder[0][0]+1
	else:
		SourceOrder=1

	PMPDict['Class']=ModelClass
	PMPDict['StandardName']=StandardName
	PMPDict['StandardVersion']=StandardVersion
	PMPDict['Study']=Study
	PMPDict['DSName']=IGDName
	PMPDict['RSType']=RSType
	PMPDict['SourceName']=SourceName
	PMPDict['SourceDescription']=SourceDescription
	PMPDict['SourceOrder']=SourceOrder
	PMPDict['Instruction']=Instruction
	

	return render(request,'StandardDeveloper1/'+form+'.html', PMPDict)


def VarGroupList(DSName,Class):
	if DSName == 'ADSL':
		return ['Study Identifiers','Subject Demographics']

	elif Class == 'OCCURRENCE DATA STRUCTURE':
		return ['Identifier Variables','Dictionary Coding Variables for MedDRA','Dictionary Coding Variables for MedDRA']

	elif Class == 'BASIC DATA STRUCTURE':
		return ['Subject Identifier Variable','Treatment Variables']


def ChangeVarGroup(request):
	DSName=request.POST['DSName']
	Study=request.POST['Study']
	StandardName=request.POST['StandardName']
	StandardVersion=request.POST['StandardVersion']
	Class=request.POST['Class']
	VarGroup=request.POST['VarGroup']
	NextVarGroup=request.POST['NextVarGroup']
	IGDSource=request.POST['IGDSource']

	VGList=VarGroupList(DSName,Class)
	VarGroup=NextVarGroup

	if NextVarGroup:
		NGVIndex=VGList.index(NextVarGroup)
		if NGVIndex == len(VGList)-1:
			NextVarGroup=''
		else:
			NextVarGroup=VGList[NGVIndex+1]

	return render(request,'StandardDeveloper1/variablelist.html', {'Study':Study,'DSName':DSName,'StandardName':StandardName,'StandardVersion':StandardVersion,'Class':Class,'VarGroup':VarGroup,'NextVarGroup':NextVarGroup,'IGDSource':IGDSource})



def JoinSource(request):
	DSName=request.POST['DSName']
	Study=request.POST['Study']
	StandardName=request.POST['StandardName']
	StandardVersion=request.POST['StandardVersion']
	Class=request.POST['Class']
	Sources=request.POST['sources']

	# Get list of study variables
	StudyVarsList=QStudyVars(Study,DSName) 

	return render(request,'StandardDeveloper1/mergepredotherlist.html',{'Study':Study,'DSName':DSName,'StandardName':StandardName,'StandardVersion':StandardVersion,'Class':Class,'Sources':Sources,'VarDefType':'PRED'})

def Record2Join(request):
	DSName=request.POST['DSName']
	Study=request.POST['Study']
	StandardName=request.POST['StandardName']
	StandardVersion=request.POST['StandardVersion']
	Class=request.POST['Class']
	return render(request,'StandardDeveloper1/JoinSourceDef.html',{'Action':'Add','Study':Study,'IGDName':DSName,'VarDefType':'PRED',\
		'StandardName':StandardName,'StandardVersion':StandardVersion,'Class':Class,'SourceForm':'SetMDVar','ReturnTo':'mergepredotherlist'})

def EditStudyVar(request):
	DSName=request.POST['DSName']
	Study=request.POST['Study']
	StandardName=request.POST['StandardName']
	StandardVersion=request.POST['StandardVersion']
	VarSource=request.POST['VarSource']
	VarName=request.POST['VarName']
	VarLabel=request.POST['VarLabel']
	VarSASType=request.POST['VarSASType']
	VarOrigin=request.POST['VarOrigin']
	VarDataType=request.POST['VarDataType']
	VarOrderNumber=request.POST['VarOrderNumber']
	VarMandatory=request.POST['VarMandatory']
	VarSASLength=request.POST['VarSASLength']
	RSType=request.POST['RSType']
	methodchoice=request.POST['methodchoice']
	VLMOID=int(request.POST['WCOID'])
	ModelClass=request.POST['Class']

	if not VLMOID:
		VLM=False
		# get the variable metadata
		VName=request.POST['VarName']
		VLabel=request.POST['VarLabel']
		VSASType=request.POST['VarSASType']
		VOrigin=request.POST['VarOrigin']
		VSASLength=request.POST['VarSASLength']
		VMandatory=request.POST['VarMandatory']
		VOrderNumber=request.POST['VarOrderNumber']
		VDataType=request.POST['VarDataType']

	else:
		VLM=True
		VName=request.POST['VLMName']
		VLabel=request.POST['VLMLabel']
		VOrigin=request.POST['VLMOrigin']
		VLength=request.POST['VLMLength']
		VMandatory=request.POST['VLMMandatory']
		VDataType=request.POST['VLMDataType']

	# MODIFY THE METADATA IN THE DATABASE FOR THIS VARIABLE

	# Get the variable metadata, the sources, the method, and the controlled terminology.
	# Turn each recordlist into a dataframe containing only the columns necessary for comparison

	if VLM:
		VLMString='-(VLD:ValueListDef)-[:ItemRef {WCRefOID:'+str(VLMOID)+'}]->(:ItemDef)-'
		VarMDDic = QStudyVarMD(Study,DSName,VarName)[1]
		VLMDDic = QStudyVarMD(Study,DSName,VarName,VLMOID=VLMOID)[1]
		MethodList = QMethod('Study',Study,0,DSName,VarName,VLMOID=VLMOID)
	else:
		VLMString=''
		VarMDDic = QStudyVarMD(Study,DSName,VarName)[1]
		MethodList = QMethod('Study',Study,0,DSName,VarName)

	DataSources = graph.cypher.execute('match (s:Study {Name:"'+Study+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+DSName+'"})-[:ItemRef]->(c:ItemDef \
		{Name:"'+VarName+'"})-'+VLMString+'[:FromSource]->(d:Source) with s,d match (s)--(:ItemGroupDef)--(e:ItemDef)-[:FromSource]->(d) with s,d,count(e) as Count1 match (s)--(b1:ItemGroupDef)-\
		[:ItemRef {SourceRef:"'+DSName+'"}]->(e1:ItemDef {Name:"'+VarName+'"}) with d,Count1,count(b1) as Count2 return d.Order as Order,Count1>1 or Count2>1 as Count')

	SrcBeforeDF = pd.DataFrame(DataSources.records,columns=DataSources.columns)

	MethodRL = MethodList[0]
	if MethodList[1]==2:
		# See how many times each condition node is being used
		# Then merge that with the method results (from before any changes were made)
		ConditionCountsRL=graph.cypher.execute('match (s:Study {Name:"'+Study+'"})--(:ItemGroupDef {Name:"'+DSName+'"})--(:ItemDef {Name:"'+VarName+'"}) \
			-'+VLMString+'-(:MethodDef)--(mc:MethodCondition) match (m:MethodDef)--(mc) return mc.If as If2,count(m.OID) as Count')
		ConditionCountsDF=pd.DataFrame(ConditionCountsRL.records,columns=ConditionCountsRL.columns)
		MethodBeforeDF1=pd.DataFrame(MethodRL.records,columns=MethodRL.columns)[['If2','ElseFL2','Order2','Then2']]
		MethodBeforeDF=pd.merge(MethodBeforeDF1,ConditionCountsDF,how='left',on='If2')
	else:
		MethodBeforeDF=pd.DataFrame(columns=['If2','ElseFL2','Order2','Then2','Count'])


	CTRL = graph.cypher.execute('match (:Study {Name:"'+Study+'"})-[:ItemGroupRef]->(:ItemGroupDef {Name:"'+DSName+'"})-[:ItemRef]-(:ItemDef {Name:"'+VarName+'"})\
		-'+VLMString+'[:CodeListRef]->(a:CodeList)-[:ContainsCodeListItem]->(b:CodeListItem) match (c:CodeList)-[:ContainsCodeListItem]->(b) return b.CodedValue as CodedValue, count(c) as Count')
	CTBeforeDF = pd.DataFrame(CTRL.records,columns=CTRL.columns)


	# Create dataframes from the collected metadata for comparison to what is currently in the database
	SrcAfter=[]
	for k,v in request.POST.iteritems():
		if k[0:4] == 'src_':
			SrcAfter.append({'Order':int(k[4:]),'New':False})
		elif k[0:7] == 'newsrc_':
			SrcAfter.append({'Order':int(k[7:]),'New':True})

	SrcAfterDF = pd.DataFrame(SrcAfter)
	SrcMergeDF = pd.merge(SrcBeforeDF,SrcAfterDF,how='outer',indicator=True)

	print 'SRCMERGEDF: '
	print SrcMergeDF

	CTAfter=[]
	for k,v in request.POST.iteritems():
		if k[0:8] == 'stdterm_':
			CTAfter.append({'CodedValue':k[8:],'Ext':False,'New':False})
		elif k[0:8] == 'extterm_':
			CTAfter.append({'CodedValue':k[8:],'Ext':True,'New':False})
		elif k[0:8] == 'newterm_':
			CTAfter.append({'CodedValue':k[8:],'Ext':False,'New':True})

	if CTAfter:
		CTAfterDF = pd.DataFrame(CTAfter).sort_values(by='CodedValue')
	else:
		CTAfterDF = pd.DataFrame(columns=['CodedValue'])

	CTMergeDF = pd.merge(CTBeforeDF,CTAfterDF,how='outer',indicator=True)
	if not CTMergeDF.empty:
		CTName=request.POST['CTName']
		CTExt=request.POST['CTExtensible']
		CTAlias=request.POST['CTAlias']
		CTDataType=request.POST['CTDataType']
		CTStdYN=request.POST['CTStdYN']
		PCGCode=request.POST['ParentCodelistGlobal']

	MethodAfter=[]
	if methodchoice == 'conditions':
		cond={}
		res={}
		elseres=request.POST['else']
		for key,val in request.POST.iteritems():
			if key[0:4] == 'cond':
				Order=key[4:]
				MethodAfter.append({'If2':val,'ElseFL2':'N','Order2':int(Order),'Then2':request.POST['res'+Order]})
		MethodAfter.append({'ElseFL2':'Y','Order2':99999,'Then2':request.POST['else']})
		MethodAfterDF=pd.DataFrame(MethodAfter)
	else:
		MethodAfterDF=pd.DataFrame(columns=['If2','ElseFL2','Order2','Then2'])

	MethodMergeDF=pd.merge(MethodBeforeDF,MethodAfterDF,how='outer',indicator=True,suffixes=['_before','_after'],on='If2')
	print 'METHODMERGEDF: '
	print MethodMergeDF

	# Determine if anything changed
	if VLM:
		VLMDefChg = VLMDDic['Name']!=VName or VLMDDic['Label']!=VLabel or VLMDDic['DataType']!=VDataType \
			or VLMDDic['Origin']!=VOrigin or VLMDDic['Length']!=VLength

		VLMRefChg = VLMDDic['Mandatory']!=VMandatory

		# Determine what SASLength should be based on max value at the value level, and what Mandatory should be based on value level
		propsRL=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef {Name:"'+DSName+'"})--(:ItemDef {Name:"'+VarName+'"})--\
			(:ValueListDef)-[r:ItemRef]->(a:ItemDef) where r.WCRefOID <> '+str(VLMOID)+' return max(a.Length) as SASLength,case when "No" in collect(r.Mandatory) then "No" else \
			"Yes" end as Mandatory')
		VarSASLength=max(int(propsRL[0]['SASLength']),int(VLength))
		if propsRL[0]['Mandatory'] == 'No' or VMandatory == 'No':
			VarMandatory='No'
		else:
			VarMandatory='Yes'

		# Note that this definition does not include a change in Origin
		VarDefChg = VarMDDic['Name']!=VarName or VarMDDic['Label']!=VarLabel or VarMDDic['SASType']!=VarSASType or VarMDDic['DataType']!=VarDataType \
			or VarMDDic['SASLength']!=VarSASLength

	else:
		VLMDefChg=False
		VLMRefChg=False

		VarDefChg = VarMDDic['Name']!=VarName or VarMDDic['Label']!=VarLabel or VarMDDic['SASType']!=VarSASType or VarMDDic['DataType']!=VarDataType \
			or VarMDDic['Origin']!=VarOrigin or VarMDDic['SASLength']!=VarSASLength

	VarRefChg = VarMDDic['OrderNumber']!=int(VarOrderNumber) or VarMDDic['Mandatory']!=VarMandatory

	SrcChg = False
	for k,v in SrcMergeDF.iterrows():
		if v['_merge'] in ['left_only','right_only']:
			SrcChg = True


	CTChg = False
	for k,v in CTMergeDF.iterrows():
		if v['_merge'] in ['left_only','right_only']:
			CTChg = True


	MethodChg = False
	if methodchoice=='freetext':
		if request.POST['free'] != MethodRL[0]['Description3']:
			print 'DESCRIPTION BEFORE: '
			print MethodRL[0]['Description3']
			print 'DESCRIPTION AFTER: '
			print request.POST['free']
			MethodChg = True

	for k,v in MethodMergeDF.iterrows():
		if v['_merge'] in ['left_only','right_only'] or (v['_merge'] == 'both' and (v['Order2_before']!=v['Order2_after'] or v['ElseFL2_before']!=v['ElseFL2_after'] or v['Then2_before']!=v['Then2_after'])):
			MethodChg=True

	# Build the Cypher query

	# Count the number of data sets that contain this variable.  If only one, then we can just modify properties, or point it to a new code list or method. 
	# Otherwise, we have to create a new node
	DSVarCount=graph.cypher.execute('match (a:Study {Name:"'+Study+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+DSName+'"})-[:ItemRef]->(c:ItemDef {Name:"'+VarName+'"})\
		with a,c match (a)--(b1:ItemGroupDef)--(c) return count(*)')[0][0]

	# How many variables point to the same codelist that this variable points to
	CLCount=graph.cypher.execute('match (a:Study {Name:"'+Study+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+DSName+'"})-[:ItemRef]->(c:ItemDef {Name:"'+VarName+'"})\
		-'+VLMString+'[:CodeListRef]->(d:CodeList) with (d) match (e:ItemDef)-[:CodeListRef]->(d) return count(e)')[0][0]

	# Now do the same with methods
	MethodCount=graph.cypher.execute('match (s:Study {Name:"'+Study+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+DSName+'"})-[:ItemRef]->(c:ItemDef {Name:"'+VarName+'"})\
		-'+VLMString+'[:MethodRef]->(d:MethodDef) with s,d match (s)--(b1:ItemGroupDef)-[ir:ItemRef]->(c1:ItemDef) where ir.MethodOID=d.OID return count(ir) as count')[0][0]

	# Start generating the statement
	if VLM:
		stmt0='match (st:Study {Name:"'+Study+'"})-[:ItemGroupRef]->(b1:ItemGroupDef {Name:"'+DSName+'"})-[IR0:ItemRef]->(c0:ItemDef {Name:"'+VarName+'"}) \
			--(:ValueListDef)-[IR1:ItemRef {WCRefOID:'+str(VLMOID)+'}]->(c1:ItemDef) '
	else:
		stmt0='match (st:Study {Name:"'+Study+'"})-[:ItemGroupRef]->(b1:ItemGroupDef {Name:"'+DSName+'"})-[IR1:ItemRef]->(c1:ItemDef {Name:"'+VarName+'"}) '

	# If Variable definition property changes or codelist changes, and other data sets point at this variable, then generate a new ItemDef node.
	if (VarDefChg or CTChg) and DSVarCount > 1:
		stmt=stmt0+'optional match (c1)-[:BasedOn]->(cc:ItemDef) with b1,c1,IR1,collect(cc) as cccoll create (b1)-[IR2:ItemRef {Mandatory:"'+VarMandatory+'",Order:'+VarOrderNumber+',MethodRef:"'+DSName+'",SourceRef:"'+DSName+'"}]->\
			(c2:ItemDef {Name:"'+VarName+'",Label:"'+VarLabel+'",SASType:"'+VarSASType+'",SASLength:"'+VarSASLength+'",\
			Origin:"'+VarOrigin+'",DataType:"'+VarDataType+'"}) with b1,c1,c2,cccoll,IR1,IR2 delete IR1 with b1,c1,c2,cccoll,IR2 foreach(x in cccoll|create (c2)-[:BasedOn]->(x)) '
		NewItemDef=True
		AliasNumber='2'
		withlist=['st','c1','c2','IR2']

	else:
		# New ItemDef is not necessary.  Just modify the original one
		if VarDefChg or VarRefChg or VLMDefChg or VLMRefChg:
			stmt=stmt0+'set '
			CommaStr=''
			if VLM:
				IDAlias='c0'
				IRAlias='IR0'
			else:
				IDAlias='c1'
				IRAlias='IR1'

			if VLMDefChg:
				stmt=stmt+'c1.Label="'+VLabel+'" , c1.Length="'+VLength+'" , c1.DataType="'+VarDataType+'" , c1.Origin="'+VOrigin+'" '
				CommaStr=', '

			if VarDefChg:
				stmt=stmt+CommaStr+IDAlias+'.Label="'+VarLabel+'" , '+IDAlias+'.SASType="'+VarSASType+'" , '+IDAlias+'.SASLength="'+str(VarSASLength)+'" , '+IDAlias+'.DataType="'+VarDataType+'" , '+IDAlias+'.Origin="'+VarOrigin+'" '
				CommaStr=', '

			if VarRefChg:
				stmt=stmt+CommaStr+IRAlias+'.Mandatory="'+VarMandatory+'" , '+IRAlias+'.SourceRef="'+DSName+'" , '+IRAlias+'.MethodRef="'+DSName+'" , '+IRAlias+'.Order='+VarOrderNumber+' '
				CommaStr=', '

			if VLMRefChg:
				stmt=stmt+CommaStr+'IR1.Mandatory="'+VMandatory+'" , IR1.SourceRef="'+DSName+'" , IR1.MethodRef="'+DSName+'" '

			withlist=['st','c1','IR1']

			if VLM:
				withlist.append('c0')
				withlist.append('IR0')

		else:
			stmt=''

		NewItemDef=False
		AliasNumber='1'

	# Determine if a new codelist node is needed
	if CTChg and (DSVarCount > 1 or CLCount > 1 or CTBeforeDF.empty):
		NewCLDef=True
	else:
		NewCLDef=False

	# Determine if a new method node is needed
	if MethodChg and MethodCount>1:
		NewMethodDef=True
	else:
		NewMethodDef=False

	# Add Sources

	for k,v in SrcMergeDF.iterrows():
		# Drop an existing source
		if v['_merge']=='left_only':
			if stmt:
				stmt=stmt+WithClause(withlist)+'match (c1) '
			else:
				stmt=stmt0

			stmt=stmt+'-[FS1:FromSource]->(f1:Source {Order:'+str(v['Order'])+'}) '

			if v['Count']:
				stmt=stmt+'delete FS1 '
				if VLM:
					stmt=stmt+WithClause(withlist)+',f1 match (c0)-[FS1:FromSource]->(f1) delete FS1 '
			else:
				stmt=stmt+'detach delete f1 '

		# Keep an existing source
		elif v['_merge']=='both' and NewItemDef:
			if stmt:
				stmt=stmt+WithClause(withlist)+'match (c1) '
			else:
				stmt=stmt0

			stmt=stmt+'-[FS1:FromSource]->(f1:Source {Order:'+str(v['Order'])+'}) create (c2)-[:FromSource {SourceRef:"'+DSName+'",Type:"'+RSType+'"}]->(f1) delete FS1 '

		# Add a new source
		elif v['_merge'] == 'right_only':
			if stmt:
				stmt=stmt+WithClause(withlist)
			else:
				stmt=stmt0+'with c1 '
				withlist=['c1']

			if v['New']:
				SourceOrder=v['Order']
				SourceName=request.POST['newds_newsrc_'+str(SourceOrder)]
				SourceDescription=request.POST['newdesc_newsrc_'+str(SourceOrder)]
				SourceJoin=request.POST['newjoin_newsrc_'+str(SourceOrder)]

				stmt=stmt+'create (c'+AliasNumber+')-[:FromSource {Type:"'+RSType+'", SourceRef:"'+DSName+'"}]->(src:Source {Name:"'+SourceName+'",\
					Description:"'+SourceDescription+'",Join:"'+SourceJoin+'",Order:'+str(SourceOrder)+'}) '

			else:
				stmt=stmt+'match (:Study {Name:"'+Study+'"})-[:ItemGroupRef]->(:ItemGroupDef)-[:ItemRef]->(:ItemDef)-[:FromSource]\
					->(src:Source {Order:'+str(v['Order'])+'}) with c'+AliasNumber+',src limit 1 create (c'+AliasNumber+')-[:FromSource {SourceRef:"'+DSName+'",Type:"'+RSType+'"}]->(src) '

			if VLM:
				stmt=stmt+WithClause(withlist)+',src create (c0)-[:FromSource {Type:"'+RSType+'", SourceRef:"'+DSName+'"}]->(src) '

	# Methods

	# Start by creating or matching the MethodDef node and adding or updating its Description property
	if NewItemDef or MethodChg:
		if not stmt:
			stmt=stmt0
			withlist=['c1','IR1']

		# If a new MethodDef node is needed, create it here
		if NewMethodDef:
			# Calculate new OID
			maxOIDRL=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)--(md1:MethodDef) with max(md1.OID) as max1 \
				optional match (:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)--(:ValueListDef)--(:ItemDef)--(md2:MethodDef) return max1,max(md2.OID) \
				as max2')
			if maxOIDRL:
				MethodOID=max(maxOIDRL[0]['max1'],maxOIDRL[0]['max2'])+1
			else:
				MethodOID=1

			# Create the new MethodDef
			stmt=stmt+WithClause(withlist)+'match (c1)-[mr1:MethodRef]->(m1:MethodDef) '
			withlist.append('mr1')
			withlist.append('m1')

			# If the method was connected to conditions, connect the new MethodDef to the same conditions
			# We'll delete them later if necessary
			if not MethodBeforeDF.empty:
				stmt=stmt+'-[cc1:ContainsConditions]->(mt:MethodCondition)-[it:IfThen]->(x) where it.MethodOID=m1.OID '
				withlist.append('mt')

			stmt=stmt+WithClause(withlist)+'set '

			if not MethodBeforeDF.empty:
				stmt=stmt+'it.MethodOID='+str(MethodOID)+', '

			stmt=stmt+'IR'+AliasNumber+'.MethodOID='+str(MethodOID)+' create (c'+AliasNumber+')-[:MethodRef]->(m:MethodDef {OID:'+str(MethodOID)

			if request.POST['free']:
				stmt=stmt+', Description:"'+request.POST['free']+'"'
			stmt=stmt+'}) '

			if not MethodBeforeDF.empty:
				stmt=stmt+'-[cc:ContainsConditions]->(mt) '
				withlist.remove('mt')

			# delete the link to the old MethodDef and conditions if necessary
			if DSVarCount == 1:
				stmt=stmt+'delete mr1 '

			withlist.remove('mr1')
			withlist.remove('m1')

		else:
			stmt=stmt+WithClause(withlist)+'match (c1)--(m:MethodDef) '

			if NewItemDef:
				stmt=stmt+'create (c2)-[:MethodRef]->(m) set IR2.MethodOID=m.OID '

			if methodchoice == 'freetext':
				stmt=stmt+'set m.Description="'+request.POST['free']+'" '
			elif MethodRL[0]['Description3']:
				stmt=stmt+'remove m.Description '

		withlist.append('m')

		# If conditions are involved
		if not MethodMergeDF.empty:
			for k,v in MethodMergeDF.iterrows():
				# If something changed about a condition...
				if v['_merge']!='both' or v['Order2_before']!=v['Order2_after'] or v['ElseFL2_before']!=v['ElseFL2_after'] or v['Then2_before']!=v['Then2_after']:
					# Find the condition
					if v['_merge']!='right_only':
						stmt=stmt+WithClause(withlist)+'match (m)-[mr:ContainsConditions]->(cond:MethodCondition {Order:'+str(v['Order2_before'])+'})-[it:IfThen]-> '
						withlist.append('cond')
						withlist.append('mr')
						if MethodList[1]==2:
							stmt=stmt+'(mt:MethodThen) where it.MethodOID=m.OID '
							withlist.append('mt')
						elif MethodList[1]==1:
							stmt=stmt+'(cli:CodeListItem)<-[ccli:ContainsCodeListItem]-(:CodeList) where it.MethodOID=m.OID '
							withlist.append('cli')
							withlist.append('ccli')

					# drop unwanted conditions
					if v['_merge']=='left_only':
						if NewMethodDef:
							stmt=stmt+WithClause(withlist)+'delete mr '
							withlist.remove('mr')
						else:
							if MethodCount==1 and v['Count']==1:
								stmt=stmt+WithClause(withlist)+' detach delete cond,mt '
								withlist.remove('cond')
								withlist.remove('mt')
							else:
								stmt=stmt+WithClause(withlist)+' detach delete mt '
								withlist.remove('mt')

					# add/change conditions
					else:
						if MethodCount==1 and v['Count']==1 and v['_merge']!='right_only':
							stmt=stmt+'set cond.IF="'+v['If2']+'", cond.ElseFL="'+v['ElseFL2_after']+'", cond.Order='+str(v['Order2_after'])+', mt.Then="'+v['Then2_after']+'" '
						else:
							stmt=stmt+WithClause(withlist)+'create (m)-[:ContainsConditions]->(:MethodCondition {Order:'+str(v['Order2_after'])+',If:"'+v['If2']+'",\
								ElseFL:"'+v['ElseFL2_after']+'"})-[it:IfThen]->'
							if request.POST['methods']=='m2':
								stmt=stmt+'(:MethodThen {Then:"'+v['Then2_after']+'"}) '
							stmt=stmt+'set it.MethodOID=m.OID '

		withlist.remove('m')

	# Derived record methods
	if not VLM:
		DTMethodsBefeore=graph.cypher.execute('match (s:Study {Name:"'+Study+'"})--(:ItemGroupDef {Name:"'+DSName+'"})--(:ItemDef {Name:"'+VarName+'"})-[dmr:DerivedMethodRef {DS:"'+DSName+'"}]->\
			(dmd:DerivedMethodDef) with dmr,dmd match (s)--(:ItemGroupDef)--(:ItemDef)-[dmr2:DerivedMethodRef]-(dmd) \
			return dmr.PARAMCD as paramcd,dmr.DTYPE as dtype,dmd.Type as type,dmd.Description as description,count(dmr2) as count')

		if DTMethodsBefore:
			DTMethodsBeforeDF=pd.DataFrame(DTMethodsBefore.records,columns=DTMethodsBefore.columns)

		else:
			DTMethodsBeforeDF=pd.DataFrame(columns=['paramcd','dtype','type','description'])

		# Now see what was collected, store in a list of dictionaries, and merge with the dataframe above
		DataList=[]
		DTMethodsAfterDF=pd.DataFrame(columns=['paramcd','dtype','type','description'])
		for k,v in request.POST.iteritems():
			if k[:3] == 'DT_':
				k3=k[3:]
				dic={'paramcd':k3.split('_')[0],'type':v}
				if len(k3.split('_')) == 2:
					dic['dtype']=k3.split('_')[1]
				if 'DTO_'+k[:3] in dic:
					dic['Description'] = request.POST['DTO_'+k[:3]]
				DataList.append(dic)

		DTMethodsAfterDF=pd.DataFrame(DataList)
		DTMethodsMergeDF=pd.merge(DTMethodsBeforeDF,DTMethodsAfterDF,how='outer',indicator=True,suffixes=['_before','_after'],on=['paramcd','dtype'])

		if not DTMethodsMergeDF.empty:
			if not stmt:
				stmt=stmt0

			for k,v in DTMethodsMergeDF.iterrows():
				# If no longer wanted, delete the relationship from the original ItemDef node, or the whole DMD node if only one pointer to it
				if v['_merge'] == 'left_only':
					stmt=stmt+WithClause(withlist)+'match (c1)-[dmr:DerivedMethodRef]->(dmd:DerivedMethodDef) '
					if v['count'] > 1:
						stmt=stmt+'detach delete dmd '
					else:
						stmt=stmt+'delete dmr '

				# Keep an existing relationship DMD, but delete the original relationship if creating a new ItemDef
				elif v['_merge'] == 'both' and NewItemDef:
					stmt=stmt+WithClause(withlist)+'match (c1)-[dmr:DerivedMethodRef]->(dmd:DerivedMethodDef) create (c2)-[:DerivedMethodRef \
						{DS:"'+DSName+'",PARAMCD:"'+v['paramcd']+'",DTYPE:"'+v['dtype']+'"}]->(dmd) delete dmr '

				elif v['_merge'] == 'right_only':
					stmt=stmt+WithClause(withlist)+'create (c'+AliasNumber+')-[:DerivedMethodRef {DS:"'+DSName+'",PARAMCD:"'+v['paramcd']+'",DTYPE:"'+v['dtype']+'"}]\
					->(:DerivedMethodDef {Type:"'+v['type']+'",Description:"'+v['description']+'"})'

	# Add controlled terminology if desired
	if not CTMergeDF.empty:
		stdterms=''
		newterms={}
		extterms=''
		delterms=''

		if CTAfterDF.empty and not NewItemDef:
			if stmt:
				stmt=stmt+WithClause(withlist)+'match (c1) '
			else:
				stmt=stmt0

			if CLCount < 2:
				stmt=stmt+'-[:CodeListRef]->(e:CodeList) detach delete e '
			else:
				stmt=stmt+'-[CR:CodeListRef]->(e:CodeList) delete CR '

		elif not CTAfterDF.empty:
			for k,v in CTMergeDF.iterrows():
				if (v['_merge'] == 'right_only' or (v['_merge'] == 'both' and NewCLDef)) and not v['New'] and not v['Ext']:
					if stdterms:
						stdterms=stdterms+', '
					else:
						stdterms='['
					stdterms=stdterms+"'"+v['CodedValue']+"'"

				elif (v['_merge'] == 'right_only' or (v['_merge'] == 'both' and NewCLDef)) and v['Ext']:
					if extterms:
						extterms=extterms+', '
					else:
						extterms='['
					extterms=extterms+"'"+v['CodedValue']+"'"

				elif v['New'] and not pd.isnull(v['New']):
					newterms[v['CodedValue']]=request.POST['newterm_'+v['CodedValue']] 

				elif v['_merge'] == 'left_only' and not NewCLDef:
					if delterms:
						delterms=delterms+', '
					else:
						delterms='['
					delterms=delterms+"'"+v['CodedValue']+"'"

			if stdterms:
				stdterms=stdterms+'] '

				if stmt and not CTBeforeDF.empty:
					stmt=stmt+WithClause(withlist)+'match (c1)-[CR:CodeListRef]->(d:CodeList) '+WithClause(withlist)+', d, CR '
					withlist.append('d')
					withlist.append('CR')

				elif not stmt:
					withlist.append('c1')
					if not CTBeforeDF.empty:
						withlist.append('d')
						withlist.append('CR')
						stmt=stmt0+'-[CR:CodeListRef]->(d:CodeList) '+WithClause(withlist)
					else:
						stmt=stmt0+WithClause(withlist)

				else:
					stmt=stmt+WithClause(withlist)

				if VarSource=='Standard' and CTStdYN=='Y':
					stmt=stmt+'match (:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})-[:ItemGroupRef]->(:ItemGroupDef {Name:"'+DSName+'"})-[:ItemRef]->(:ItemDef {Name:"'+VarName+'"}) \
						-[:CodeListRef]->(:CodeList)-[:ContainsCodeListItem]->(es:CodeListItem) where es.CodedValue in '+stdterms+WithClause(withlist)+', collect(distinct es) as colles '
					withlist.append('colles')
				elif VarSource=='Model' and CTStdYN=='Y':
					stmt=stmt+'match (a:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})-[:BasedOn]->(:Model)-[:ItemGroupRef]->(:ItemGroupDef {Name:"'+ModelClass+'"})-[:ItemRef]->(:ItemDef {Name:"'+VarName+'"}) \
						-[:CodeListRef]->(:CodeList)-[:ContainsCodeListItem]->(es:CodeListItem) where es.CodedValue in '+stdterms+WithClause(withlist)+', collect(distinct es) as colles '
					withlist.append('colles')
				elif PCGCode:
					stmt=stmt+'match (:CT)--(:CodeList {AliasName:"'+PCGCode+'"})--(es:CodeListItem) where es.CodedValue in '+stdterms+WithClause(withlist)+', collect(distinct es) as colles '
					withlist.append('colles')

			if extterms:
				if stmt:
					if stdterms:
						stmt=stmt+'match (d)-[:ContainsCodeListItem]'
					else:
						stmt=stmt+WithClause(withlist)+'match (c1)-[CR:CodeListRef]->(d:CodeList)-[:ContainsCodeListItem]'
						withlist.append('d')
						withlist.append('CR')
				else:
					stmt=stmt0+'-[CR:CodeListRef]->(d:CodeList)-[:ContainsCodeListItem]'
					withlist.append('d')
					withlist.append('CR')
					withlist.append('c1')

				extterms=extterms+'] '
				stmt=stmt+'->(ee:CodeListItem) where ee.CodedValue in '+extterms+WithClause(withlist)+', collect(ee) as collee '
				withlist.append('collee')

			if delterms:
				delterms=delterms+']'

			if NewCLDef:
				# Delete pointers to relationships if not creating a new ItemDef but are changing an existing codelist
				if not CTBeforeDF.empty and not NewItemDef:
					if stmt:
						if not stdterms and not extterms:
							stmt=stmt+WithClause(withlist)+'match (c1)-[CR:CodeListRef]->(d:CodeList) '+WithClause(withlist)+', d '
							withlist.append('d')

					else:
						withlist.append('c1')
						withlist.append('d')
						stmt=stmt0+'-[CR:CodeListRef]->(d:CodeList) '+WithClause(withlist)

					stmt=stmt+'delete CR '
					withlist.remove('CR')

				# Create a new codelist node
				if not stmt:
					stmt=stmt0
					withlist.append('c1')

				stmt=stmt+WithClause(withlist)+'create (c'+AliasNumber+')-[:CodeListRef]->(e1:CodeList {Name:"'+CTName+'",DataType:"'+CTDataType+'",Extensible:"'+CTExt+'",AliasName:"'+CTAlias+'"}) '
				withlist.append('e1')
				# if stdterms or extterms or newterms:
				# 	stmt=stmt+WithClause(withlist)

				AliasNumber2='e1'

			elif (CTChg or NewItemDef):
				# Find the existing codelist node
				if stmt:
					if not stdterms and not extterms:
						stmt=stmt+WithClause(withlist)+'match (c1)-[CR:CodeListRef]->(d:CodeList) '
						withlist.append('d')
				else:
					withlist.append('d')
					withlist.append('c1')
					stmt=stmt0+'-[CR:CodeListRef]->(d:CodeList) '

				if delterms:
					stmt=stmt+WithClause(withlist)+'match (d)-[rd:ContainsCodeListItem]->(ed:CodeListItem) where ed.CodedValue in '+delterms+' '+WithClause(withlist)+', collect(rd) as collrd \
						foreach(x in collrd|delete x) '

				if NewItemDef:
					stmt=stmt+WithClause(withlist)+'create (c2)-[:CodeListRef]->(d) '

				AliasNumber2='d'

			if stdterms or extterms or newterms:
				stmt=stmt+WithClause(withlist)

			if stdterms:
				stmt=stmt+'foreach(x in colles|create ('+AliasNumber2+')-[:ContainsCodeListItem]->(x)) '

			if extterms:
				stmt=stmt+'foreach(x in collee|create ('+AliasNumber2+')-[:ContainsCodeListItem]->(x)) '

			if newterms:
				for k,v in newterms.iteritems():
					stmt=stmt+'create ('+AliasNumber2+')-[:ContainsCodeListItem]->(:CodeListItem {CodedValue:"'+k+'",Decode:"'+v+'"}) '


	print 'STMT: '+stmt
	tx=graph.cypher.begin()
	tx.append(stmt)
	tx.commit()


	# NOW GET READY FOR RENDERING THE PAGE
	if 'SaveSubset' in request.POST:
		PredStdYN=request.POST['PredStdYN']

		# Create VarDic
		VarDic={'Name':VarName,'Label':VarLabel,'SASType':VarSASType,'OrderNumber':VarOrderNumber,'Mandatory':VarMandatory,'DataType':VarDataType}

		return render(request,'StandardDeveloper1/mdvar.html',{'Study':Study,'DSName':DSName,'VarMD':VarDic,'RSType':RSType,\
			'PredStdYN':PredStdYN,'CTStdYN':'N','Action':"Edit",\
			'InstructionCT':'Optionally, you may define a list of allowable values for this variable called a Codelist.  Click one of the buttons to get started.',\
			'StandardName':StandardName,'Class':ModelClass,'StandardVersion':StandardVersion,'CondExist':True,'FromVar':True})

	else:
		# Get data set metadata
		MDDS=QStudyDSMD(Study,DSName)

		# Get data set class
		Class=QDSClass(Study,DSName)

		# Get Variable list
		Variables=QStudyDSVarList(Study,DSName)[0]

		return render(request,'StandardDeveloper1/datasethome.html',{'MDDS':MDDS,'Variables':Variables,'Class':Class,'Study':Study,'DSName':DSName,'StandardName':StandardName,'StandardVersion':StandardVersion})


def WithClause(list):
	clause='with '
	for x,y in enumerate(set(list)):
		if x>0:
			clause=clause+', '
		clause=clause+y+' '
	return clause


def NewVar(request):
	DSName=request.POST['DSName']
	Study=request.POST['Study']
	StandardName=request.POST['StandardName']
	StandardVersion=request.POST['StandardVersion']
	Class=request.POST['Class']
	IGDSource=request.POST['IGDSource']
	VarGroup=request.POST['VarGroup']
	NextVarGroup=request.POST['NextVarGroup']
	ModelName=request.POST['ModelName']
	ModelVersion=request.POST['ModelVersion']

	VGList=VarGroupList(DSName,Class)
	MD=json.loads(request.POST['MD'])
	Method=json.loads(request.POST['Method'])
	Sources=''
	CT=''

	if request.POST['CT']:
		CT=json.loads(request.POST['CT'])

	if request.POST['Sources']:
		Sources=json.loads(request.POST['Sources'])


	# Calculate the numeric OID value of the method being created
	# by calculating the max value of the current OIDs in the study
	maxOIDRL=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)--(md1:MethodDef) with max(md1.OID) as max1 \
		optional match (:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)--(:ValueListDef)--(:ItemDef)--(md2:MethodDef) return max1,max(md2.OID) \
		as max2')
	if maxOIDRL[0][0]:
		MethodOID=max(maxOIDRL[0]['max1'],maxOIDRL[0]['max2'])+1
	else:
		MethodOID=1

	# Create the ItemDef
	stmt='match (study:Study {Name:"'+Study+'"})--(igd:ItemGroupDef {Name:"'+DSName+'"}) create (igd)-[ir:ItemRef {Mandatory:"'+MD['Mandatory']+'", OrderNumber:'+str(MD['OrderNumber'])+', MethodOID:'+str(MethodOID)+'}]->\
		(id:ItemDef {Name:"'+MD['Name']+'",Label:"'+MD['Label']+'",SASType:"'+MD['SASType']+'",SASLength:'+str(MD['SASLength'])+',DataType:"'+MD['DataType']+'",Origin:"'+MD['Origin']+'"}) '
	
	withlist=['study,igd,id']

	# Create links back to sources
	# Sources is a list of dictionaries, each representing a source, but one may be a dictionary that represents all record sources.  This is identified by the entry RECORDSOURCEYN:Y
	# In this case, Datasets, Models, and Subsets keys will have lists as their values, with an entry in that list per record source.
	if Sources:
		for sourcedic in Sources:
			RecordSource = ('RecordSourceYN' in sourcedic)
			if RecordSource:
				# The following means that the current iteration of the loop represents record sources
				if sourcedic['RecordSourceYN'] == 'Y':
					for x,y in enumerate(sourcedic['Datasets']):
						stmt=stmt+'with '+', '.join(withlist)+' match (igd)-[:RecordSource]->(igdrs:ItemGroupDef {Name:"'+y+'"}) create (id)-[:RecordSource {TargetDS:"'+DSName+'" '
						if MD['Origin'] == 'Predecessor':
							stmt=stmt+',Var:"'+sourcedic['Vars'][x]+'"'
						stmt=stmt+'}]->(igdrs) '

				# The following means that the current iteration represents a join source already defined in the study
				else:
					stmt=stmt+'with '+', '.join(withlist)+' match (igd)-[:JoinSource]-(igdjs:ItemGroupDef {Name:"'+sourcedic['Datasets']+'"}) create (id)-[:JoinSource TargetDS:"'+DSName+'" '
					if MD['Origin'] == 'Predecessor':
						stmt=stmt+',Var:"'+sourcedic['Vars']+'"'
					stmt=stmt+'}]->(igdjs) '

			# The following represents the addition of new join sources
			else:
				if sourcedic['Models'] == 'SDTM':
					stmt=stmt+'with '+', '.join(withlist)+' create (igd)-[:JoinSource {Subset:"'+sourcedic['Subsets']+'",Join:"'+sourcedic['JConditions']+'"}]->(igdjs:ItemGroupDef {Name:"'+sourcedic['Datasets']+'"}) '
				else:
					stmt=stmt+'with '+', '.join(withlist)+' match (study)--(igdjs:ItemGroupDef {Name:"'+sourcedic['Datasets']+'"}) create (igd)-[:JoinSource {Subset:"'+sourcedic['Subsets']+'",Join:"'+sourcedic['JConditions']+'"}]->(igdjs) '

				stmt=stmt+'with '+', '.join(withlist)+',igdjs create (id)-[:JoinSource {TargetDS:"'+DSName+'" '
				if MD['Origin'] == 'Predecessor':
					stmt=stmt+',Var:"'+sourcedic['Vars']+'"'
				stmt=stmt+'}]->(igdjs) '

	# Create CT
	if CT:
		if 'Code' in CT:
			CodeString='AliasName:"'+CT['Code']+'",'
		else:
			CodeString=''

		# If match an already-exising codelist
		if CT['Match']:
			print 'MATCH!!!'
			# If matching a global list
			if CT['MatchType'] == 'Standard':
				#Copy the codelist node
				stmt=stmt+'with '+', '.join(withlist)+' match (:CT)--(cl1:CodeList {AliasName:"'+CT['Code']+'",Name:"'+CT['Name']+'"}) create (id)-[:CodeListRef]->(cl:CodeList) set cl=cl1 '
				withlist.append('cl1')
				withlist.append('cl')
				stmt=stmt+'with '+', '.join(withlist)+' match (cl1)--(cli1:CodeListItem) with '+', '.join(withlist)+',collect(cli1) as coll match (cl) foreach(x in coll|create (cl)-[:ContainsCodeListItem]->(cli:CodeListItem) set cli=x) '
				withlist.remove('cl1')
				withlist.remove('cl')
			# If matching an existing study codelist
			elif CT['MatchType'] == 'Study':
				stmt=stmt+'with '+', '.join(withlist)+' match (study)--(:ItemGroupDef)--(:ItemDef)--(cl:CodeList {'+CodeString+'Name:"'+CT['Name']+'"}) create (id)-[:CodeListRef]->(cl) '
		# No match to anything
		else:
			print 'NO MATCH :('
			stmt=stmt+'with '+', '.join(withlist)+' create (id)-[:CodeListRef]->(cl:CodeList {'+CodeString+'Name:"'+CT['Name']+'",Extensible:"'+CT['Extensible']+'",DataType:"'+CT['DataType']+'"}) '
			withlist.append('cl')
			CTList=CT['CT']
			for x in CTList:
				if 'Decode' in x:
					DecodeString='Decode:"'+x['Decode']+'",'
				else:
					DecodeString=''
				if 'ItemCode' in x:
					ItemCodeString='AliasName:"'+x['ItemCode']+'",'
				else:
					ItemCodeString=''
				stmt=stmt+'with '+', '.join(withlist)+' create (cl)-[:ContainsCodeListItem]->(:CodeListItem {'+DecodeString+ItemCodeString+'CodedValue:"'+x['CodedValue']+'"}) '
			withlist.remove('cl')

	# Create method
	MethodType=Method['MethodType']
	stmt=stmt+'with '+', '.join(withlist)+' create (id)-[:MethodRef]->(md:MethodDef {OID:'+str(MethodOID)
	if MethodType == 'FreeText':
		stmt=stmt+',Description:"'+Method['MethodValue']+'"}) '
	else:
		stmt=stmt+'}) '
		# When a method is defined by conditions
		MethodChoice=request.POST['methodchoice']
		# Determine if conditions were defined with CT (a) 
		if MethodChoice in ['b','c']:
			ConditionList=Method['MethodValue']
			withlist.append('md')
			for x,y in enumerate(ConditionList):
				stmt=stmt+'with '+', '.join(withlist)+' create (md)-[:ContainsConditions]->(:MethodCondition {Order:'
				if y['IfElse'] == 'Else':
					stmt=stmt+'99999,ElseFL:"Y"}) '
				else:
					stmt=stmt+str(x)+',ElseFL:"N",If:"'+y['Condition']+'"}) '
				stmt=stmt+'-[:IfThen {MethodOID:'+str(MethodOID)+'}]->(:MethodThen {Then:"'+y['Result']+'"}) '

		elif MethodChoice == 'a':
			pass


	print 'NEWVAR STMT: '+stmt 
	tx=graph.cypher.begin()
	tx.append(stmt)
	tx.commit()

	return render(request,'StandardDeveloper1/variablelist.html', {'Study':Study,'DSName':DSName,'StandardName':StandardName,'StandardVersion':StandardVersion,'Class':Class,\
		'VarGroup':VGList[0],'NextVarGroup':VGList[1],'IGDSource':IGDSource,'ModelName':ModelName,'ModelVersion':ModelVersion})



def OldNewVar(request):
	ReturnTo = request.POST['ReturnTo']
	VarDefType=request.POST['VarDefType']
	DSName=request.POST['DSName']
	Study=request.POST['Study']
	StandardName=request.POST['StandardName']
	StandardVersion=request.POST['StandardVersion']
	ModelClass=request.POST['Class']
	VarSource=request.POST['VarSource']
	Action=request.POST['Action']
	ParentCLGlobal=request.POST['ParentCodelistGlobal']
	ParentCLStudy=request.POST['ParentCodelistStudy']
	Condition=request.POST['Condition']
	VarName=request.POST['VarName']
	VarLabel=request.POST['VarLabel']
	VarSASType=request.POST['VarSASType']
	VarMandatory=request.POST['VarMandatory']
	VarDataType=request.POST['VarDataType']
	VarOrderNumber=request.POST['VarOrderNumber']
	VarNameSub=request.POST['VarNameSub']
	BeforeVarName=request.POST['BeforeVarName']
	Sources=request.POST['Sources']
	SourcesOut=request.POST['SourcesOut']

	if not Condition:
		VLM=False
		# get the variable metadata
		VName=request.POST['VarName']
		VLabel=request.POST['VarLabel']
		VSASType=request.POST['VarSASType']
		VOrigin=request.POST['VarOrigin']
		VSASLength=request.POST['VarSASLength']
		VMandatory=request.POST['VarMandatory']
		VOrderNumber=request.POST['VarOrderNumber']
		VDataType=request.POST['VarDataType']

	else:
		VLM=True
		VName=request.POST['VLMName']
		VLabel=request.POST['VLMLabel']
		VOrigin=request.POST['VLMOrigin']
		VLength=request.POST['VLMLength']
		VMandatory=request.POST['VLMMandatory']
		VDataType=request.POST['VLMDataType']

	# Calculate the numeric OID value of the method being created
	# by calculating the max value of the current OIDs in the study

	maxOIDRL=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)--(md1:MethodDef) with max(md1.OID) as max1 \
		optional match (:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)--(:ValueListDef)--(:ItemDef)--(md2:MethodDef) return max1,max(md2.OID) \
		as max2')
	if maxOIDRL[0][0]:
		MethodOID=max(maxOIDRL[0]['max1'],maxOIDRL[0]['max2'])+1
	else:
		MethodOID=1


	# Make variables for query parts
	QIGD='(study:Study {Name:"'+Study+'"})--(IGD:ItemGroupDef {Name:"'+DSName+'"}) '
	#QSRC='(src:Source {Name:"'+SourceName+'", Description:"'+SourceDescription+'",Order:'+str(SourceOrder)+',Join:"'+SourceJoin+'"}) '


	# Create ItemDef node
	# For variable-level metadata, set all the properties.  For VLM, set only the name, but also create the ValueListDef and the value-level ItemDef
	# Also with variable-level, see if the ItemDef with all its properties has already been defined for another data set.  If so, use it.
	# This won't be done for VLM.

	stmt='match '+QIGD
	withlist=['study','IGD']
	VarMDDic={}

	if not VLM:
		ItemDefExist=graph.cypher.execute('return exists((:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef {Name:"'+VName+'", Label:"'+VLabel+'",SASType:"'+VSASType+'",Origin:"'+VOrigin+'",SASLength:"'+VSASLength+'",DataType:"'+VDataType+'"}))')[0][0]
		QID='(ID:ItemDef {Name:"'+VName+'", Label:"'+VLabel+'",SASType:"'+VSASType+'",Origin:"'+VOrigin+'",SASLength:"'+VSASLength+'",DataType:"'+VDataType+'"}) '
		QIR='[:ItemRef {Mandatory:"'+VMandatory+'", Order:'+VOrderNumber+',SourceRef:"'+DSName+'", MethodOID:'+str(MethodOID)+'}] '

		if ItemDefExist:
			stmt=stmt+WithClause(withlist)+'match (study)--(:ItemGroupDef)--'+QID
			withlist.append('ID')
			stmt=stmt+WithClause(withlist)+'limit 1 create (IGD)-'+QIR+'->(ID) '
		else:
			stmt=stmt+WithClause(withlist)+' create (IGD)-'+QIR+'->'+QID
			withlist.append('ID')

	else:

		# In case the user chooses a where clause originally defined for another variable, rather than create a new one
		WCOID=request.POST['WCOID']

		# If the variable-level node already exists, then determine if any property values need changing
		VarMDDic = QStudyVarMD(Study,DSName,BeforeVarName)[1]
		if VarMDDic:
			# Determine what SASLength should be based on max value at the value level, and what Mandatory should be based on value level
			propsRL=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef {Name:"'+DSName+'"})--(:ItemDef {Name:"'+VarName+'"})--\
				(:ValueListDef)-[r:ItemRef]->(a:ItemDef) return max(a.Length) as SASLength,case when "No" in collect(r.Mandatory) then "No" else \
				"Yes" end as Mandatory')
			VarSASLength=max(propsRL[0]['SASLength'],int(VLength))
			if propsRL[0]['Mandatory'] == 'No' or VMandatory == 'No':
				VarMandatory='No'
			else:
				VarMandatory='Yes'

			stmt=stmt+'-[IR1:ItemRef]->(ID1:ItemDef {Name:"'+BeforeVarName+'"})-[VLR:ValueListRef]->(VLD:ValueListDef)'
			if VarName != VarMDDic['Name'] or VarLabel != VarMDDic['Label'] or VarSASType != VarMDDic['SASType'] or VarDataType != VarMDDic['DataType'] or VarOrderNumber != VarMDDic['OrderNumber'] or VarSASLength != VarMDDic['SASLength'] or VarMandatory != VarMDDic['Mandatory'] :
				stmt=stmt+' set IR1.Order='+VarOrderNumber+',ID1.Name="'+VarName+'",ID1.Label="'+VarLabel+'",ID1.SASType="'+VarSASType+'",ID1.DataType="'+VarDataType+'" \
					,ID1.SASLength="'+str(VarSASLength)+'",IR1.Mandatory="'+VarMandatory+'" '

		else:
			stmt=stmt+WithClause(withlist)+'create (IGD)-[IR1:ItemRef {Order:'+VarOrderNumber+',Mandatory:"'+VMandatory+'",SourceRef:"'+DSName+'",MethodOID:'+str(MethodOID)+'}]->(ID1:ItemDef {Name:"'+VarName+'",Label:"'+VarLabel+'",SASType:"'+VarSASType+'",\
				DataType:"'+VarDataType+'",SASLength:"'+str(VLength)+'"})-[VLR:ValueListRef]->(VLD:ValueListDef), (ID1)-[:MethodRef]->(:MethodDef {OID:'+str(MethodOID)+',Description:"See Value Level"}) '

		withlist.append('IR1')
		withlist.append('ID1')
		withlist.append('VLD')

		stmt=stmt+WithClause(withlist)+'create (VLD)-[IR:ItemRef {Mandatory:"'+VMandatory+'",MethodOID:'+str(MethodOID)+'}]->(ID:ItemDef {Name:"'+VName+'", Label:"'+VLabel+'",Origin:"'+VOrigin+'",Length:"'+VLength+'",DataType:"'+VDataType+'"}) '
		withlist.remove('VLD')
		withlist.append('IR')
		withlist.append('ID')

		if WCOID:
			stmt=stmt+WithClause(withlist)+'match (study)--(:ItemGroupDef)--(:ItemDef)--(:ValueListDef)--(:ItemDef)--(WC:WhereClauseDef {OID:'+WCOID+'}) with WC limit 1 \
				create (ID)-[:WhereClauseRef]->(WC) '
		else:
			# Calculate a numeric value for the WhereClauseDef OID value by finding the currently existing max value in the study and then adding 1
			WCOIDmax=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)--(:ValueListDef)--(:ItemDef)--(w:WhereClauseDef) return max(w.OID) as OID')[0][0]
			if WCOIDmax:
				WCOID=WCOIDmax+1
			else:
				WCOID=1

			stmt=stmt+'-[:WhereClauseRef]->(WC:WhereClauseDef {OID:'+str(WCOID)+'}) '
			withlist.append('WC')

			# Create the RangeCheck and CheckValue nodes
			for rc in Condition.split(';'):
				cond=rc.split('|')
				stmt=stmt+WithClause(withlist)+'create (WC)-[:RangeRef]->(RC:RangeCheck {Operator:"'+cond[1]+'"}) '
				withlist.append('RC')
				stmt=stmt+WithClause(withlist)+'match (study)--(:ItemGroupDef {Name:"'+DSName+'"})--(ID2:ItemDef {Name:"'+cond[0]+'"}) create (RC)-[:Range2Item]->(ID2) '
				for cv in cond[2].split(','):
					stmt=stmt+WithClause(withlist)+'create (RC)-[:CheckRef]->(:CheckValue {Value:"'+cv+'"}) '
				withlist.remove('RC')

			stmt=stmt+'set IR.WCRefOID='+str(WCOID)+' '


	# Sources

	Sources=json.loads(request.POST['Sources'])
	SourcesList=Sources['sources']

	print 'SOURCES (NEWVAR): '+request.POST['Sources']
	Sources2=json.dumps(Sources)
	print 'SOURCES2 (NEWVAR): '+Sources2 
	print 'SOURCESLIST: '
	print SourcesList

		# Separate SDTM from ADaM sources
		# NOTE:  For now, we are just creating an SDTM ItemGroupDef node without properties.  We will update once SDTM is built into the app.
		# sdtmlist=[]
		# adamlist=[]

		# for x in SourcesList:
		# 	if x['Model'] == 'SDTM':
		# 		sdtmlist.append(x)
		# 	else:
		# 		adamlist.append(x)

		# stmt='match (s:Study {Name:"'+Study+'"})--(igd0:ItemGroupDef {Name:"'+DSName+'"}) '

		# for x in adamlist:
		# 	stmt=stmt+'with s,igd0 match (s)--(igd:ItemGroupDef {Name:"'+x['Dataset']+'"}) create (igd0)-[:RecordSource '
		# 	if x['Subset']:
		# 		stmt=stmt+'{Subset:"'+x['Subset']+'"} '
		# 	stmt=stmt+']->(igd) '

		# for x in sdtmlist:
		# 	stmt=stmt+'with s,igd0 create (igd0)-[:RecordSource '
		# 	if x['Subset']:
		# 		stmt=stmt+'{Subset:"'+x['Subset']+'"} '
		# 	stmt=stmt+']->(:ItemGroupDef {Name:"'+x['Dataset']+'"}) '


	# SrcObj=json.loads(SourcesOut)
	# SrcLst=SrcObj['sourceout']
	for x in SourcesList:
		model=x['Model']
		if model == 'ADaM':
			# Connect to already existing nodes
			if 'JCondition' in x:
				# This tests whether the current source is a record source or a join source
				stmt=stmt+WithClause(withlist)+'match (study)--(igdJ:ItemGroupDef {Name:"'+x['Dataset']+'"}) create (ID)-[:JoinSource {Subset:"'+x['Subset']+'",TargetDS:"'+DSName+'",Join:"'+x['JCondition']+'"}]->(igdJ) '
				if VLM:
					stmt=stmt+', (ID1)-[:JoinSource {Subset:"'+x['Subset']+'",TargetDS:"'+DSName+'",Join:"'+x['JCondition']+'"}]->(igdJ)'
			else:
				stmt=stmt+WithClause(withlist)+'match (study)--(igdR:ItemGroupDef {Name:"'+x['Dataset']+'"}) create (ID)-[:RecordSource {Subset:"'+x['Subset']+'",TargetDS:"'+DSName+'"}]->(igdR) '
				if VLM:
					stmt=stmt+', (ID1)-[:RecordSource {Subset:"'+x['Subset']+'",TargetDS:"'+DSName+'"}]->(igdR) '
		else:
			# For now, we're just creating a new SDTM ItemGroupDef node to connect to
			if 'JCondition' in x:
				stmt=stmt+WithClause(withlist)+'create (ID)-[:JoinSource {Subset:"'+x['Subset']+'",Join:"'+x['JCondition']+'"}]->(:ItemGroupDef {Name:"'+x['Dataset']+'"}) '
				if VLM:
					stmt=stmt+', (ID1)-[:JoinSource {Subset:"'+x['Subset']+'",Join:"'+x['JCondition']+'"}]->(:ItemGroupDef {Name:"'+x['Dataset']+'"}) '
			else:
				stmt=stmt+WithClause(withlist)+'create (ID)-[:RecordSource {Subset:"'+x['Subset']+'",TargetDS:"'+DSName+'"}]->(:ItemGroupDef {Name:"'+x['Dataset']+'"}) '
				if VLM:
					stmt=stmt+', (ID1)-[:RecordSource {Subset:"'+x['Subset']+'",TargetDS:"'+DSName+'"}]->(:ItemGroupDef {Name:"'+x['Dataset']+'"}) '


	# Create DerivedMethod nodes and relationships
	TypesList=[]
	for k,v in request.POST.iteritems():
		if k[:3] == 'DT_':
			k3=k[3:]
			paramcd=k3.split('_')[0]

			otherDesc=request.POST['DTO_'+k3]
			if otherDesc:
				stmt1='return exists((:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)--(:DerivedMethodDef {Type:"'+v+'",Description:"'+otherDesc+'"}))'
			else:
				stmt1='return exists((:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)--(:DerivedMethodDef {Type:"'+v+'"}))'

			# See if the DerivedMethodDef node exists in this study
			drvExist=graph.cypher.execute(stmt1)[0][0] or otherDesc in TypesList or v in TypesList
			if drvExist:
				stmt=stmt+WithClause(withlist)+'match (study)--(:ItemGroupDef)--(:ItemDef)--(drv:DerivedMethodDef {Type:"'+v+'"'
				withlist.append('drv')
				if otherDesc:
					stmt=stmt+',Description:"'+otherDesc+'"'
				stmt=stmt+'}) '+WithClause(withlist)+'limit 1 create (ID)-[:DerivedMethodRef {PARAMCD:"'+paramcd+'",DS:"'+DSName+'"'
				if len(k3.split('_')) == 2:
					stmt=stmt+',DTYPE:"'+k3.split('_')[1]+'"'
				stmt=stmt+'}]->(drv) '
				withlist.remove('drv')

			else:
				stmt=stmt+WithClause(withlist)+'create (ID)-[:DerivedMethodRef {PARAMCD:"'+paramcd+'",DS:"'+DSName+'"'
				if len(k3.split('_')) == 2:
					stmt=stmt+',DTYPE:"'+k3.split('_')[1]+'"'
				stmt=stmt+'}]->(:DerivedMethodDef {Type:"'+v+'"'
				if otherDesc:
					stmt=stmt+',Description:"'+otherDesc+'"'
					TypesList.append(otherDesc)
				else:
					TypesList.append(v)

				stmt=stmt+'}) '


	# Create methods

	if ReturnTo == 'mainpredlist':
		SrcObj=json.loads(SourcesOut)
		SrcDic=SrcObj['sourceout'][0]
		MethodDesc='Copy '+', '.join(k+'.'+v for k,v in SrcDic.iteritems())

	elif VarDefType == 'PRED':
		SrcObj=json.loads(SourcesOut)
		SrcDic=SrcObj['sourceout'][0]
		MethodDesc='Copy '+SrcDic['Dataset']+'.'+SrcDic['Variable']

	else:
		MethodDesc=''
		methodchoice=request.POST['methodchoice']

		if methodchoice == 'freetext':
			MethodDesc=request.POST['free']

		elif methodchoice == 'conditions':
			cond={}
			res={}
			elseres=request.POST['else']

			for key,val in request.POST.iteritems():
				if key[0:4] == 'cond':
					cond[key]=val
				elif key[0:3] == 'res':
					res[key]=val

			stmt=stmt+WithClause(withlist)+'create (ID)-[:MethodRef]->(MD:MethodDef {OID:'+str(MethodOID)+'}) '
			withlist.append('MD')

			for key,val in cond.iteritems():
				Order=key[4:]
				stmt=stmt+WithClause(withlist)+'create (MD)-[:ContainsConditions]->(:MethodCondition {Order:'+Order+',If:"'+val+'",ElseFL:"N"})-[:IfThen {MethodOID:'+str(MethodOID)+'}]->(:MethodThen {Then:"'+res['res'+Order]+'"}) '

			stmt=stmt+WithClause(withlist)+'create (MD)-[:ContainsConditions]->(:MethodCondition {ElseFL:"Y", Order:99999})-[:IfThen {MethodOID:'+str(MethodOID)+'}]->(:MethodThen {Then:"'+elseres+'"}) '
			withlist.remove('MD')


	# For methods with a description, see if one exists already.  If so, connnect to it.  Otherwise, create it.
	if MethodDesc:
		MethodExistRL=graph.cypher.execute('return exists((:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)--(:MethodDef {Description:"'+MethodDesc+'"})) as ID1,\
			exists((:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)--(:ValueListDef)--(:ItemDef)--(:MethodDef {Description:"'+MethodDesc+'"})) as ID2')
		if MethodExistRL[0][0]:
			stmt=stmt+WithClause(withlist)+'match (:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)--(MD:MethodDef {Description:"'+MethodDesc+'"}) '+WithClause(withlist)+',MD limit 1 create (ID)-[:MethodRef]->(mt) '
		elif MethodExistRL[0][1]:
			stmt=stmt+WithClause(withlist)+'match (:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)--(:ValueListDef)--(:ItemDef)--(MD:MethodDef {Description:"'+MethodDesc+'"}) '+WithClause(withlist)+',MD limit 1 create (ID)-[:MethodRef]->(mt) '
		else:
			stmt=stmt+WithClause(withlist)+'create (ID)-[:MethodRef]->(:MethodDef {Description:"'+MethodDesc+'",OID:'+str(MethodOID)+'}) '


	# Create the BasedOn relationship from the study variable to the Standard/Model variable
	if not VarMDDic:
		if VLM:
			Alias='ID1'
		else:
			Alias='ID'

		if VarNameSub:
			OriginalVarName=BeforeVarName
		else:
			OriginalVarName=VarName 

		if VarSource == "Model":
			stmt=stmt+WithClause(withlist)+'match (:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})-[:BasedOn]->(:Model)-[:ItemGroupRef]->(:ItemGroupDef {Name:"'+ModelClass+'"})-[:ItemRef]->(ID2:ItemDef {Name:"'+OriginalVarName+'"}) '+WithClause(withlist)+',ID2 \
				create ('+Alias+')-[:BasedOn]->(ID2) '
		elif VarSource == "Standard":
			stmt=stmt+WithClause(withlist)+'match (:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})-[:ItemGroupRef]->(:ItemGroupDef {Name:"'+DSName+'"})-[:ItemRef]->(ID2:ItemDef {Name:"'+OriginalVarName+'"}) '+WithClause(withlist)+',ID2 create ('+Alias+')-[:BasedOn]->(ID2) '


	# Controlled Terminology
	
	stdterms=''
	newterms={}

	# Put the standard terms into a list
	for x in request.POST:
		if x[0:8] == 'stdterm_':
			if stdterms:
				stdterms=stdterms+', '
			else:
				stdterms='['
			stdterms=stdterms+"'"+x[8:]+"'"

	if stdterms:
		stdterms=stdterms+'] '
		print 'STDTERMS: '+stdterms

	# Put the extended terms into a dictionary
	for k,v in request.POST.iteritems():
		if k[0:8] == 'newterm_':
			newterms[k[8:]]=v
	print 'NEWTERMS: '
	print newterms


	# If there is any controlled terminology...
	if stdterms or newterms:
		CTExt=request.POST['CTExtensible']
		CTAlias=request.POST['CTAlias']
		CTDataType=request.POST['CTDataType']
		CTName=request.POST['CTName']

		# A variable whose standard does not associate CT, but a custom list is based on some standard codelist
		if ParentCLGlobal:
			stdstmt='match (:CT)--(a:CodeList {AliasName:"'+ParentCLGlobal+'"})--(e:CodeListItem) where e.CodedValue in '+stdterms

		elif ParentCLStudy:
			pass

		# CT associated with the variable at the standard level or custom CT
		else:
			# Find the items from the standard that were chosen by the user
			if VarSource == "Standard":
				stdstmt = 'match (:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})-[:ItemGroupRef]->(:ItemGroupDef {Name:"'+DSName+'"})-[:ItemRef]->(:ItemDef {Name:"'+VarName+'"}) \
					-[:CodeListRef]->(:CodeList)-[:ContainsCodeListItem]->(CLI:CodeListItem) where CLI.CodedValue in '+stdterms
			elif VarSource == "Model":
				stdstmt = 'match (:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})-[:BasedOn]->(:Model)-[:ItemGroupRef]->(:ItemGroupDef {Name:"'+ModelClass+'"})-[:ItemRef]->(:ItemDef {Name:"'+VarName+'"}) \
					-[:CodeListRef]->(:CodeList)-[:ContainsCodeListItem]->(CLI:CodeListItem) where CLI.CodedValue in '+stdterms

		# Create the study codelist
		stmt=stmt+WithClause(withlist)+'create (ID)-[:CodeListRef]->(CL:CodeList {Extensible:"'+CTExt+'", DataType:"'+CTDataType+'",Name:"'+CTName+'",AliasName:"'+CTAlias+'"}) '
		withlist.append('CL')

		# Connect the study codelist to the items
		if stdterms:
			stmt=stmt+WithClause(withlist)+stdstmt+' create (CL)-[:ContainsCodeListItem]->(CLI) '

		# Connect the study codelist to extended items
		for k,v in newterms.iteritems():
			stmt=stmt+WithClause(withlist)+'create (CL)-[:ContainsCodeListItem]->(:CodeListItem {CodedValue:"'+k+'",Decode:"'+v+'"})'

	print 'NEWVAR STATEMENT: '+stmt
	print 'REQUEST.POST: '
	print request.POST ;

	tx=graph.cypher.begin()
	tx.append(stmt)
	tx.commit()

	if 'SaveSubset' in request.POST:
		PredStdYN=request.POST['PredStdYN']

		# Create VarDic
		VarDic={'Name':VarName,'Label':VarLabel,'SASType':VarSASType,'OrderNumber':VarOrderNumber,'Mandatory':VarMandatory,'DataType':VarDataType}

		return render(request,'StandardDeveloper1/mdvar.html',{'Study':Study,'DSName':DSName,'VarMD':VarDic,'RSType':RSType,\
			'PredStdYN':PredStdYN,'CTStdYN':'N','CTExt':'Yes','ReturnTo':ReturnTo,'Action':"Add",\
			'InstructionCT':'Optionally, you may define a list of allowable values for this variable called a Codelist.  Click one of the buttons to get started.',\
			'DecodeYN':'','StandardName':StandardName,'Class':ModelClass,\
			'SourceOrder':SourceOrder,'StandardVersion':StandardVersion,'CondExist':True,'FromVar':True})

	elif ReturnTo == 'datasethome':
		MDDS=QStudyDSMD(Study,DSName)

		# Get Variable list
		Variables=QStudyDSVarList(Study,DSName)[0]

		return render(request,'StandardDeveloper1/datasethome.html',{'MDDS':MDDS,'Variables':Variables,'Class':ModelClass,'Study':Study,'DSName':DSName,'StandardName':StandardName,'StandardVersion':StandardVersion})

	else:
		if ReturnTo == 'mainpredlist':
			PMPDict=PrepMainPredList(StandardName,StandardVersion,Study,DSName,ModelClass,VarSource)

		elif VarDefType:
			#PMPDict={'StudyVarsRL':QStudyVarsandSources(Study,DSName,'d.Name="'+SourceName+'" and d.Description="'+SourceDescription+'" and c.Origin="Predecessor"')}
			PMPDict={'StudyVarsRL':QStudyVars(Study,DSName,'c.Origin="Predecessor"')}

		elif RSType == "MODEL":
			PMPDict=PrepModelLists(StandardName,StandardVersion,Study,DSName,ModelClass)

		# elif RSType == 'OTHER':
		# 	OtherVars=graph.cypher.execute('match (a:Study {Name:"'+Study+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+DSName+'"})-[r:ItemRef]->(c:ItemDef) \
		# 		where c.Origin<>"Predecessor" and not (c)-[:BasedOn]->(:ItemDef) return c.Name as Name,c.Label as Label')
		# 	PMPDict={}
		# 	PMPDict['StudyVarsRL']=OtherVars

		print 'STUDYVARSRL: '
		print PMPDict['StudyVarsRL']

		PMPDict['Class']=ModelClass
		PMPDict['StandardName']=StandardName
		PMPDict['StandardVersion']=StandardVersion
		PMPDict['Study']=Study
		PMPDict['DSName']=DSName
		PMPDict['CondExist']=VLM
		PMPDict['Sources']=request.POST['Sources']
		PMPDict['VarDefType']=VarDefType

		return render(request,'StandardDeveloper1/'+ReturnTo+'.html',PMPDict)

def PrepMainPredList(Standard,Version,Study,DSName,Class,VarSource):

	# Query the database for standard predecessors
	if VarSource != "Model":
		StandardPredsRL=QStandardVars(Standard,Version,DSName,'c.Origin="Predecessor"')[0]
		if StandardPredsRL:
			VarSource='Standard'
		else:
			VarSource='Model'

	# If the standard does not define the data set (assuming that all data sets have at least one predecessor), then go to the model
	if VarSource == "Model":
		ModelInfoRL=graph.cypher.execute('match (a:Standard {Name:"'+Standard+'",Version:"'+Version+'"})-[:BasedOn]->(b:Model) return b.name as Name,b.version as Version')
		ModelName=ModelInfoRL[0][0]
		ModelVersion=ModelInfoRL[0][1]
		StandardPredsRL=QModelVars(ModelName,ModelVersion,Class,'c.Origin="Predecessor"')[0]

	# Get list of study variables
	StudyVarsRL=QStudyVars(Study,DSName,'c.Origin="Predecessor"')

	# Remove from the standards list variables already in the study
	StandardPredsList=EliminateDups(StandardPredsRL,StudyVarsRL)

	PMPDict={'StandardPredsList':StandardPredsList,'StudyVarsRL':StudyVarsRL,'VarSource':VarSource}
	return PMPDict

def PrepModelLists(StandardName,StandardVersion,Study,DSName,ModelClass):
	""" This function creates two lists and a record list.  The record list contains all study model variables.  The lists contain variables in Model
	and standard not in the study (model variables not in the standard)"""

	# Get variables to display
	# Start by getting model variables 
	# To do this, we need to know the model and the class
	model=graph.cypher.execute('match (:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})-[:BasedOn]->(a:Model) return a.name,a.version')
	modelname=model[0][0]
	modelversion=model[0][1]

	ModelRL=QModelVars(modelname,modelversion,ModelClass)[0]

	# Now get standard variables
	StandardRL=QStandardVars(StandardName,StandardVersion,DSName)[0]

	# Now get study variables
	StudyRL=graph.cypher.execute('match (a:Study {Name:"'+Study+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+DSName+'"})-[r:ItemRef]->(c:ItemDef) \
		where (c.Origin<>"Predecessor" or c.Origin IS NULL) and (c)-[:BasedOn]->(:ItemDef) return c.Name as Name order by r.OrderNumber,c.Origin as Origin')

	StandardVarList=EliminateDups(StandardRL,StudyRL)
	ModelVarList=[]
	for x in ModelRL:
		ModelVarList.append(x[0])
	for x in StandardVarList:
		if x in ModelVarList:
			ModelVarList.remove(x)
	for x in StudyRL:
		if x[0] in ModelVarList:
			ModelVarList.remove(x[0])

	return {'StudyVarsRL':StudyRL,'StandardList':StandardVarList,'ModelList':ModelVarList}



def QueryStandardDSMD(request):
	StandardName=request.GET["StandardName"]
	StandardVersion=request.GET["StandardVersion"]
	Study=request.GET["Study"]
	Rendered=False

	# If a general structure was selected, then determine which one it was
	# Get the classes for the model
	Classes=graph.cypher.execute('match (:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})-[:BasedOn]->(:Model)-[:ItemGroupRef]->(a:ItemGroupDef)-\
		[:ItemRef]->(:ItemDef) return distinct a.Name')
	# Turn Classes into proper case 
	for x1 in Classes:
		Class=''
		for x2 in x1[0].split():
			Class=Class+x2[0:1]+x2[1:].lower()+' '
		if Class.strip() in request.GET:
			Rendered=True
			return render(request,'StandardDeveloper1/mdds.html',{'StandardName':StandardName,'StandardVersion':StandardVersion,'Study':Study,'StandardYN':"N",'Class':Class.strip().upper()})
			break

	if not Rendered:
	# If editing a study data set
		if 'editmdds' in request.GET:
			DSName=request.GET['DSName']
			# Get study data set metadata
			# If the following returns anything, then it came from a standard data set
			MDDS=graph.cypher.execute('match (a:Study {Name:"'+Study+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+DSName+'"})-[:BasedOn]->(:ItemGroupDef)-[:BasedOn]->(c:ItemGroupDef)<-[:ItemGroupRef]\
				-(:Model) return c.Name as Class,b.Name as DSName,b.Structure as Structure,b.Repeating as Repeat,b.IsReferenceData as Reference,b.Label as Label')
			if MDDS:
				StandardYN='Y'
			
			# Otherwise, if the following returns anything, then we know that the data set was defined from a general model class, not a specific standard data set
			else:
				MDDS=graph.cypher.execute('match (a:Study {Name:"'+Study+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+DSName+'"})-[:BasedOn]->(c:ItemGroupDef)<-[:ItemGroupRef]-(:Model) return c.Name as Class,b.Name as DSName,b.Structure as Structure,\
					b.Repeating as Repeat,b.IsReferenceData as Reference,b.Label as Label')
				StandardYN='N'
				# The following will return standard data set metadata, which will tell us which properties the standard has defined.
				# Properties not defined by the standard data set will be left to the user to. populate
			DSMDRL=QStandardDSMD(StandardName,StandardVersion,DSName)

			return render(request,'StandardDeveloper1/mdds.html',{'StandardName':StandardName,'StandardVersion':StandardVersion,'Study':Study,'StandardYN':StandardYN,'StudyDSProps':MDDS,'StandardDSProps':DSMDRL,'DSName':DSName,'Class':MDDS[0]['Class']})

		# If a specific data set was chosen for defining, get the its metadata
		else:
			for k,v in request.GET.iteritems():
				if k[0:4] == "add_":
					DSName=k[4:]
					# Get Standard data set Metadata
					DSMDRL=QStandardDSMD(StandardName,StandardVersion,DSName)
					return render(request,'StandardDeveloper1/mdds.html',{'StandardName':StandardName,'StandardVersion':StandardVersion,'Study':Study,'StandardYN':"Y",'StandardDSProps':DSMDRL,'DSName':DSName,'Class':DSMDRL[0]['Class']})

def SetMDVar(request):
	StandardName=request.POST["StandardName"]
	StandardVersion=request.POST["StandardVersion"]
	Study=request.POST["Study"]
	ReturnTo=request.POST["ReturnTo"]
	DSName=request.POST["DSName"]
	Class=request.POST["Class"]

	Sources=''
	VarDefType=''

	if ReturnTo == 'mergepredotherlist':
		VarDefType=request.POST['VarDefType']

	if VarDefType == 'PRED' or ReturnTo == 'mainpredlist':
		Sources=request.POST['Sources']


	for k in request.POST:
		if k in ["add_rspred","add_other","add_pred"]:

			# Add a custom predecessor or other custom variable
			return render(request,'StandardDeveloper1/mdvar.html',{'Study':Study,'StandardName':StandardName,'DSName':DSName,'VarDefType':VarDefType,'Sources':Sources,'Class':Class,\
				'ReturnTo':ReturnTo,'PredStdYN':"N",'Action':'Add','InstructionCT':'Optionally, you may define a list of allowable values for this variable called a Codelist.  Click one of the buttons to get started.',\
				'CTStdYN':'N','CTExt':'Y','CTTable':'','DecodeYN':'','StandardVersion':StandardVersion,'CondExist':False})

		elif k[:3] == 'add':
			if k[0:6] == 'addstd':
				# From modelvarlist, when adding a standard variable 
				varname = k[7:]
				VarSource = 'Standard'
				PredStdYN = ''
			elif k[0:6] == 'addmod':
				# from modelvarlist, when adding a model variable
				varname = k[7:]
				VarSource = 'Model'
				PredStdYN = ''
			else:
				varname=k[4:]
				VarSource=request.POST['VarSource']
				PredStdYN = 'Y'

			# Determine if the variable name has lower case letters or hyphens in it.  If so, turn on the flag
			rpname=re.compile('[a-z]+|-+')
			VarNameSub=False
			if rpname.search(varname):
				VarNameSub=True

			# Determine if the label requires substitutions by looking for lone y or zz, or the label ends in of
			rplabel=re.compile("of$|\sy$|\sy\s|\szz$|\szz\s")
			VarLabelSub=False
			# Query the database for the variable metadata 
			if VarSource == "Standard":
				VarDic=QStandardVar(StandardName,StandardVersion,DSName,varname)[1]
				CTRL=QStandardCT(StandardName,StandardVersion,DSName,varname)
				MethodList=QMethod('Standard',StandardName,StandardVersion,DSName,varname)
				if rplabel.search(VarDic['Label']):
					VarLabelSub=True


			elif VarSource == "Model":
				ModelInfoRL=graph.cypher.execute('match (a:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})-[:BasedOn]->(b:Model) return b.name as Name,b.version as Version')
				ModelName=ModelInfoRL[0][0]
				ModelVersion=ModelInfoRL[0][1]
				VarDic=QModelVar(ModelName,ModelVersion,ModelClass,varname)[1]
				CTRL=QModelCT(ModelName,ModelVersion,ModelClass,varname)
				MethodList=QMethod('Model',ModelName,ModelVersion,ModelClass,varname)
				if rplabel.search(VarDic['Label']):
					VarLabelSub=True

			if CTRL:
				if CTRL[0].Decode:
					DecodeYN='Y'
				else:
					DecodeYN='N'
				
				InstructionCT = 'The standard associates a list of allowable values called a codelist with the variable '+varname+'.  '
				InstructionCT = InstructionCT+'Check only those values relevant in this context.'
				CTExt = CTRL[0]['Extensible']
				CTStdYN = 'Y'
			else:
				InstructionCT = 'Optionally, you may define a list of allowable values for this variable called a Codelist.  Click one of the buttons to get started.'
				CTStdYN = 'N'
				CTExt = 'Yes'
				DecodeYN=''

			return render(request,'StandardDeveloper1/mdvar.html',{'Study':Study,'DSName':DSName,'VarMD':VarDic,'CTTable':CTRL,'MethodMD':MethodList[0],'MethodType':MethodList[1],'Sources':Sources,\
				'PredStdYN':PredStdYN,'CTStdYN':CTStdYN,'CTExt':CTExt,'ReturnTo':ReturnTo,'Action':"Add",'InstructionCT':InstructionCT,'DecodeYN':DecodeYN,'StandardName':StandardName,\
				'VarSource':VarSource,'Class':Class,'StandardVersion':StandardVersion,'CondExist':False,'VarNameSub':VarNameSub,'VarLabelSub':VarLabelSub})


def QueryStudyVarMD(request):
	StandardName=request.POST["StandardName"]
	StandardVersion=request.POST["StandardVersion"]
	Study=request.POST["Study"]
	DSName=request.POST["DSName"]
	Class=request.POST['Class']
	Instruction="Here we define the metadata for the chosen variable"
	StandardPath='(:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})-[:ItemGroupRef]->(:ItemGroupDef {Name:"'+DSName+'"})'
	ModelPath='(:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})-[:BasedOn]->(:Model)-[:ItemGroupRef]->(:ItemGroupDef {Name:"'+Class+'"})'
	StudyPath='(:Study {Name:"'+Study+'"})-[:ItemGroupRef]->(:ItemGroupDef {Name:"'+DSName+'"})'
	CLPath='-[:CodeListRef]->(:CodeList)'
	CLItemPath='-[:CodeListRef]->(d:CodeList)-[:ContainsCodeListItem]->(e:CodeListItem) return d.Name as Name,d.Extensible as Extensible,d.DataType as DataType,\
					d.AliasName as AliasCL,e.CodedValue as CodedValue,e.AliasName as AliasItem,e.Decode as Decode'
	CLstmt=''


	for k,v in request.POST.iteritems():
		if k[0:8] == 'varedit_':
			Level='Var'
			varname=k[8:]
			VarPath='-[:ItemRef]->(:ItemDef {Name:"'+varname+'"})'

			# Determine if the variable is from the standard, the model, or neither
			StdVarExist=graph.cypher.execute('return exists('+StandardPath+VarPath+')')[0][0]
			if StdVarExist:
				VarSource='Standard'
			else:
				ModelVarExist=graph.cypher.execute('return exists('+ModelPath+VarPath+')')[0][0]
				if ModelVarExist:
					VarSource='Model'
				else:
					VarSource=''

			# Get variable metadata 
			VarMDDic=QStudyVarMD(Study,DSName,varname)[1]
			# Determine if this variable is defined with VLM
			VLM = VarMDDic['VLM']

			if not VLM:
				# Get Data Sources
				DataSources=graph.cypher.execute('match (a:Study {Name:"'+Study+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+DSName+'"})-[:ItemRef]->(c:ItemDef {Name:"'+varname+'"}) \
					-[r:FromSource]->(d:Source) return distinct r.Type as RSType,d.Name as Name,d.Order as Order order by d.Order')

				RSType=DataSources[0]['RSType']

				# These are needed for predecessors
				SourceName=DataSources[0]['Name']
				SourceOrder=DataSources[0]['Order']

				StudyMethodRL=QMethod('Study',Study,0,DSName,varname)
				StudyCLRL=graph.cypher.execute('match '+StudyPath+VarPath+CLItemPath)

				# Begin to generate the Cypher statement
				if StudyCLRL:
					CLstmt='match '+StudyPath+VarPath+CLItemPath

			else:
				RSType=''
				SourceName=''
				SourceOrder=''
				MethodMD=''
				MethodType=''
				StudyCLRL=''
				StudyMethodRL=['','']

		elif k[0:8] == 'vlmedit_':
			Level='VLM'
			VLMOID=k[8:]
			varname=request.POST['VarName']
			varordernumber=request.POST['VarOrderNumber2']
			vardatatype=request.POST['VarDataType2']

			if 'VarSASType2' in request.POST:
				varsastype=request.POST['VarSASType2']
			else:
				varsastype=request.POST['VarSASType']

			if 'VarLabel2' in request.POST:
				varlabel=request.POST['VarLabel2']
			else:
				varlabel=request.POST['VarLabel']

			VarPath='-[:ItemRef]->(:ItemDef {Name:"'+varname+'"})'
			VLMPath='-[:ItemRef]->(:ItemDef {Name:"'+varname+'"})--(:ValueListDef)-[IR:ItemRef {WCRefOID:'+VLMOID+'}]->(:ItemDef)'
			VarSource=request.POST['VarSource']
			VLMDic=QStudyVarMD(Study,DSName,varname,VLMOID=int(VLMOID))[1]
			AllConds=json.loads(request.POST['AllConds'])
			ChosenCondition=AllConds['wc'+VLMOID]

			# Get Data Sources
			DataSources=graph.cypher.execute('match (a:Study {Name:"'+Study+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+DSName+'"})-[:ItemRef]->(c:ItemDef {Name:"'+varname+'"}) \
				--(:ValueListDef)-[:ItemRef {WCRefOID:'+str(VLMOID)+'}]->(:ItemDef)-[r:FromSource]->(d:Source) return distinct r.Type as RSType,d.Name as Name,d.Order as Order order by d.Order')

			RSType=DataSources[0]['RSType']

			# These are needed for predecessors
			SourceName=DataSources[0]['Name']
			SourceOrder=DataSources[0]['Order']

			StudyMethodRL=QMethod('Study',Study,0,DSName,varname,VLMOID=int(VLMOID))
			StudyCLRL=graph.cypher.execute('match '+StudyPath+VLMPath+CLItemPath)

			# Begin to generate the Cypher statement
			if StudyCLRL:
				CLstmt='match '+StudyPath+VLMPath+CLItemPath


	# STUDY-LEVEL CONTROLLED TERMINOLOGY AND METHOD
	# Initiate some values and get the study codelist
	CTStdYN='N'
	DecodeYN='N'
	CTTable=''
	CTExt='Yes'
	StudyList=[]
	StudyAlias=''

	if StudyCLRL:
		# Put the study terms into a list to pass to the template so we know which ones to check
		for x in StudyCLRL:
			StudyList.append(x['CodedValue'])


	# STANDARD-LEVEL CONTROLLED TERMINOLOGY AND METHOD
	if VarSource == 'Standard':
		# Determine if a codelist is assocated with the variable at the standard level
		StdCLExist=graph.cypher.execute('return exists('+StandardPath+VarPath+CLPath+')')[0][0]
		if StdCLExist:
			if CLstmt:
				CLstmt=CLstmt+' union '
			CLstmt=CLstmt+'match '+StandardPath+VarPath+CLItemPath
			CTStdYN='Y'

	elif VarSource == 'Model':
		ModelCLExist=graph.cypher.execute('return exists('+ModelPath+VarPath+CLPath+')')[0][0]
		if ModelCLExist:
			if CLstmt:
				CLstmt=CLstmt+' union '
			CLstmt=CLstmt+'match '+ModelPath+VarPath+CLItemPath
			CTStdYN='Y'

	# Now see if a custom codelist was built from another standard codelist, if the codelist is not associated with the standard variable
	if CTStdYN == 'N' and StudyCLRL:
		# Get the codelist alias of the study codelist
		StudyAlias=StudyCLRL[0]['AliasCL']

		# If StudyAlias exists, then the study codelist must be based on a global codelist
		# If that is the case, then get that codelist and combine it with the study codelist
		if StudyAlias:
			CLstmt=CLstmt+' union match (:CT)--(d:CodeList {AliasName:"'+StudyAlias+'"})--(e:CodeListItem) return d.Name as Name,d.Extensible as Extensible,d.DataType as DataType,\
				d.AliasName as AliasCL,e.CodedValue as CodedValue,e.AliasName as AliasItem,e.Decode as Decode'

	if CLstmt:
		print 'CLSTMT: '+CLstmt
		CTTable=graph.cypher.execute(CLstmt)
		CTExt=CTTable[0]['Extensible']
		if CTTable[0]['Decode']:
			DecodeYN='Y'

	if Level == 'Var':
		return render(request,'StandardDeveloper1/mdvar.html',{'Study':Study,'DSName':DSName,'Instruction':Instruction,'VarMD':VarMDDic,'CTTable':CTTable,'MethodMD':StudyMethodRL[0],'MethodType':StudyMethodRL[1],'RSType':RSType,\
			'PredStdYN':'','CTStdYN':CTStdYN,'CTExt':CTExt,'SourceName':SourceName,'Action':"Edit",'DecodeYN':DecodeYN,'StandardName':StandardName,'ReturnTo':'datasethome',\
			'VarSource':VarSource,'Class':Class,'StudyList':StudyList,'SourceOrder':SourceOrder,'StandardVersion':StandardVersion,'ParentCodelistGlobal':StudyAlias,'CondExist':VLM,'FromVar':True})

	elif Level == 'VLM':
		return render(request,'StandardDeveloper1/mdvar.html',{'Study':Study,'DSName':DSName,'Instruction':Instruction,'VarMD':{'Name':varname,'Label':varlabel,'SASType':varsastype,'OrderNumber':varordernumber,'DataType':vardatatype},'ReturnTo':'datasethome',\
			'VLMD':VLMDic,'CTTable':CTTable,'MethodMD':StudyMethodRL[0],'MethodType':StudyMethodRL[1],'RSType':RSType,'PredStdYN':'','CTStdYN':CTStdYN,'CTExt':CTExt,'SourceName':SourceName,'Action':"Edit",'DecodeYN':DecodeYN,\
			'StandardName':StandardName,'ChosenCondition':ChosenCondition,'VarSource':VarSource,'Class':Class,'StudyList':StudyList,'SourceOrder':SourceOrder,'StandardVersion':StandardVersion,'ParentCodelistGlobal':StudyAlias,'CondExist':True,'FromVar':False})

def QStandardDS(standardname,standardversion=''):
	""" This function returns a single-column Record List of data sets attached to the given standard """
	stmt='match (a:Standard {Name:"'+standardname+'"'
	if standardversion:
		stmt=stmt+',Version:"'+standardversion+'"'
	stmt=stmt+'})-[:ItemGroupRef]->(c:ItemGroupDef) return c.Name as name'
	standarddatasetsRL=graph.cypher.execute(stmt)
	return standarddatasetsRL

def QStandardDSMD(standardname,standardversion,dsname):
	""" This function gets the metadata, including class, for a single data set """
	standarddatasetRL=graph.cypher.execute('match (a:Standard {Name:"'+standardname+'",Version:"'+standardversion+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+dsname+'"})-[:BasedOn]->(c:ItemGroupDef) \
		return b.Name as DSName,b.Repeating as Repeat,b.Reference as Reference,b.Structure as Structure,b.Label as Label,c.Name as Class')
	return standarddatasetRL 

def QStandardVars(standardname,standardversion,dsname,filter=''):
	# Get standard variables
	statement='match (a:Standard {Name:"'+standardname+'",Version:"'+standardversion+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+dsname+'"})\
		-[r:ItemRef]->(c:ItemDef) '
	if filter:
		statement=statement+'where '+filter
	statement=statement+' return c.Name as Name,c.Group as VarGroup order by r.OrderNumber '
	print 'STATEMENT FROM QSTANDARVARS: '+statement
	stdvars=graph.cypher.execute(statement)
	stdvarsDF=pd.DataFrame(stdvars.records,columns=stdvars.columns)
	# stdvars=graph.cypher.execute('match (a:Standard {name:"'+standardname+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+dsname+'"})\
	# 	-[:ItemRef]->(c:ItemDef) return distinct c.Name')
	return [stdvars,stdvarsDF]

def QStandardVar(standardname,standardversion,dsname,varname):
	# This function returns a set of properties for a chosen variable in a chosen data set, in the form of a record list and a dictionary
	RL=graph.cypher.execute('match (a:Standard {Name:"'+standardname+'",Version:"'+standardversion+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+dsname+'"})-[r:ItemRef]->(c:ItemDef {Name:"'+varname+'"}) \
		return distinct c.Name as Name,c.Label as Label,c.SASType as SASType,c.DataType as DataType,c.Origin as Origin,c.MaxLength as SASLength, \
		c.Core as Core,c.CodeList as CodeList,r.OrderNumber as OrderNumber,r.Mandatory as Mandatory')
	df=pd.DataFrame(RL.records,columns=RL.columns)
	print 'STANDARDNAME: '+standardname+', STANDARDVERSION: '+standardversion+', DSNAME: '+dsname+', VARNAME: '+varname 
	print 'DF: '
	print df 
	dic={}
	for k,v in df.iloc[0].iteritems():
		dic[k]=v

	return [RL,dic]


def QStandardCT(standardname,standardversion,dsname,varname):
	return graph.cypher.execute('match (a:Standard {Name:"'+standardname+'",Version:"'+standardversion+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+dsname+'"})-[r:ItemRef]->(c:ItemDef {Name:"'+varname+'"}) \
		-[:CodeListRef]->(d:CodeList)-[:ContainsCodeListItem]->(e:CodeListItem) return d.Name as Name, \
		d.Extensible as Extensible,d.DataType as DataType,d.AliasName as AliasCL,e.CodedValue as CodedValue,e.AliasName as AliasItem,e.Decode as Decode') 

def QMethod(Level,LevelName,LevelVersion,IGDName,VarName,VLMOID=''):
	# This function determine method type and returns a record list whose structure depends on the type of method as well as a method type indicator
	# Level is either Standard or Model or Study
	if Level == "Standard":
		LevelString='match (a:Standard {Name:"'+LevelName+'",Version:"'+LevelVersion+'"})'
	elif Level == "Model":
		LevelString='match (a:Model {name:"'+LevelName+'",version:"'+LevelVersion+'"})'
	elif Level == "Study":
		LevelString='match (a:Study {Name:"'+LevelName+'"})'

	stmt=LevelString+'-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+IGDName+'"})-[:ItemRef]->'

	if VLMOID:
		stmt=stmt+'(c0:ItemDef {Name:"'+VarName+'"})-[:ValueListRef]->(:ValueListDef)-[r1:ItemRef {WCRefOID:'+str(VLMOID)+'}]->(c:ItemDef) '
	else:
		stmt=stmt+'(c:ItemDef {Name:"'+VarName+'"}) '

	stmt=stmt+'optional match (c)-[:MethodRef]->(d:MethodDef) optional match (d)-[:ContainsConditions]->(b2:MethodCondition)-[e2:IfThen]->(c2:MethodThen) where d.OID=e2.MethodOID optional match (d)-[:ContainsConditions]->\
		(b1:MethodCondition)-[e1:IfThen]->(c1:CodeListItem) where d.OID=e1.MethodOID return d.Description as Description3,b1.If as If1,b1.Order as Order1,b1.ElseFL as ElseFL1,\
		b2.If as If2,b2.Order as Order2,b2.ElseFL as ElseFL2,c2.Then as Then2,c1.CodedValue as CodedValue1,d.OID as MethodDefOID,e1.MethodOID as ConditionOID1,\
		e2.MethodOID as ConditionOID2 order by b1.Order,b2.Order'

	print 'METHOD STATEMENT: '+stmt

	result=graph.cypher.execute(stmt)

	print 'METHOD RESULTS: '
	print result 

	if result[0][3]:
		return [result,1]
	elif result[0][6]:
		return [result,2]
	elif result[0][0]:
		return [result,3]
	else:
		return [result,'']


def QModelVar(modelname,modelversion,classname,varname):
	# Get all properties of a model variable and return them in a record list and a dictionary
	RL=graph.cypher.execute('match (a:Model {name:"'+modelname+'",version:"'+modelversion+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+classname+'"})-[r:ItemRef]->(c:ItemDef {Name:"'+varname+'"}) \
		return distinct c.Name as Name,c.Label as Label,c.SASType as SASType,c.DataType as DataType,c.Origin as Origin,c.MaxLength as SASLength, \
		c.Core as Core,c.CodeList as CodeList,r.OrderNumber as OrderNumber,r.Mandatory as Mandatory')
	df=pd.DataFrame(RL.records,columns=RL.columns)
	dic={}
	for x,y in df.iloc[0].iteritems():
		dic[x]=y

	return [RL,dic]

def QModelCT(modelname,modelversion,classname,varname):
	return graph.cypher.execute('match (a:Model {name:"'+modelname+'",version:"'+modelversion+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+classname+'"})-[r:ItemRef]->(c:ItemDef {Name:"'+varname+'"}) \
		-[:CodeListRef]->(d:CodeList)-[:ContainsCodeListItem]->(e:CodeListItem) return d.Name as Name, \
		d.Extensible as Extensible,d.DataType as DataType,d.AliasName as AliasCL,e.CodedValue as CodedValue,e.AliasName as AliasItem,e.Decode as Decode') 



def QModelVars(modelname,modelversion,classname,filter=''):
	statement='match (a:Model {name:"'+modelname+'",version:"'+modelversion+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+classname+'"})-[r:ItemRef]->(c:ItemDef)'
	if filter:
		statement=statement+' where '+filter
	statement=statement+' return c.Name as Name,c.Group as VarGroup order by r.OrderNumber '
	ModelVarsRL=graph.cypher.execute(statement)
	ModelVarsDF=pd.DataFrame(ModelVarsRL.records,columns=ModelVarsRL.columns)
	#ModelVarsRL=graph.cypher.execute('match (a:Model {name:"'+modelname+'",version:"'+modelversion+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+classname+'"})-[r:ItemRef]->(c:ItemDef)')
	return [ModelVarsRL,ModelVarsDF]


def QStudyVarMD(studyname,dsname,varname,VLMOID=''):
	# This function returns ItemDef metadata.  If VLMOID is specified, it will return VLM for the condition specified by this OID.
	# When no VLMOID is given, then variable-level metadata is returned, as well as a boolean (VLM) that indicates whether or not this variable is defined with VLM.
	stmt='match (a:Study {Name:"'+studyname+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+dsname+'"})-[r:ItemRef]-> '
	if VLMOID:
		stmt=stmt+'(c0:ItemDef {Name:"'+varname+'"})-[:ValueListRef]->(:ValueListDef)-[r1:ItemRef {WCRefOID:'+str(VLMOID)+'}]->(c:ItemDef) \
			return distinct c.Name as Name,c.Label as Label,c.DataType as DataType,c.Origin as Origin,c.Length as Length, \
			r1.Mandatory as Mandatory,r1.WCRefOID as VLMOID'
	else:
		stmt=stmt+'(c:ItemDef {Name:"'+varname+'"}) return distinct c.Name as Name,c.Label as Label,c.SASType as SASType,c.DataType as DataType,c.Origin as Origin,c.SASLength as SASLength, \
		r.Order as OrderNumber,r.Mandatory as Mandatory,exists((c)--(:ValueListDef)) as VLM'

	RL=graph.cypher.execute(stmt)
	df=pd.DataFrame(RL.records,columns=RL.columns)
	dic={}
	if not df.empty:
		for x,y in df.iloc[0].iteritems():
			dic[x]=y

	return [RL,dic]


def QStudyVars(studyname,dsname,filter=''):
	""" This function returns a record list of all study variables """
	statement='match (a:Study {Name:"'+studyname+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+dsname+'"})-[r:ItemRef]->(c:ItemDef) '
	if filter:
		statement=statement+'where '+filter
	statement=statement+'optional match (c)-[rsr:RecordSource]->(igdr:ItemGroupDef) optional match (c)-[jsr:JoinSource]->(igdj:ItemGroupDef) return c.Name as Name,c.Label as Label,r.Order,\
		collect(igdr.Name) as RecordDatasets,collect(igdj.Name) as JoinDatasets order by r.Order'
	return graph.cypher.execute(statement)

# def QStudyVars(studyname,dsname):
# 	return graph.cypher.execute('match (a:Study {Name:"'+studyname+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+dsname+'"})-[r:ItemRef]->(c:ItemDef) return c.Name as Name,c.Label as Label,c.Origin as Origin')

def QStudyDS(studyname):
	""" This function returns a two-column Record List of study name, data set names for the given study """
	studydatasetsRL=graph.cypher.execute('match (a:Study {Name:"'+studyname+'"})-[:ItemGroupRef]->(c:ItemGroupDef) return c.Name as dsname,a.Name as studyname')
	return studydatasetsRL

def QStudyDSMD(studyname,dsname):
	return graph.cypher.execute('match (a:Study {Name:"'+studyname+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+dsname+'"}) return b.Name as DSName,b.Structure as Structure,\
		b.Repeating as Repeat,b.IsReferenceData as Reference,b.Label as Label')

def QDSClass(Study,DSName):
	# Returns the class, as a string, to which a study's data set is associated
	return graph.cypher.execute('match (a:Study {Name:"'+Study+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+DSName+'"})-[:BasedOn]->(:ItemGroupDef)-[:BasedOn]->(c:ItemGroupDef)<-[:ItemGroupRef] \
		-(:Model) return c.Name union match (a:Study {Name:"'+Study+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+DSName+'"})-[:BasedOn]->(c:ItemGroupDef)<-[:ItemGroupRef]-(:Model) return c.Name')[0][0]

# def QDSSources(Study,DSName):
# 	# Get all sources associated with a study's data sets
# 	return graph.cypher.execute('match (:Study {Name:"'+Study+'"})-[:ItemGroupRef]->(:ItemGroupDef {Name:"'+DSName+'"})-[:ItemRef]->(:ItemDef)-[:FromSource]->(a:Source) \
# 		return distinct a.Name as Name, a.Description as Description, a.Join as Join, a.Order as Order')

def QStudyDSVarList(Study,DSName):
	# Get a study dataset's variables
	RL=graph.cypher.execute('match (:Study {Name:"'+Study+'"})-[:ItemGroupRef]->(:ItemGroupDef {Name:"'+DSName+'"})-[:ItemRef]->(a:ItemDef) return a.Name as Name, a.Label as Label')
	DF=pd.DataFrame(RL.records,columns=RL.columns)
	return [RL,DF]

# def QStudyPreds(studyname,dsname):
# 	""" This function returns a single-column record list of all study variables """
# 	studypredsRL=graph.cypher.execute('match (a:Study {Name:"'+studyname+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+dsname+'"})-[:ItemRef]->(c:ItemDef {Origin:"Predecessor"}) return distinct c.Name as Name')
# 	return studypredsRL 


def EliminateDups(RL1,RL2):
	""" This function returns a list of values from the first column of RecordList1 that are not in the first column of RecordList 2 """
	EDList=[]
	for x in RL1:
		EDList.append(x[0])
	for x in RL2:
		if x[0] in EDList:
			EDList.remove(x[0])
	return EDList


def Test1(request):
	Parms=request.POST['parms']
	ParmsDict=json.loads(Parms)
	ParmsList=ParmsDict['parms']
	print 'ParmsDict: '
	print ParmsDict
	print 'ParmsList: '
	print ParmsList
	for i,v1 in enumerate(ParmsList):
		print 'Row '+str(i)
		for k,v2 in v1.iteritems():
			if k not in ['parcats','dtypes']:
				print k+': '+v2
			else:
				for k2,v3 in v2.iteritems():
					print k2+': '+v3

	return render(request,'StandardDeveloper1/test.html')






def GetAllGlobalCL(request):
	result=graph.cypher.execute('match (:CT)--(a:CodeList)--(b:CodeListItem) with a,b return a.Name as CodeListName,a.AliasName as CLCode,a.Extensible as Extensible,a.DataType as DataType,collect(b.Decode is not null)[0] as DecodeYN')
	df=pd.DataFrame(result.records,columns=result.columns)
	return HttpResponse(df.to_json(orient='records'),content_type='application/json')

# Get terms of a chosen global codelist
def Get1GlobalCodeList(request):
	Code=request.GET['CLCode']
	Name=request.GET['StandardCLName']
	result=graph.cypher.execute('match (:CT)--(a:CodeList {AliasName:"'+Code+'",Name:"'+Name+'"})--(b:CodeListItem) return b.CodedValue as CodedValue,b.Decode as Decode,b.AliasName as ItemCode')
	df=pd.DataFrame(result.records,columns=result.columns)
	return HttpResponse(df.to_json(orient='records'),content_type='application/json')

# Get names and codes of all codelists in a given study
def GetAllStudyCL(request):
	Study=request.GET.get('Study')
	result=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGrouDef)--(:ItemDef)--(cl:CodeList)--(cli:CodeListItem) return cl.AliasName as CLCode,cl.Name as CodeListName,cl.Extensible as Extensible,cl.DataType as DataType,collect(cli.Decode is null)[0] as DecodeYN union \
		match (:Study {Name:"'+Study+'"})--(:ItemGrouDef)--(:ItemDef)--(:ValueListDef)--(:ItemDef)--(clv:CodeList)--(cliv:CodeListItem) return clv.AliasName as CLCode,clv.Name as CodeListName,clv.Extensible as Extensible,clv.DataType as DataType,collect(cliv.Decode is null)[0] as DecodeYN')

	df1=pd.DataFrame(result.records,columns=result.columns) 
	df2=df1.drop_duplicates().sort_values(by=['CLCode','CLName'])
	return HttpResponse(df2.to_json(orient='records'),content_type='application/json')

# Get names of all study codelists with a given code
def GetStudyCodelistsByCode(request):
	Study=request.GET['Study']
	CLCode=request.GET['CLCode']
	stmt='match (:Study {Name:"'+Study+'"})--(:ItemGrouDef)--(:ItemDef)--(cl:CodeList {AliasName:"'+CLCode+'"}) return cl.Name as CLName union \
		match (:Study {Name:"'+Study+'"})--(:ItemGrouDef)--(:ItemDef)--(:ValueListDef)--(:ItemDef)--(clv:CodeList {AliasName:"'+CLCode+'"}) return clv.Name as CLName'
	print 'GETSTUDYCODELISTSBYCODE: '+stmt
	RL=graph.cypher.execute(stmt)

	df1=pd.DataFrame(RL.records,columns=RL.columns) 
	df2=df1.drop_duplicates().sort_values(by='CLName')
	return HttpResponse(df2.to_json(orient='records'),content_type='application/json')

# Get terms of a chosen study codelist
def Get1StudyCodeList(request):
	Study=request.GET['Study']
	CLCode=request.GET['CLCode']
	Name=request.GET['StudyCLName']
	stmt='match (:Study {Name:"'+Study+'"})--(:ItemGrouDef)--(:ItemDef)--(cl:CodeList {AliasName:"'+CLCode+'",Name:"'+Name+'"})--(cli:CodeListItem) return cli.CodedValue as CodedValue,cli.Decode as Decode,cli.AliasName as ItemCode union \
		match (:Study {Name:"'+Study+'"})--(:ItemGrouDef)--(:ItemDef)--(:ValueListDef)--(:ItemDef)--(clv:CodeList {AliasName:"'+CLCode+'",Name:"'+Name+'"})--(clvi:CodeListItem) return clvi.CodedValue as CodedValue,clvi.Decode as Decode,clvi.AliasName as ItemCode'
	print 'Get1StudyCodeList: '+stmt
	RL=graph.cypher.execute(stmt)

	df1=pd.DataFrame(RL.records,columns=RL.columns) 
	df2=df1.drop_duplicates().sort_values(by=['CLCode','CLName'])
	return HttpResponse(df2.to_json(orient='records'),content_type='application/json')

# Get terms of a study codelist combined with terms of associated standard codelist
def Get1StudyWStandardCodeList(request):
	Study=request.GET['Study']
	CLCode=request.GET['CLCode']
	StudyCLName=request.GET['StudyCLName']
	StandardCLName=request.GET['StandardCLName']

	stmtstudy='match (:Study {Name:"'+Study+'"})--(:ItemGrouDef)--(:ItemDef)--(cl:CodeList {AliasName:"'+CLCode+'",Name:"'+StudyCLName+'"})--(cli:CodeListItem) return cli.CodedValue as CodedValue,cli.Decode as Decode,cli.AliasName as ItemCode union \
		match (:Study {Name:"'+Study+'"})--(:ItemGrouDef)--(:ItemDef)--(:ValueListDef)--(:ItemDef)--(clv:CodeList {AliasName:"'+CLCode+'",Name:"'+StudyCLName+'"})--(clvi:CodeListItem) return clvi.CodedValue as CodedValue,clvi.Decode as Decode,clvi.AliasName as ItemCode'
	print 'STUDY STATEMENT: '+stmtstudy
	RLstudy=graph.cypher.execute(stmtstudy)
	dfstudy=pd.DataFrame(RLstudy.records,columns=RLstudy.columns).drop_duplicates()

	stmtstandard='match (:CT)--(a:CodeList {AliasName:"'+CLCode+'",Name:"'+StandardCLName+'"})--(b:CodeListItem) return b.CodedValue as CodedValue,b.Decode as Decode,b.AliasName as ItemCode'
	print 'STANDARD STATEMENT: '+stmtstandard
	RLstandard=graph.cypher.execute(stmtstandard)
	dfstandard=pd.DataFrame(RLstandard.records,columns=RLstandard.columns)

	df=pd.merge(dfstandard,dfstudy,how='outer',on=['CodedValue','ItemCode'],indicator=True)
	# Flag as True the terms in both lists.  These should be pre-checked in the display.
	df2=df1.assign(state=(df1['_merge'] == 'both')).sort_values(by='ItemCode')

	print 'DF2: '
	print df2
	return HttpResponse(df2.to_json(orient='records'),content_type='application/json')

# This compares newly defined CT to already existing study codelists as well as a global list to see if it matches
# Information passeed back to browser:
	# Does the new CT match current study CT or a global list?  Match (boolean)
	# Does it match a study codelist or a global list?  MatchType (Study|Standard|null)
	# Code of the matched list or of the global parent  Code
	# Name of the matched list or of the global parent  Name
def CompareCT(request):
	Study=request.GET['Study']
	StandardCLName=request.GET['StandardCLName']
	StandardCode=request.GET['StandardCode']
	NewCTDF=pd.read_json(request.GET['CT'],orient='records')

	# Get all study codelists
	stmt='match (:Study {Name:"'+Study+'"})--(:ItemGrouDef)--(:ItemDef)--(cl:CodeList)--(cli:CodeListItem) \
		return cl.AliasName as CLCode,cl.Name as CodeListName,cli.CodedValue as CodedValue,cli.Decode as Decode,cli.AliasName as ItemCode union \
		match (:Study {Name:"'+Study+'"})--(:ItemGrouDef)--(:ItemDef)--(:ValueListDef)--(:ItemDef)--(clv:CodeList)--(cliv:CodeListItem) \
		return clv.AliasName as CLCode,clv.Name as CodeListName,cliv.CodedValue as CodedValue,cliv.Decode as Decode,cliv.AliasName as ItemCode'
	print 'STMT from CompareCT: '+stmt
	result=graph.cypher.execute(stmt)

	StudyCTDF=pd.DataFrame(result.records,columns=result.columns).drop_duplicates()

	Both1DF=pd.merge(NewCTDF,StudyCTDF,how='outer',on='CodedValue',indicator=True)
	# Create a boolean column (match) that indicates if the term is in both the study and new codelist
	Both2DF=Both1DF.assign(match=Both1DF['_merge'] == 'both')
	# Reduce down to one line per codelist, taking the minimum of match to indicate whether or not all terms are in both the new and study codelists
	gpb=Both2DF.groupby(['CLCode','CodeListName'])
	summDF=gpb.agg({'match':min})
	# Keep only the codelist that matches exactly with the new codelist
	matchDF=summDF[summDF['match'] == True]

	if matchDF.empty and StandardCLName and StandardCode:
		# i.e. new codelist does not match any current study codelist, then check to see if it matches the standard codelist that was passed, if there is one
		result=graph.cypher.execute('match (:CT)--(a:CodeList {AliasName:"'+StandardCode+'",Name:"'+StandardCLName+'"})--(b:CodeListItem) return b.CodedValue as CodedValue,b.Decode as Decode,b.AliasName as ItemCode')
		StandardDF=pd.DataFrame(result.records,columns=result.columns)
		Both3DF=pd.merge(StandardDF,NewCTDF,how='outer',on='CodedValue',indicator=True)
		Match=True

		for x,y in Both3DF.iterrows():
			if y['_merge'] != 'both' and Match:
				Match=False

		if Match:
			print 'CONDITION 1'
			return HttpResponse(pd.Series({'Match':True,'MatchType':'Standard','Code':StandardCode,'Name':StandardCLName}).to_json(),content_type='application/json')
		else:
			print 'CONDITION 2'
			return HttpResponse(pd.Series({'Match':False,'MatchType':'','Code':StandardCode,'Name':StandardCLName}).to_json(),content_type='application/json')

	elif matchDF.empty:
		print 'CONDITION 3'
		return HttpResponse(pd.Series({'Match':False,'MatchType':'','Code':'','Name':''}).to_json(),content_type='application/json')

	else:
		print 'CONDITION 4'
		return HttpResponse(pd.Series({'Match':True,'MatchType':'Study','Code':matchDF.iloc[0]['CLCode'],'Name':matchDF.iloc[0]['CodeListName']}).to_json(),content_type='application/json')


# def GetStudySources(request):
# 	Study=request.GET['Study']
# 	DSName=request.GET['DSName']
# 	VarName=request.GET['VarName']
# 	VLMOID=request.GET['VLMOID']

# 	stmt='match (s:Study {Name:"'+Study+'"})--(igd:ItemGroupDef)--(id1:ItemDef)-[f1:FromSource]->(s1:Source) return distinct s1.Name,s1.Description,s1.Join,s1.Order,'

# 	if VLMOID:
# 		stmt=stmt+'exists((s)--(:ItemGroupDef {Name:"'+DSName+'"})--(:ItemDef {Name:"'+VarName+'"})--(:ValueListDef)-[:ItemRef {WCRefOID:'+str(VLMOID)+'}]\
# 			->(:ItemDef)--(s1))'

# 	else:
# 		stmt=stmt+'exists((s)--(:ItemGroupDef {Name:"'+DSName+'"})--(:ItemDef {Name:"'+VarName+'"})--(s1))'

# 	print 'SOURCE STMT: '+stmt 

# 	result=graph.cypher.execute(stmt)

# 	resultlist=[]
# 	for x in result:
# 		resultlist.append({'Name':x[0],'Description':x[1],'Join':x[2],'Order':x[3],'VarIn':x[4]})

# 	return JsonResponse(resultlist,safe=False)

def GetStudySources(request):
	Study=request.GET.get('Study')
	DSName=request.GET.get('DSName')
	# Study='oct1835'
	# DSName='ADSL'
	RSources=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef {Name:"'+DSName+'"})-[r:RecordSource]->(igd:ItemGroupDef) with igd.Name as Dataset,r.Subset as Subset,\
		case when exists((igd)-[:BasedOn]->(:ItemGroupDef)) then "ADAM" else "SDTM" end as Model order by Dataset \
		with Dataset,Subset,Model,case when Subset is not null then Model+"."+Dataset+" (where "+Subset+")" else Model+"."+Dataset end as Display order by Dataset \
		with collect(Dataset) as Datasets,collect(Subset) as Subsets,collect(Model) as Models,collect(Display) as Display return Datasets,Subsets,Models,reduce(s=head(Display), x in tail(Display)|s+"; "+x) as Display,"Y" as RecordSourceYN ')

	JSources=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef {Name:"'+DSName+'"})--(:ItemDef)-[js:JoinSource]->(igd:ItemGroupDef) \
		with distinct igd.Name as Datasets,js.Subset as Subsets,case when exists((igd)-[:BasedOn]->(:ItemGrouDef)) then "ADAM" else "SDTM" end as Models, js.Join as JConditions \
		return Datasets,Subsets,Models,case when Subsets is not null then Models+"."+Datasets+" (where "+Subsets+")" else Models+"."+Datasets end as Display,JConditions,"N" as RecordSourceYN')

	RSDF=pd.DataFrame(RSources.records,columns=RSources.columns)
	JSDF=pd.DataFrame(JSources.records,columns=JSources.columns)

	DF5=RSDF.append(JSDF)
	print 'DataFrame combining record sources and join sources: '
	print DF5
	print 'JSON DF5: '+DF5.to_json(orient='records')
	# return JsonResponse(DF5.to_json(orient='records', force_ascii=False),safe=False)
	return HttpResponse(DF5.to_json(orient='records'),content_type='application/json')

def GetStandards(request):
	Model=request.GET['Model']
	result=graph.cypher.execute('match (a:Standard {Name:"'+Model+'"}) return a.Name,a.Version')
	resultlist=[]
	for x in result:
		resultlist.append({'Name':x[0],'Version':x[1]})
	return JsonResponse(resultlist,safe=False)

def GetStudies(request):
	Model=request.GET['Model']
	result=graph.cypher.execute('match (a:Study)-[:BasedOn]->(b:Standard {Name:"'+Model+'"}) return a.Name')
	resultlist=[]
	for x in result:
		resultlist.append({'Name':x[0]})
	return JsonResponse(resultlist,safe=False)

def GetPotentialWhereConditions(request):
	# This function builds a set of potential where clauses based on all parameters defined and dtypes defined for those parameters.
	# Those that have already been defined in the study for a given variable are identified (_MERGE variable in FINAL dataframe)
	Study=request.GET['Study']
	DSName=request.GET['DSName']
	VarName=request.GET['VarName']

	print 'VARNAME: '+VarName
	# This first query captures all potential conditions for parameters that have DTYPEs
	dtparms=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef {Name:"'+DSName+'"})--(:ItemDef {Name:"DTYPE"})--(:ValueListDef)\
		--(a:ItemDef)--(:WhereClauseDef)--(:RangeCheck)--(b:CheckValue) match (a)--(:CodeList)--(c:CodeListItem) return b.Value as PARAMCD,c.CodedValue as DTYPE')
	# Turn it into a dataframe
	dtparmsdf=pd.DataFrame(dtparms.records,columns=dtparms.columns)
	# For each of these parameters, we need to add the possibility that DTYPE is null too.  
	# To do that, we'll create a dataframe with each unique parameter from DTPARMSDF, but with DTYPE=NULL.
	# We'll then append that to DTPARMSDF.
	dtparmsuniquelist=[]
	for x in dtparmsdf.PARAMCD.unique():
		dtparmsuniquelist.append({'PARAMCD':x,'DTYPE':'NULL'})

	dtparmsuniquedf=pd.DataFrame(dtparmsuniquelist)
	dtparmswithnulldf=dtparmsdf.append(dtparmsuniquedf)
	print 'DTPARMSWITHNULLDF AFTER APPEND: '
	print dtparmswithnulldf

	# Now get all parameters
	allparms=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef {Name:"'+DSName+'"})--(:ItemDef {Name:"PARAMCD"})\
		--(:CodeList)--(a:CodeListItem) return a.CodedValue as PARAMCD')
	# Turn it into a dataframe
	allparmsdf=pd.DataFrame(allparms.records,columns=allparms.columns)
	# Now merge the two data frames together, giving us an idea of which parameters have DTYPEs (_MERGE=both) and which don't (_MERGE=left_only)
	apmrgdf=pd.merge(allparmsdf,dtparmswithnulldf,how='left',indicator=True,on='PARAMCD')
	# Since we'll have to merge again, change the name _MERGE to MERGE1
	newcols=[]
	for x in apmrgdf.columns:
		if x == '_merge':
			newcols.append('merge1')
		else:
			newcols.append(x)

	apmrgdf.columns=newcols

	print "APMRGDF: "
	print apmrgdf

	# Now see which PARAMCD/DTYPE combintations exist in where clauses (without regard to the operator) for this variable 
	WC = graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef {Name:"'+DSName+'"})--(:ItemDef {Name:"'+VarName+'"})--(:ValueListDef)\
		--(:ItemDef)--(w:WhereClauseDef) match (w)--(rp:RangeCheck)--(:ItemDef {Name:"PARAMCD"}) match (rp)--(cp:CheckValue) optional match (w)\
		--(rd:RangeCheck)--(:ItemDef {Name:"DTYPE"}) with w,rd,cp optional match (rd)--(cd:CheckValue) return w.OID as wcOID,cp.Value as PARAMCD, cd.Value as DTYPE')

	WCdf = pd.DataFrame(WC.records,columns=WC.columns)

	print "WCDF: "
	print WCdf
	# Now merge with the above data frame.  Matches (_merge=both) represent conditions that already exist for this variable
	final=pd.merge(apmrgdf,WCdf,how='left',indicator=True,on=['PARAMCD','DTYPE'])

	print "FINAL: "
	print final

	return JsonResponse(final.to_json(orient='records'),safe=False)

def GetStudyVarConditions(request):
	# This function retrieves all where clauses that have been defined by the study for a given variable
	Study=request.GET['Study']
	DSName=request.GET['DSName']
	VarName=request.GET['VarName']

	results=graph.cypher.execute('match (a:Study {Name:"'+Study+'"})--(:ItemGroupDef {Name:"'+DSName+'"})--(:ItemDef {Name:"'+VarName+'"})--(:ValueListDef)--(:ItemDef)\
		--(wc:WhereClauseDef)--(rc:RangeCheck) with wc,rc match (rc)-[:CheckRef]->(cv:CheckValue) with wc,rc,collect(cv.Value) as values \
		match (rc)-[:Range2Item]->(ID:ItemDef) with wc,rc,values,ID  return wc.OID as wcoid,collect(ID.Name+"|"+rc.Operator) as coll1,\
		collect(values) as coll2')

	resultlist=[]
	for x in results:
		dic={}
		dic['WCOID']=x['wcoid']
		string=''
		for y1,y2 in enumerate(x['coll1']):
			if y1 != 0:
				string=string+' AND '
			cond1=y2.split('|')
			string=string+cond1[0]+' '+cond1[1]+' '
			for z1,z2 in enumerate(x['coll2'][y1]):
				if z1 != 0:
					string=string+', '
				string=string+z2
		dic['Condition']=string 
		resultlist.append(dic)

	return JsonResponse(resultlist,safe=False)


def GetParmList(request):
	Study=request.GET['Study']
	DSName=request.GET['DSName']

	allparms=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef {Name:"'+DSName+'"})--(:ItemDef {Name:"PARAMCD"})\
		--(:CodeList)--(a:CodeListItem) return a.CodedValue as PARAMCD')
	allparmsdf=pd.DataFrame(allparms.records,columns=allparms.columns)
	return JsonResponse(allparmsdf.to_json(orient='records'),safe=False)

def GetDerivedRows(request):
	Study=request.GET['Study']
	DSName=request.GET['DSName']
	VarName=request.GET['VarName']

	# Get parameter/dtype combinations
	allresults1=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef {Name:"'+DSName+'"})--(:ItemDef {Name:"DTYPE"})--(:ValueListDef)\
		--(id:ItemDef)--(:WhereClauseDef)--(r:RangeCheck)--(c:CheckValue) match (id)--(:CodeList)--(cli:CodeListItem) return c.Value as paramcd,cli.CodedValue as dtype')
	allresults1df=pd.DataFrame(allresults1.records,columns=allresults1.columns)

	# Get parameters with at least two sources
	allresults2=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef {Name:"'+DSName+'"})--(:ItemDef {Name:"PARAMCD"})--(:CodeList)\
		--(cli:CodeListItem)-[ps:ParmSource]->(src:Source) with cli,count(*) as count where count > 1 return cli.CodedValue as paramcd')

	allresults2df=pd.DataFrame(allresults2.records,columns=allresults2.columns)
	allresults1df=allresults1df.append(allresults2df)

	# Now determine how any of the above methods were defined for this study
	studyresults=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef {Name:"'+DSName+'"})--(:ItemDef {Name:"'+VarName+'"})-[dmr:DerivedMethodRef]->\
		(dmd:DerivedMethodDef) return dmr.DTYPE as dtype,dmr.PARAMCD as paramcd,dmd.Type as type,dmd.Description as desc')

	studyresultsdf=pd.DataFrame(studyresults.records,columns=studyresults.columns)

	mergeresultsdf=pd.merge(allresults1df,studyresultsdf,how='left',indicator=True,on=['paramcd','dtype'])

	return JsonResponse(mergeresultsdf.to_json(orient='records'),safe=False)

def GetStandardDS(request):
	# Get dataset metadata properties for a given standard data set
	DSName=request.GET['DSName']
	StandardName=request.GET['StandardName']
	StandardVersion=request.GET['StandardVersion']

	stmt='match (a:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})--(igd:ItemGroupDef {Name:"'+DSName+'"})-[:BasedOn]->(igdm:ItemGroupDef) \
		return igd.Name as Name,igd.Label as Label,igd.Structure as Structure,igd.Repeating as Repeating,igd.Reference as Reference,igd.Purpose as Purpose,igdm.Name as Class'

	print 'GETSTANDARDDS STMT: '+stmt 
	RL=graph.cypher.execute(stmt)

	df=pd.DataFrame(RL.records,columns=RL.columns)
	print 'GETSTANDARDDS DF: '
	print df
	return HttpResponse(df.to_json(orient='records'),content_type='application/json')


def GetStudyDatasets(request):
	Study=request.GET['Study']
	print "STUDY: "+Study 
	DSRL=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(igd:ItemGroupDef) return igd.Name as Dataset')
	DSDF=pd.DataFrame(DSRL.records,columns=DSRL.columns)
	print 'GETSTUDYDATASETS DSDF: '
	print DSDF
	return HttpResponse(DSDF.to_json(orient='records'),content_type='application/json')

def GetStudyDSMD(request):
	# Get data set metadata for all data sets.  Includes CLASS and record sources.  One record returned, with records sources separated by commas
	Study=request.GET['Study']
	stmt='match (:Study {Name:"'+Study+'"})--(igd:ItemGroupDef)-[:RecordSource]->(igdr:ItemGroupDef) optional match (igd)-[:BasedOn]->(igdm:ItemGroupDef)<-[:ItemGroupRef]-(:Model) \
		optional match (igd)-[:BasedOn]->(:ItemGroupDef)-[:BasedOn]->(igds:ItemGroupDef)<-[:ItemGroupRef]-(:Model) \
		with igd.Name as Dataset,igd.Label as Label,igd.Structure as Structure,igd.Repeating as Repeating,igd.IsReferenceData as Reference,igdm.Name as ClassFromModel,igds.Name as ClassFromStd, \
		reduce(acc=head(collect(igdr.Name)), x in tail(collect(igdr.Name))|acc+", "+x) as RecordSources return Dataset,Label,Structure,Repeating,Reference,\
		case when ClassFromModel is null then ClassFromStd else ClassFromModel end as Class,RecordSources'

	DSRL=graph.cypher.execute(stmt)
	DSDF=pd.DataFrame(DSRL.records,columns=DSRL.columns)

	return HttpResponse(DSDF.to_json(orient='records'),content_type='application/json')

def GetStudyDS(request):
	# Get data set metadata for a specific data set. Includes CLASS and a boolean that indicates whether it comes from a standard data set or not, but no record sources
	Study=request.GET['Study']
	DSName=request.GET['DSName']

	stmt='match (:Study {Name:"'+Study+'"})--(igd:ItemGroupDef {Name:"'+DSName+'"}) optional match (igd)-[:BasedOn]->(igdm:ItemGroupDef)<-[:ItemGroupRef]-(:Model) \
		optional match (igd)-[:BasedOn]->(:ItemGroupDef)-[:BasedOn]->(igds:ItemGroupDef)<-[:ItemGroupRef]-(:Model) \
		with igd,igdm.Name as ClassFromModel,igds.Name as ClassFromStd \
		return igd.Name as Dataset,igd.Label as Label,igd.Structure as Structure,igd.Repeating as Repeating,igd.IsReferenceData as Reference,\
		case when ClassFromModel is null then ClassFromStd else ClassFromModel end as Class, \
		exists((igd)-[:BasedOn]->(:ItemGroupDef)<-[:ItemGroupRef]-(:Standard)) as StandardTF'

	RL=graph.cypher.execute(stmt)

	df=pd.DataFrame(RL.records,columns=RL.columns)
	return HttpResponse(df.to_json(orient='records'),content_type='application/json')

def GetRecordSources(request):
	# Get record sources for all data sets - one record per record source
	Study=request.GET['Study']
	DSName=request.GET['DSName']
	# For now, SDTM is differentiated from ADaM by assuming that any data set connected through BasedOn to another must be ADaM
	RSRL=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(igd:ItemGroupDef)-[rs:RecordSource]->(igdr:ItemGroupDef) \
		return distinct igdr.Name as dataset,rs.Subset as subset,case when exists((igdr)-[:BasedOn]->(:ItemGroupDef)) then "ADAM" else "SDTM" end as model, \
		"'+DSName+'" in collect(igd.Name) as state')

	RSDF=pd.DataFrame(RSRL.records,columns=RSRL.columns)

	print 'GetRecordSources DF: '
	print RSDF

	return HttpResponse(RSDF.to_json(orient='records'),content_type='application/json')


def GetJoinSources(request):
	Study=request.GET['Study']
	DSName=request.GET['DSName']
	JSRL=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef)--(:ItemDef)-[js:JoinSource]->(igd:ItemGroupDef) \
		return distinct igd.Name as Dataset,js.Subset as Subset,js.Join as JCondition,case when exists((igd)-[:BasedOn]->(:ItemGrouDef)) then "ADAM" else "SDTM" end as Model')
	JSDF=pd.DataFrame(JSRL.records,columns=JSRL.columns)
	return JsonResponse(JSDF.to_json(orient='records'),safe=False)

def GetStandardVarswoStudy(request):
	Study=request.GET['Study']
	DSName=request.GET['DSName']
	Class=request.GET['Class']
	IGDSource=request.GET['IGDSource']
	StandardName=request.GET['StandardName']
	StandardVersion=request.GET['StandardVersion']
	VarGroup=request.GET['VarGroup']

	# Get Standard variables for the current variable group
	if IGDSource == 'Standard':
		if VarGroup:
			AllStandardVars=QStandardVars(StandardName,StandardVersion,DSName,filter='c.Group="'+VarGroup+'"')[1]
		else:
			AllStandardVars=QStandardVars(StandardName,StandardVersion,DSName,filter='')[1]

	elif IGDSource == 'Model':
		# First get model information
		ModelInfoRL=graph.cypher.execute('match (:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})-[:BasedOn]->(m:Model) return m.name as name,m.version as version')
		if VarGroup:
			AllStandardVars=QModelVars(ModelInfoRL[0][0],ModelInfoRL[0][1],Class,filter='c.Group="'+VarGroup+'"')[1]
		else:
			AllStandardVars=QModelVars(ModelInfoRL[0][0],ModelInfoRL[0][1],Class,filter='')[1]

	# Get Study variables
	StudyVars=QStudyDSVarList(Study,DSName)[1]

	# Now merge the dataframes
	mrg=pd.merge(AllStandardVars,StudyVars,how='left',on='Name',indicator=True)

	# Keep only the variables that are not study variables
	mrg2=mrg[mrg['_merge'] == 'left_only']

	if not VarGroup:
		print 'MRG2: '
		print mrg2 

	return JsonResponse(mrg2.to_json(orient='records'),safe=False)

def GetStandardDSwoStudy(request):
	Study=request.GET['Study']
	StandardName=request.GET['StandardName']
	StandardVersion=request.GET['StandardVersion']

	StandardDS=graph.cypher.execute('match (:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})--(igd:ItemGroupDef) return igd.Name as Name')
	StudyDS=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(igd:ItemGroupDef) return igd.Name as Name')

	StandardDSDF=pd.DataFrame(StandardDS.records,columns=StandardDS.columns)
	StudyDSDF=pd.DataFrame(StudyDS.records,columns=StudyDS.columns)

	mrg=pd.merge(StandardDSDF,StudyDSDF,how='left',on='Name',indicator=True)
	mrg2=mrg[mrg['_merge'] == 'left_only']
	return HttpResponse(mrg2.to_json(orient='records'),content_type='application/json')

def GetAllVarGroups(request):
	DSName=request.GET['DSName']
	Class=request.GET['Class']
	IGDSource=request.GET['IGDSource']
	StandardName=request.GET['StandardName']
	StandardVersion=request.GET['StandardVersion']

	if IGDSource == 'Standard':
		GroupVars=QStandardVars(StandardName,StandardVersion,DSName,filter='')[1]
		GroupVars2=pd.DataFrame(GroupVars.VarGroup.unique())
		GroupVars2.columns=['VarGroup']

	elif IGDSource == 'Model':
		ModelInfoRL=graph.cypher.execute('match (:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})-[:BasedOn]->(m:Model) return m.name as name,m.version as version')
		GroupVars=QModelVars(ModelInfoRL[0][0],ModelInfoRL[0][1],Class,filter='')[1]
		GroupVars2=pd.DataFrame(GroupVars.VarGroup.unique())
		GroupVars2.columns=['VarGroup']
	print 'GETALLVARGROUPS: '
	print GroupVars2

	return HttpResponse(GroupVars2.to_json(orient='records'))

def GetStandardVar(request):
	IGDName=request.GET['IGDName']
	VarName=request.GET['VarName']
	StandardName=request.GET['StandardName']
	StandardVersion=request.GET['StandardVersion']
	StandardType=request.GET['StandardType']

	if StandardType == 'Standard':
	# This function returns a set of properties for a chosen variable in a chosen data set
		stmt='match (a:Standard {Name:"'+StandardName+'",Version:"'+StandardVersion+'"})'

	elif StandardType == 'Model':
		stmt='match (a:Model {name:"'+StandardName+'",version:"'+StandardVersion+'"})'

	stmt=stmt+'-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+IGDName+'"})-[r:ItemRef]->(c:ItemDef {Name:"'+VarName+'"}) \
			optional match (c)-[:CodeListRef]->(cl:CodeList)--(cli:CodeListItem) return c.Name as Name,c.Label as Label,c.SASType as SASType,c.DataType as DataType,c.Origin as Origin,c.MaxLength as SASLength, \
			r.OrderNumber as OrderNumber,r.Mandatory as Mandatory,cl.Name as CodeListName,cl.Extensible as Extensible,cl.AliasName as CLCode,cli.Decode is not null as DecodeYN limit 1 '

	print 'GetStandardVar STMT: '+stmt

	RL=graph.cypher.execute(stmt)

	df=pd.DataFrame(RL.records,columns=RL.columns)
	return HttpResponse(df.to_json(orient='records'))

def GetStudyVar(request):
	DSName=request.GET['DSName']
	VarName=request.GET['VarName']
	Study=request.GET['Study']
	VLMOID=request.GET['VLMOID']

	# This function returns ItemDef metadata for a single variable.  If VLMOID is specified, it will return VLM for the condition specified by this OID.
	# When no VLMOID is given, then variable-level metadata is returned, as well as a boolean (VLM) that indicates whether or not this variable is defined with VLM.
	stmt='match (a:Study {Name:"'+Study+'"})-[:ItemGroupRef]->(b:ItemGroupDef {Name:"'+DSName+'"})-[r:ItemRef]-> '
	if VLMOID:
		stmt=stmt+'(c0:ItemDef {Name:"'+VarName+'"})-[:ValueListRef]->(:ValueListDef)-[r1:ItemRef {WCRefOID:'+str(VLMOID)+'}]->(c:ItemDef) \
			return distinct c.Name as Name,c.Label as Label,c.DataType as DataType,c.Origin as Origin,c.Length as Length, \
			r1.Mandatory as Mandatory,r1.WCRefOID as VLMOID'
	else:
		stmt=stmt+'(c:ItemDef {Name:"'+VarName+'"}) optional match (c)-[:CodeListRef]->(clStudy:CodeList)--(cliStudy:CodeListItem) \
		optional match (c)-[:BasedOn]->(c2:ItemDef)-[:CodeListRef]->(clStandard:CodeList)--(cliStandard:CodeListItem) \
		return distinct c.Name as Name,c.Label as Label,c.SASType as SASType,c.DataType as DataType,c.Origin as Origin,c.SASLength as SASLength, \
		r.Order as OrderNumber,r.Mandatory as Mandatory,exists((c)--(:ValueListDef)) as VLM,exists((c)-[:BasedOn]->(x) as CustomYN,clStandard.Name as CodeListName,\
		c1Study.Extensible as Extensible,c1Study.AliasName as CLCode,c1Study.Decode is not null as DecodeYN,clStudy.Name as CodeListNameStudy limit 1 '

	print 'GETSTUDYVAR STMT: '+stmt

	RL=graph.cypher.execute(stmt)
	df=pd.DataFrame(RL.records,columns=RL.columns)

	return HttpResponse(df.to_json(orient='records'))


def GetADaMStudyVars(request):
	# All variable metadata for a chosen data set in the study - does not include sources, methods, CT, or ItemRef properties
	Study=request.GET['Study']
	DSName=request.GET['DSName']
	RL=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(:ItemGroupDef {Name:"'+DSName+'"})-[ir:ItemRef]-(id:ItemDef) return id.Name as VarName,id.Label as Label,id.SASType as SASType,id.SASLength as SASLength,id.DataType as DataType,\
		id.Origin as Origin order by ir.OrderNumber')
	df=pd.DataFrame(RL.records,columns=RL.columns)
	return HttpResponse(df.to_json(orient='records'),content_type='application/json')

def GetADaMStudyDS(request):
	# List of all study data sets for the study
	Study=request.GET['Study']
	RL=graph.cypher.execute('match (:Study {Name:"'+Study+'"})--(igd:ItemGroupDef) return igd.Name as DSName')
	df=pd.DataFrame(RL.records,columns=RL.columns)
	return HttpResponse(df.to_json(orient='records'))












def modds(request):
	for k,v in request.GET.iteritems():
		if k == 'add_BDS':
			DSClass="BASIC DATA STRUCTURE"
			modelitemdefs = graph.cypher.execute('match (:Model {name:"'+parentmodel+'"})-[:ItemGroupRef]->(:ItemGroupDef {Name:"'+DSClass+'"})- \
				[r:ItemRef]->(b:ItemDef) return b.Name as Name,b.OID as OID order by r.OrderNumber')
			return render(request,'StandardDeveloper1/vars1.html',{'Itemdefs':modelitemdefs,'Name':'','Label':'','Class':DSClass,'Structure':'One record per parameter per time point per subject',\
				'Reference':'No','Repeating':'Yes'})

		elif k == 'add_Occur':
			DSClass="OCCURRENCE DATA STRUCTURE"
			modelitemdefs = graph.cypher.execute('match (:Model {name:"'+parentmodel+'"})-[:ItemGroupRef]->(:ItemGroupDef {Name:"'+DSClass+'"})- \
				[r:ItemRef]->(b:ItemDef) return b.Name as Name,b.OID as OID order by r.OrderNumber')
			return render(request,'StandardDeveloper1/vars1.html',{'Itemdefs':modelitemdefs,'Name':'','Label':'','Class':DSClass,'Structure':'One record per occurrence per subject',\
				'Reference':'No','Repeating':'Yes'})

		elif k[0:4] == 'add_':
			# Determine which data set is chosen
			dataset=k[4:]
			# Determine if the data set is attached directly to a standard.  If so, then its properties are read-only
			readonlyRL=graph.cypher.execute('match (:Standard {name:"'+parentstandard+'"})-[:ItemGroupRef]->(:ItemGroupDef {Name:"'+dataset+'"})')
			if readonlyRL:
				readonly='Y'
			else:
				readonly='N'
			# Now get the data set properties from the  standard
			dsprops = graph.cypher.execute('match (:Standard {name:"'+parentstandard+'"})-[:ItemGroupRef]->(a:ItemGroupDef {Name:"'+dataset+'"})-[:BasedOn]->(c:ItemGroupDef) \
				return a.Name as IGDName,a.Label as Label,a.Repeating as Repeating,a.Reference as Reference,a.Structure as Structure,a.Purpose as Purpose,c.Name as Class')
			return render(request,'StandardDeveloper1/mdds.html',{'Study':study,'mdds':dsprops,'readonly':readonly,'Action':'Add'})

		elif k[0:5] == 'edit_':
			dataset=k[5:]
			# Determine if the data set is attached directly to a standard.  If so, then its properties are read-only
			readonlyRL=graph.cypher.execute('match (:Standard {name:"'+parentstandard+'"})-[:ItemGroupRef]->(:ItemGroupDef {Name:"'+dataset+'"})')
			if readonlyRL:
				readonly='Y'
			else:
				readonly='N'
			# Now get the data set properties from the study
			dsprops = graph.cypher.execute('match (:Study {Name:"'+study+'"})-[:ItemGroupRef]->(a:ItemGroupDef {Name:"'+dataset+'"})-\
				return a.Name as IGDName,a.Label as Label,a.Repeating as Repeating,a.Reference as Reference,a.Structure as Structure,a.Purpose as Purpose,a.DSClass as Class')
			mainsource = graph.cypher.execute('match (:Study {Name:"'+study+'"})-[:ItemGroupRef]->(a:ItemGroupDef {Name:"'+dataset+'"})-->\
				(b:RecordSourceContainer {Type:"Original"})-->(c:RecordSource {Type:"Main"}) return c.Name as Name,c.Description as Description')
			mergesources = graph.cypher.execute('match (:Study {Name:"'+study+'"})-[:ItemGroupRef]->(a:ItemGroupDef {Name:"'+dataset+'"})-->\
				(b:RecordSourceContainer {Type:"Original"})-->(c:RecordSource {Type:"Merge"}) return c.Name as Name')
			return render(request,'StandardDeveloper1/mdds.html',{'Study':study,'mdds':dsprops,'MainSource':mainsource,'MergeSources':mergesources,'readonly':readonly,'Action':'Edit'})

		elif k[0:7] == 'remove_':
			pass

		elif k == 'Export':
			# Get standard data sets
			getStandardDS()
			# Get standard variables
			standardvarsRL = graph.cypher.execute('match (a:Standard {name:"'+standard+'"})-[:ItemGroupRef]->(b:ItemGroupDef)-[r:ItemRef]->(c:ItemDef) \
				return r.OrderNumber as OrderNumber,b.Name as DSName,c.Name as VarName,c.Label,c.DataType,c.MaxLength as Length,r.Mandatory as \
				Mandatory,c.CodeList as CodeList,c.Origin as Origin,r.MethodOID as MethodOID,c.Predecessor as Predecessor order by b.Name,r.OrderNumber')
			# Create Excel workbook
			wb = openpyxl.Workbook()
			sheet = wb.active
			sheet.title = "Datasets"
			sheet['A1'] = 'Dataset'
			sheet['B1'] = 'Description'
			sheet['C1'] = 'Class'
			sheet['D1'] = 'Structure'
			sheet['E1'] = 'Purpose'
			sheet['F1'] = 'Key Variables'
			sheet['G1'] = 'Repeating'
			sheet['H1'] = 'Reference Data'
			sheet['I1'] = 'Comment'
			row = 1
			for x1 in standarddatasetsRL:
				row = row+1
				sheet.cell(row=row,column=1).value=x1[0]
				sheet.cell(row=row,column=2).value=x1[1]
				sheet.cell(row=row,column=3).value=x1[5]
				sheet.cell(row=row,column=4).value=x1[3]
				sheet.cell(row=row,column=5).value="Analysis"
				sheet.cell(row=row,column=7).value=x1[2]
				sheet.cell(row=row,column=8).value=x1[4]

			vars = wb.create_sheet('Variables')
			vars['A1'] = 'Order'
			vars['B1'] = 'Dataset'
			vars['C1'] = 'Variable'
			vars['D1'] = 'Label'
			vars['E1'] = 'Data Type'
			vars['F1'] = 'Length'
			vars['G1'] = 'Significant Digits'
			vars['H1'] = 'Format'
			vars['I1'] = 'Mandatory'
			vars['J1'] = 'Codelist'
			vars['K1'] = 'Origin'
			vars['L1'] = 'Pages'
			vars['M1'] = 'Method'
			vars['N1'] = 'Predecessor'
			vars['O1'] = 'Role'
			vars['P1'] = 'Comment'
			row=1
			lastds=''
			for x1 in standardvarsRL:
				if x1["DSName"] == lastds:
					order=order+1
				else:
					order=1
				row=row+1
				lastds=x1["DSName"]
				vars.cell(row=row,column=1).value=order
				vars.cell(row=row,column=2).value=x1["DSName"]
				vars.cell(row=row,column=3).value=x1["VarName"]
				vars.cell(row=row,column=4).value=x1["c.Label"]
				vars.cell(row=row,column=5).value=x1["c.DataType"]
				vars.cell(row=row,column=6).value=x1["Length"]
				vars.cell(row=row,column=9).value=x1["Mandatory"]
				vars.cell(row=row,column=10).value=x1["Codelist"]
				vars.cell(row=row,column=11).value=x1["Origin"]
				vars.cell(row=row,column=13).value=x1["MethodOID"]
				vars.cell(row=row,column=14).value=x1["Predecessor"]

			filename = standard+'-'+datetime.now().isoformat()+'.xlsx'
			wb.save(filename)

			return render(request,'StandardDeveloper1/datasets.html',{'Standard':standard,'AddDatasets':parentdatasetsLIST, \
				'StandardDatasets':standarddatasetsRL,'Message':'Standard '+standard+' has been saved to '+filename})


